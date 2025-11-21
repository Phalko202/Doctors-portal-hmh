"""
Generate a blank Excel template for doctor schedules
Use this as a starting template to fill in your schedule data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Doctor Schedule"

# Define styles
title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
title_alignment = Alignment(horizontal='center', vertical='center')

section_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
section_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
section_alignment = Alignment(horizontal='center', vertical='center')

header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Get current date and 14 days ahead
today = datetime.now()
end_date = today + timedelta(days=13)
date_range = f"{today.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"

# Column headers
columns = [
    'Date',
    'Doctor',
    'Start Time',
    'Room',
    'Total Patients',
    'Before Break OPD Timing',
    'Before Break OPD Patients',
    'Breaks',
    'After Break OPD Timing',
    'After Break OPD Patients',
    'Status'
]

# Set column widths
col_widths = [12, 35, 12, 10, 15, 20, 20, 15, 20, 20, 12]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Row 1: Title (NO MERGE across frozen column). Fill A..last and center text using centerContinuous.
title_text = f'Doctor Schedule Template ({date_range})'
last_col_letter = get_column_letter(len(columns))
for col_idx in range(1, len(columns) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = title_fill
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
# Put text in column B and center across using centerContinuous so it appears centered even with freeze panes
ws['B1'] = title_text
ws['B1'].font = title_font
ws['B1'].fill = title_fill
ws['B1'].alignment = Alignment(horizontal='centerContinuous', vertical='center')
ws.row_dimensions[1].height = 30

def ordinal(n:int)->str:
    return f"{n}{'th' if 11<=n%100<=13 else {1:'st',2:'nd',3:'rd'}.get(n%10,'th')}" 

def add_date_banner(row_num, d_obj):
    label = f"{ordinal(d_obj.day)} {d_obj.strftime('%A %Y %d/%m/%Y')}"
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(columns))
    cell = ws.cell(row=row_num, column=1)
    cell.value = label
    cell.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='00B388', end_color='00B388', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 22
    return row_num + 1

current_row = 2

# List of specialties
specialties = [
    'Internal Medicine',
    'Paediatrics',
    'Orthopaedics',
    'Obstetrics & Gynaecology',
    'General Surgery',
    'ENT',
    'Ophthalmology',
    'Dental'
]

# Add each specialty section with column headers and 5 blank rows
def add_specialty_section(row_num, specialty_name, blank_rows=10):
    """Insert a specialty block.
    Improvements:
      - No merging across frozen column → prevents horizontal drift.
      - Use centerContinuous alignment from column B to last for a clean centered header.
      - Uniform green band background across the full width A..last.
      - Reduced vertical padding & removed extra spacer explosion.
    """
    # Specialty header row styling across all columns
    for col_idx in range(1, len(columns) + 1):
        hcell = ws.cell(row=row_num, column=col_idx)
        hcell.fill = section_fill
        # Put the specialty label only once (col B) but visually center across row
        if col_idx == 2:
            hcell.value = specialty_name
            hcell.font = section_font
            hcell.alignment = Alignment(horizontal='centerContinuous', vertical='center')
        else:
            hcell.value = ''
            hcell.font = section_font
            hcell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 24
    row_num += 1

    # Column headers row
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    ws.row_dimensions[row_num].height = 30
    row_num += 1

    # Blank data entry rows
    for _ in range(blank_rows):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_num, column=col_idx, value='')
            cell.border = border
            # Left-align Doctor column for readability; center others
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row_num].height = 20
        row_num += 1

    # Minimal single blank spacer row (visual separation only)
    for col_idx in range(1, len(columns) + 1):
        spacer = ws.cell(row=row_num, column=col_idx, value='')
        spacer.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 6
    row_num += 1
    return row_num

# Build grouped-by-day skeleton for the next 14 days
cur = today
while cur <= end_date:
    current_row = add_date_banner(current_row, cur)
    for specialty in specialties:
        current_row = add_specialty_section(current_row, specialty, blank_rows=8)
    cur += timedelta(days=1)

# Freeze panes: freeze only the title row + Date column (A) so date banners scroll beneath
ws.freeze_panes = 'B2'

# Add instructions in a separate sheet
instructions = wb.create_sheet("Instructions")
instructions['A1'] = "HOW TO USE THIS TEMPLATE"
instructions['A1'].font = Font(size=16, bold=True, color='1F4788')

instructions_text = [
    "",
    "📋 COLUMN DEFINITIONS:",
    "",
    "• Date: Format as DD/MM/YYYY (e.g., 06/11/2025) or YYYY-MM-DD (e.g., 2025-11-06)",
    "• Doctor: Full doctor name with prefix (e.g., Dr. Aminath Munaza)",
    "• Start Time: 24-hour format (e.g., 8:00 or 08:00)",
    "• Room: Room number or code (e.g., R1, Room 1)",
    "• Total Patients: Total number of patients for the day",
    "• Before Break OPD Timing: Time range before break (e.g., 8:00-11:00)",
    "• Before Break OPD Patients: Number of patients before break",
    "• Breaks: Break time range (e.g., 11:00-14:00) or 'NO BREAK'",
    "• After Break OPD Timing: Time range after break (e.g., 14:00-17:00)",
    "• After Break OPD Patients: Number of patients after break",
    "• Status: ON DUTY, OFF, ON CALL, POST ON CALL",
    "",
    "✅ IMPORTANT TIPS:",
    "",
    "• Use separate columns for timing and patient counts (NOT combined)",
    "• Fill in dates for each row (or use date declaration rows)",
    "• Keep section headers as they are (don't modify specialty names)",
    "• Use consistent date format throughout the file",
    "• Leave Status empty if ON DUTY (it's the default)",
    "• Use 'OFF' status for off-duty doctors",
    "",
    "🔧 FIXES APPLIED:",
    "",
    "✓ Title row is centered and fixed",
    "✓ Section headers are centered in green",
    "✓ Frozen panes prevent header sliding",
    "✓ Proper column separation for timing and patients",
    "",
    "💾 SAVE & UPLOAD:",
    "",
    "1. Fill in the schedule data in the 'Doctor Schedule' sheet",
    "2. Save the file",
    "3. Upload to the Schedule Management System",
    "4. Check the preview statistics to verify all rows parsed correctly",
    "5. Click 'Apply Schedules' to update the system",
]

for i, text in enumerate(instructions_text, 2):
    instructions[f'A{i}'] = text
    if text.startswith('•'):
        instructions[f'A{i}'].font = Font(size=11)
    elif text.startswith('✅') or text.startswith('📋') or text.startswith('🔧') or text.startswith('💾'):
        instructions[f'A{i}'].font = Font(size=13, bold=True, color='1F4788')
    else:
        instructions[f'A{i}'].font = Font(size=11)

instructions.column_dimensions['A'].width = 100

# Save the workbook
output_file = 'Doctor_Schedule_BLANK_Template.xlsx'
wb.save(output_file)
print(f"✅ Blank Excel template created successfully: {output_file}")
print(f"\n📋 Template includes:")
print(f"   • All specialty sections with blank rows to fill in")
print(f"   • Proper formatting with centered headers")
print(f"   • Frozen panes for easy navigation")
print(f"   • Instructions sheet with detailed guide")
print(f"   • Date range: {date_range}")
print(f"\n💡 Fill in your schedule data and upload to the system!")
