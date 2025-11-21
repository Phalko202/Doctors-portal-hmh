"""
Generate a properly formatted Excel template for doctor schedules
This template fixes the sliding header and centering issues
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Doctor Schedule"

# Define styles
title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
title_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
title_alignment = Alignment(horizontal='center', vertical='center')

section_font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
section_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
section_alignment = Alignment(horizontal='center', vertical='center')

header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_alignment = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)

# Get current date and next day
today = datetime.now()
tomorrow = today + timedelta(days=1)
date_range = f"{today.strftime('%Y-%m-%d')} to {tomorrow.strftime('%Y-%m-%d')}"

# Column headers
columns = [
    'Date',
    'Doctor',
    'Start Time',
    'Room',
    'Total Patients',
    'Before Break OPD Timing',
    'Before Break OPD Patients',
    'Breaks',
    'After Break OPD Timing',
    'After Break OPD Patients',
    'Status'
]

# Set column widths
col_widths = [12, 30, 12, 10, 15, 20, 20, 15, 20, 20, 12]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Row 1: Title (NO MERGE to avoid frozen-column split). Fill A..last and center via centerContinuous.
title_text = f'Doctor Schedule Template ({date_range})'
for col_idx in range(1, len(columns) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = title_fill
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
ws['B1'] = title_text
ws['B1'].font = title_font
ws['B1'].fill = title_fill
ws['B1'].alignment = Alignment(horizontal='centerContinuous', vertical='center')
ws.row_dimensions[1].height = 30

def ordinal(n:int)->str:
    return f"{n}{'th' if 11<=n%100<=13 else {1:'st',2:'nd',3:'rd'}.get(n%10,'th')}" 

def add_date_banner(row_num, d_obj):
    label = f"{ordinal(d_obj.day)} {d_obj.strftime('%A %Y %d/%m/%Y')}"
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(columns))
    cell = ws.cell(row=row_num, column=1)
    cell.value = label
    cell.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='00B388', end_color='00B388', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 22
    return row_num + 1

current_row = 2

# Function to add a specialty section
def add_specialty_section(row_num, specialty_name, doctors_data):
    global current_row
    
    # Section header: DO NOT MERGE ACROSS FROZEN COLUMN A (prevents sliding glitch)
    # Fill column A cell to keep the green band continuous
    left_cell = ws.cell(row=row_num, column=1)
    left_cell.fill = section_fill
    # Merge from column B..last for the header text
    start_col, end_col = 2, len(columns)
    ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
    section_cell = ws.cell(row=row_num, column=start_col)
    section_cell.value = specialty_name
    section_cell.font = section_font
    section_cell.fill = section_fill
    section_cell.alignment = section_alignment
    ws.row_dimensions[row_num].height = 25
    row_num += 1
    
    # Column headers
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    ws.row_dimensions[row_num].height = 35
    row_num += 1
    
    # Data rows
    for doctor_data in doctors_data:
        for col_idx, value in enumerate(doctor_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = border
        ws.row_dimensions[row_num].height = 20
        row_num += 1
    
    # Add one blank row between sections
    row_num += 1
    return row_num

# Sample data for Internal Medicine
internal_medicine_data = [
    [today.strftime('%d/%m/%Y'), 'Dr. Aminath Munaza', '8:00', 'R1', 20, '8:00-11:00', 10, '11:00-14:00', '14:00-17:00', 10, 'ON DUTY'],
    [today.strftime('%d/%m/%Y'), 'Dr. Cho Thway Mon', '9:00', 'R2', 15, '9:00-12:00', 8, '12:00-14:00', '14:00-16:00', 7, 'ON DUTY'],
    [today.strftime('%d/%m/%Y'), 'Dr. Ei Shwe Yee Myint', '', '', '', '', '', '', '', '', 'OFF'],
    [tomorrow.strftime('%d/%m/%Y'), 'Dr. Aminath Munaza', '8:00', 'R1', 18, '8:00-11:00', 9, '11:00-14:00', '14:00-17:00', 9, 'ON DUTY'],
]

# Sample data for Paediatrics
paediatrics_data = [
    [today.strftime('%d/%m/%Y'), 'Dr. Alaa Adel Ahmed Sayed Ahmed Sobeih Elamawey', '8:00', 'P1', 25, '8:00-11:00', 12, '11:00-14:00', '14:00-17:00', 13, 'ON DUTY'],
    [today.strftime('%d/%m/%Y'), 'Dr. Alisha Joshi', '9:00', 'P2', 20, '9:00-12:00', 10, '12:00-14:00', '14:00-16:00', 10, 'ON DUTY'],
    [tomorrow.strftime('%d/%m/%Y'), 'Dr. Roshan Kumar Dikshit', '', '', '', '', '', '', '', '', 'OFF'],
]

# Sample data for Orthopaedics
orthopaedics_data = [
    [today.strftime('%d/%m/%Y'), 'Dr. Ahmed Azim Abdul Shakoor', '8:00', 'O1', 22, '8:00-11:30', 11, '11:30-14:00', '14:00-17:00', 11, 'ON DUTY'],
    [tomorrow.strftime('%d/%m/%Y'), 'Dr. Ahmed Azim Abdul Shakoor', '8:00', 'O1', 20, '8:00-11:30', 10, '11:30-14:00', '14:00-17:00', 10, 'ON DUTY'],
]

# Build grouped-by-day layout: date banner, then specialties for that date
specialties = [
    ('Internal Medicine', internal_medicine_data),
    ('Paediatrics', paediatrics_data),
    ('Orthopaedics', orthopaedics_data),
]
for d_obj in [today, tomorrow]:
    current_row = add_date_banner(current_row, d_obj)
    d_str = d_obj.strftime('%d/%m/%Y')
    for spec_name, rows in specialties:
        # Filter rows for this date
        spec_rows = [r for r in rows if (r and r[0] == d_str)]
        # If no sample rows for this date, include a couple of blank rows
        if not spec_rows:
            spec_rows = [[d_str, '', '', '', '', '', '', '', '', '', ''] for _ in range(3)]
        current_row = add_specialty_section(current_row, spec_name, spec_rows)

# Freeze panes: Keep only blue title and Date column frozen so date banners scroll with content
ws.freeze_panes = 'B2'

# Save the workbook
output_file = 'Doctor_Schedule_Template_Proper_Format.xlsx'
wb.save(output_file)
print(f"✅ Excel template created successfully: {output_file}")
print(f"\n📋 Template features:")
print(f"   • Proper column separation (timing and patient count)")
print(f"   • Centered section headers (Internal Medicine, etc.)")
print(f"   • Fixed title row (Doctor Schedule Template)")
print(f"   • Frozen panes for easy scrolling")
print(f"   • Sample data for {today.strftime('%d/%m/%Y')} and {tomorrow.strftime('%d/%m/%Y')}")
print(f"\n💡 The template is ready to upload to the schedule system!")
