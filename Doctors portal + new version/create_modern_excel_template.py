"""Modern flat Excel schedule template generator.

Design goals:
  - Clean single table (no repeated specialty header blocks) for fast editing.
  - Frozen Date column & title row for constant context.
  - White data cells, subtle borders, green header band.
  - Optional Status dropdown (OFF, ON/CALL, POST ON/CALL; blank = normal duty).
  - Auto filter enabled on header row for quick doctor/date filtering.
  - Generates rows for every doctor for each day in a 14‑day window (skips Fridays).

Output: Doctor_Schedule_MODERN_Template.xlsx
"""

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
import json, os

OUTPUT_FILE = 'Doctor_Schedule_MODERN_Template.xlsx'
DAYS_AHEAD = 13  # inclusive range: today .. today+13 = 14 days

# Columns (kept compatible with existing backend parser). Added Specialty + Shift.
COLUMNS = [
    'Date',
    'Specialty',
    'Doctor',
    'Shift',
    'Start Time',
    'Room',
    'Total Patients',
    'Before Break OPD Timing',
    'Before Break OPD Patients',
    'Breaks',
    'After Break OPD Timing',
    'After Break OPD Patients',
    'Status'
]

# Styling definitions
TITLE_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')  # deep green
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Calibri', size=11, color='000000')
ALT_ROW_FILL = PatternFill(start_color='F7FFF9', end_color='F7FFF9', fill_type='solid')  # very subtle green tint
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
THIN = Side(style='thin', color='D0D0D0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def load_doctors():
    """Load doctors list from data/doctors.json; fallback to example if unavailable."""
    path = os.path.join(os.getcwd(), 'data', 'doctors.json')
    try:
        with open(path, 'r', encoding='utf-8') as f:
            js = json.load(f)
        docs = js if isinstance(js, list) else js.get('doctors', [])
        if not docs:
            raise ValueError('empty doctors list')
        return sorted(docs, key=lambda d: d.get('name',''))
    except Exception:
        return [
            {'name': 'Dr. Example One', 'specialty': 'Internal Medicine'},
            {'name': 'Dr. Example Two', 'specialty': 'Paediatrics'},
        ]

def ordinal(n:int)->str:
    return f"{n}{'th' if 11<=n%100<=13 else {1:'st',2:'nd',3:'rd'}.get(n%10,'th')}" 

def build_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Schedule'

    today = datetime.now()
    end_date = today + timedelta(days=DAYS_AHEAD)
    date_range = f"{today.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"

    # Title row (no merge; centerContinuous for resilience with frozen col A)
    for col_idx in range(1, len(COLUMNS) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.fill = TITLE_FILL
        c.font = TITLE_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'] = f'Doctor Schedule (Modern Flat) {date_range}'
    ws['B1'].font = TITLE_FONT
    ws['B1'].alignment = Alignment(horizontal='centerContinuous', vertical='center')
    ws.row_dimensions[1].height = 30

    # Header row
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[2].height = 34

    # Column widths (balanced for readability)
    widths = [12, 20, 36, 12, 11, 10, 14, 19, 18, 14, 19, 18, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Build hidden ShiftKnowledge sheet
    know = wb.create_sheet('ShiftKnowledge')
    know_headers = ['Key','Specialty','Shift','Start Time','Before Timing','Before Patients','Breaks','After Timing','After Patients','Total Patients']
    for ci, h in enumerate(know_headers, start=1):
        know.cell(row=1, column=ci, value=h).font = Font(bold=True)
    # Load knowledge from data/shift_knowledge.json if present
    def _load_knowledge():
        path_json = os.path.join(os.getcwd(), 'data', 'shift_knowledge.json')
        try:
            with open(path_json, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {
                'Internal Medicine': {
                    'Shift 1': {'start': '08:00','before_timing':'08:00-12:00','before_patients':15,'breaks':'12:00-13:00','after_timing':'13:00-16:00','after_patients':15,'total_patients':30},
                    'Shift 2': {'start': '13:00','before_timing':'13:00-18:00','before_patients':15,'breaks':'18:00-19:00','after_timing':'19:00-21:00','after_patients':15,'total_patients':30},
                },
                'Orthopaedics': {
                    'Shift 1': {'start': '08:00','before_timing':'08:00-12:30','before_patients':12,'breaks':'12:30-13:30','after_timing':'13:30-16:30','after_patients':12,'total_patients':24},
                    'Shift 2': {'start': '13:00','before_timing':'13:00-17:00','before_patients':12,'breaks':'17:00-18:00','after_timing':'18:00-21:00','after_patients':12,'total_patients':24},
                },
                'Obstetrics & Gynaecology': {
                    'Shift 1': {'start': '08:00','before_timing':'08:00-12:00','before_patients':12,'breaks':'12:00-13:00','after_timing':'13:00-16:00','after_patients':12,'total_patients':24},
                    'Shift 2': {'start': '13:00','before_timing':'13:00-18:00','before_patients':12,'breaks':'18:00-19:00','after_timing':'19:00-21:00','after_patients':12,'total_patients':24},
                },
                'Paediatrics': {
                    'Shift 1': {'start': '08:00','before_timing':'08:00-12:00','before_patients':15,'breaks':'12:00-13:00','after_timing':'13:00-16:00','after_patients':15,'total_patients':30},
                    'Shift 2': {'start': '13:00','before_timing':'13:00-18:00','before_patients':15,'breaks':'18:00-19:00','after_timing':'19:00-21:00','after_patients':15,'total_patients':30},
                }
            }
    knowledge = _load_knowledge()
    r = 2
    for spec, shifts in knowledge.items():
        for sh_name, vals in shifts.items():
            key = f"{spec}|{sh_name}"
            row_vals = [
                key, spec, sh_name,
                vals.get('start',''),
                vals.get('before_timing',''),
                vals.get('before_patients',''),
                vals.get('breaks',''),
                vals.get('after_timing',''),
                vals.get('after_patients',''),
                vals.get('total_patients',''),
            ]
            for ci, v in enumerate(row_vals, start=1):
                know.cell(row=r, column=ci, value=v)
            r += 1
    know.sheet_state = 'hidden'

    # Helper for formula ranges
    last_know_row = r-1
    key_rng = f"ShiftKnowledge!$A$2:$A${last_know_row}"
    start_rng = f"ShiftKnowledge!$D$2:$D${last_know_row}"
    btim_rng = f"ShiftKnowledge!$E$2:$E${last_know_row}"
    bpat_rng = f"ShiftKnowledge!$F$2:$F${last_know_row}"
    brks_rng = f"ShiftKnowledge!$G$2:$G${last_know_row}"
    atim_rng = f"ShiftKnowledge!$H$2:$H${last_know_row}"
    apat_rng = f"ShiftKnowledge!$I$2:$I${last_know_row}"
    tpat_rng = f"ShiftKnowledge!$J$2:$J${last_know_row}"

    # Data rows: every doctor for each date (skip Fridays)
    doctors = load_doctors()
    row = 3
    cur = today
    while cur <= end_date:
        if cur.weekday() == 4:  # Friday skip
            cur += timedelta(days=1)
            continue
        date_str_display = cur.strftime('%d/%m/%Y')
        for d in doctors:
            # Provide Date & Doctor; leave other cells blank
            ws.cell(row=row, column=1, value=date_str_display)
            ws.cell(row=row, column=2, value=d.get('specialty',''))
            ws.cell(row=row, column=3, value=d.get('name',''))
            ws.cell(row=row, column=4, value='')  # Shift (validation below)
            # Pre-fill formulas that read from ShiftKnowledge unless Shift is Custom
            # Build key: =B{row}&"|"&D{row}
            key_expr = f"$B{row}&\"|\"&$D{row}"
            # E Start Time
            ws.cell(row=row, column=5, value=f"=IF($D{row}=\"Custom\",\"\",IFERROR(XLOOKUP({key_expr},{key_rng},{start_rng},\"\"),IFERROR(INDEX({start_rng},MATCH({key_expr},{key_rng},0)),\"\")))")
            # F Room empty
            ws.cell(row=row, column=6, value='')
            # G Total Patients = sum of before/after patients if present
            ws.cell(row=row, column=7, value=f"=IF(AND($I{row}=\"\",$L{row}=\"\"),\"\",IFERROR(VALUE($I{row}),0)+IFERROR(VALUE($L{row}),0))")
            # H Before Timing
            ws.cell(row=row, column=8, value=f"=IF($D{row}=\"Custom\",\"\",IFERROR(XLOOKUP({key_expr},{key_rng},{btim_rng},\"\"),IFERROR(INDEX({btim_rng},MATCH({key_expr},{key_rng},0)),\"\")))")
            # I Before Patients
            ws.cell(row=row, column=9, value=f"=IF($D{row}=\"Custom\",\"\",IFERROR(XLOOKUP({key_expr},{key_rng},{bpat_rng},\"\"),IFERROR(INDEX({bpat_rng},MATCH({key_expr},{key_rng},0)),\"\")))")
            # J Breaks
            ws.cell(row=row, column=10, value=f"=IF($D{row}=\"Custom\",\"\",IFERROR(XLOOKUP({key_expr},{key_rng},{brks_rng},\"\"),IFERROR(INDEX({brks_rng},MATCH({key_expr},{key_rng},0)),\"\")))")
            # K After Timing
            ws.cell(row=row, column=11, value=f"=IF($D{row}=\"Custom\",\"\",IFERROR(XLOOKUP({key_expr},{key_rng},{atim_rng},\"\"),IFERROR(INDEX({atim_rng},MATCH({key_expr},{key_rng},0)),\"\")))")
            # L After Patients
            ws.cell(row=row, column=12, value=f"=IF($D{row}=\"Custom\",\"\",IFERROR(XLOOKUP({key_expr},{key_rng},{apat_rng},\"\"),IFERROR(INDEX({apat_rng},MATCH({key_expr},{key_rng},0)),\"\")))")
            # M Status blank
            ws.cell(row=row, column=13, value='')
            # Styling per row
            for col_idx in range(1, len(COLUMNS)+1):
                cell = ws.cell(row=row, column=col_idx)
                cell.font = DATA_FONT
                # Alternate subtle fill for readability
                cell.fill = ALT_ROW_FILL if (row % 2 == 1) else WHITE_FILL
                # Alignment: Doctor left, others center
                if col_idx in (3,):
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = BORDER
            ws.row_dimensions[row].height = 20
            row += 1
        cur += timedelta(days=1)

    # Add table with filtering (exclude Status validation row overflow)
    last_row = ws.max_row
    ref = f"A2:{get_column_letter(len(COLUMNS))}{last_row}"  # include header row
    table = Table(displayName='ScheduleFlat', ref=ref)
    style = TableStyleInfo(name='TableStyleMedium9', showRowStripes=False, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    # Status validation (last column)
    dv = DataValidation(type='list', formula1='"OFF,ON/CALL,POST ON/CALL"', allow_blank=True)
    dv.errorTitle = 'Invalid Status'
    dv.error = 'Choose OFF, ON/CALL, POST ON/CALL (blank = normal duty)'
    ws.add_data_validation(dv)
    status_col_letter = get_column_letter(len(COLUMNS))
    dv.add(f'{status_col_letter}3:{status_col_letter}{last_row+300}')

    # Shift validation (column D)
    dv_shift = DataValidation(type='list', formula1='"Shift 1,Shift 2,Custom"', allow_blank=True)
    dv_shift.errorTitle = 'Invalid Shift'
    dv_shift.error = 'Choose Shift 1, Shift 2 or Custom'
    ws.add_data_validation(dv_shift)
    dv_shift.add(f'D3:D{last_row+300}')

    # Freeze title row & date column (A)
    ws.freeze_panes = 'B3'  # Keeps row 1-2 and col A visible while scrolling

    # Compact instructions sheet
    inst = wb.create_sheet('Instructions')
    inst['A1'] = 'MODERN TEMPLATE QUICK GUIDE'
    inst['A1'].font = Font(size=16, bold=True, color='1F4E78')
    guide_lines = [
        '',
        'Editing Tips:',
        '• Enter Start Time as 08:00 (24h).',
        '• Breaks: e.g. 11:00-13:00 or leave blank.',
        '• Timing columns hold ranges (e.g. 08:00-11:00).',
        '• Patient columns hold counts (integers).',
        '• Use Shift to auto-fill timing and breaks from the knowledge base; choose Custom to override.',
        '• Leave Status blank for normal duty; use OFF / ON/CALL / POST ON/CALL.',
        '• Filter on Date or Doctor using header drop-downs.',
        '',
        'Colour Legend:',
        '  White / very light green alternation improves row tracking.',
        '  Green header = structure; DO NOT rename columns.',
        '',
        'Uploading:',
        '1. Fill required cells.',
        '2. Save as .xlsx.',
        '3. Upload via Admin → Bulk Excel.',
        '4. Verify preview status matrix.',
        '5. Apply schedules.',
    ]
    for i, line in enumerate(guide_lines, start=2):
        inst[f'A{i}'] = line
        if line.endswith(':'):
            inst[f'A{i}'].font = Font(size=12, bold=True, color='1F4E78')
        else:
            inst[f'A{i}'].font = Font(size=11)
    inst.column_dimensions['A'].width = 90

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE, date_range, len(doctors)

if __name__ == '__main__':
    fname, drange, doc_count = build_workbook()
    print(f'✅ Modern Excel template created: {fname}')
    print(f'   Date range: {drange}')
    print(f'   Doctors included per day: {doc_count}')
    print('   Frozen: rows 1-2 + column A')
    print('   Table: ScheduleFlat with filters')
    print('   Status validation applied (OFF, ON/CALL, POST ON/CALL)')
    print('\n💡 Ready to edit and upload.')
