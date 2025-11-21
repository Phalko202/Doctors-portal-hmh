"""
Excel Export Module for PR Portal
Formats roster data to match OneDrive Excel sheet structure
Supports Pool A&B and Pool C&D formats
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import json
import os


def load_data():
    """Load roster and staff data"""
    roster_data = {}
    staff_data = {}
    
    roster_path = 'data/pr_roster.json'
    staff_path = 'data/pr_staff.json'
    
    if os.path.exists(roster_path):
        with open(roster_path, 'r', encoding='utf-8') as f:
            roster_data = json.load(f)
    
    if os.path.exists(staff_path):
        with open(staff_path, 'r', encoding='utf-8') as f:
            pr_data = json.load(f)
            staff_data = {s['id']: s for s in pr_data.get('staff', [])}
    
    return roster_data, staff_data


def get_staff_pools(staff_id, staff_data):
    """Get pools for a staff member"""
    staff = staff_data.get(staff_id, {})
    return staff.get('pools', [])


def create_pool_ab_sheet(wb, roster_data, staff_data, start_date, end_date):
    """Create Pool A & B sheet matching the OneDrive format"""
    ws = wb.create_sheet("POOL A & B")
    
    # Colors
    pool_a_color = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    pool_b_color = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    header_color = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    time_color = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Headers
    ws['B1'] = "POOL"
    ws['C1'] = "STATIONS"
    
    # Generate date columns
    current = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    col_index = 4  # Start from column D
    
    date_columns = {}
    while current <= end:
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}1'] = current.strftime('%d-%b-%y')
        ws[f'{col_letter}2'] = current.strftime('%A').upper()[:3]
        date_columns[current.strftime('%Y-%m-%d')] = col_letter
        col_index += 1
        current += timedelta(days=1)
    
    # Pool A Stations
    row = 3
    ws[f'B{row}'] = "POOL-A"
    ws[f'B{row}'].fill = pool_a_color
    
    pool_a_stations = [
        ("DENTAL OPD", "DENT1"),
        ("GP", "GP"),
        ("KANSHI OPD", "KABI"),
        ("CASUALTY", "CYY1"),
        ("RADIOLOGY", "RI"),
        ("ADMISSION / DISCHARGE", "AD"),
        ("AIRPORT", "A1")
    ]
    
    for station_name, station_id in pool_a_stations:
        row += 1
        ws[f'C{row}'] = station_name
        
        # Fill staff data for each date
        for date_str, col in date_columns.items():
            staff_names = []
            timings = []
            
            # Find staff assigned to this station on this date
            for staff_id, dates in roster_data.items():
                if date_str in dates:
                    assignment = dates[date_str]
                    if assignment.get('station') == station_id:
                        staff = staff_data.get(staff_id, {})
                        pools = staff.get('pools', [])
                        if 'A' in pools or 'B' in pools:
                            staff_names.append(staff.get('name', ''))
                            if assignment.get('shift_start') and assignment.get('shift_end'):
                                timings.append(f"{assignment['shift_start']}-{assignment['shift_end']}")
            
            if staff_names:
                cell_value = '\n'.join(staff_names)
                if timings and len(set(timings)) == 1:  # Same timing for all
                    cell_value = f"{station_name}\n{timings[0]}\n" + '\n'.join(staff_names)
                ws[f'{col}{row}'] = cell_value
                ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Pool B Stations
    row += 2
    ws[f'B{row}'] = "POOL-B"
    ws[f'B{row}'].fill = pool_b_color
    
    pool_b_stations = [
        ("RADIOLOGY", "RI"),
        ("ADMISSION / DISCHARGE", "AD"),
        ("AIRPORT", "A1")
    ]
    
    for station_name, station_id in pool_b_stations:
        row += 1
        ws[f'C{row}'] = station_name
        
        for date_str, col in date_columns.items():
            staff_names = []
            for staff_id, dates in roster_data.items():
                if date_str in dates:
                    assignment = dates[date_str]
                    if assignment.get('station') == station_id:
                        staff = staff_data.get(staff_id, {})
                        pools = staff.get('pools', [])
                        if 'B' in pools:
                            staff_names.append(staff.get('name', ''))
            
            if staff_names:
                ws[f'{col}{row}'] = '\n'.join(staff_names)
                ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')
    
    return ws


def create_pool_cd_sheet(wb, roster_data, staff_data, start_date, end_date):
    """Create Pool C & D sheet"""
    ws = wb.create_sheet("POOL C & D")
    
    # Similar structure as Pool A & B
    pool_c_color = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    pool_d_color = PatternFill(start_color="DA9694", end_color="DA9694", fill_type="solid")
    
    ws['A1'] = "POOL"
    ws['B1'] = "STATIONS"
    
    # Date headers
    current = datetime.strptime(start_date, '%Y-%m-%d')
    end = datetime.strptime(end_date, '%Y-%m-%d')
    col_index = 3
    date_columns = {}
    
    while current <= end:
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}1'] = current.strftime('%d-%b-%y')
        date_columns[current.strftime('%Y-%m-%d')] = col_letter
        col_index += 1
        current += timedelta(days=1)
    
    # Pool C & D stations would be similar implementation
    return ws


def export_to_excel(start_date, end_date, output_path='pr_roster_export.xlsx'):
    """Main export function"""
    roster_data, staff_data = load_data()
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Create both pool sheets
    create_pool_ab_sheet(wb, roster_data, staff_data, start_date, end_date)
    create_pool_cd_sheet(wb, roster_data, staff_data, start_date, end_date)
    
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    # Test export
    export_to_excel('2025-11-16', '2025-12-15', 'test_export.xlsx')
    print("Excel export completed!")
