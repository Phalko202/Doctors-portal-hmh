import os
import json
import threading
import re
from datetime import datetime, timedelta
from time import sleep, time
from urllib.request import urlopen, Request
from urllib.parse import urlencode
from flask import Flask, request, jsonify, send_from_directory, render_template, Response, redirect, url_for, stream_with_context, session, send_file
import importlib
import sys
from collections import deque
from io import BytesIO
from hashlib import sha256
import csv
from typing import Optional

# Load .env file if present (production-friendly) without hard import dependency
try:
    _dotenv = importlib.import_module('dotenv')  # type: ignore
    _dotenv.load_dotenv()
except Exception:
    pass


# Try to load local config (optional). If CONFIG_PATH env is set, add it to sys.path first.
CONFIG_PATH = os.environ.get('CONFIG_PATH')
if CONFIG_PATH and CONFIG_PATH not in sys.path:
    sys.path.insert(0, CONFIG_PATH)
try:
    cfg = importlib.import_module('config')
except Exception:
    cfg = None

DATA_PATH = os.path.join('data', 'doctors.json')
USERS_PATH = os.path.join('data', 'users.json')
UPLOAD_DIR = os.path.join('uploads', 'doctor_photos')
PROMO_DIR = os.path.join('uploads', 'doctor_promos')  # Full-screen promotional images
MEDIA_HISTORY_DIR = os.path.join('uploads', 'media_history')  # Archive of all uploaded images
# New: persistent log file path
LOG_PATH = os.path.join('data', 'telegram.log')
# New: logo path and in-memory version for cache busting
LOGO_PATH = os.path.join('static', 'img', 'hulhumale-logo.png')
_LOGO_VERSION = 0
# Settings persistence
SETTINGS_PATH = os.path.join('data', 'settings.json')
 # Added: closure persistence path and detection patterns
CLOSURE_PATH = os.path.join('data', 'closure.json')
CLOSURE_PATTERNS = [
    re.compile(r"\b(opd\s*closed)\b", re.I),
    re.compile(r"\b(clinic\s*closed)\b", re.I),
    re.compile(r"\b(public\s*holiday)\b", re.I),
    re.compile(r"\b(hospital\s*closed)\b", re.I),
    re.compile(r"\b(opd\s*is\s*closed)\b", re.I),
]

 # --- Global configuration variables ---
ADMIN_TOKEN = os.environ.get('ADMIN_TOKEN') or (getattr(cfg, 'ADMIN_TOKEN', '') if cfg else '')  # legacy token disabled by default
SECRET_KEY = os.environ.get('SECRET_KEY', getattr(cfg, 'SECRET_KEY', 'secret!'))
def _clean_env(v: Optional[str]) -> str:
    try:
        if v is None:
            return ''
        # Trim whitespace and surrounding quotes users may include in .env
        v = str(v).strip().strip('"').strip("'")
        return v
    except Exception:
        return ''

TELEGRAM_TOKEN = _clean_env(os.environ.get('TELEGRAM_BOT_TOKEN') or (getattr(cfg, 'TELEGRAM_BOT_TOKEN', '') if cfg else ''))
TELEGRAM_CHAT_ID = _clean_env(os.environ.get('TELEGRAM_GROUP_ID') or (str(getattr(cfg, 'TELEGRAM_GROUP_ID', '')) if cfg else ''))
# Production-safe: only enable Telegram when explicitly requested via env flag
ENABLE_TELEGRAM = os.environ.get('ENABLE_TELEGRAM', '').lower() in ('1','true','yes','on')
ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}  # normalized lower-case
ACCEPT_ANY_CHAT = (os.environ.get('TELEGRAM_ACCEPT_FROM_ANY', 'false').lower() == 'true')
EXTRA_CHAT_IDS = set()
val = os.environ.get('TELEGRAM_CHAT_IDS')
if val:
    for v in val.split(','):
        v = v.strip()
        if v:
            EXTRA_CHAT_IDS.add(str(v))
if TELEGRAM_CHAT_ID:
    EXTRA_CHAT_IDS.add(str(TELEGRAM_CHAT_ID))

_recent = []

def log(msg):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    txt = f"{ts} [telegram] {msg}"
    try:
        print(txt)
    except Exception:
        pass
    try:
        _recent.append(txt)
        if len(_recent) > 200:
            del _recent[:len(_recent)-200]
    except Exception:
        pass
    try:
        os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
        with open(LOG_PATH, 'a', encoding='utf-8') as f:
            f.write(txt + "\n")
    except Exception:
        pass

# --- Settings helpers ---
_settings_cache = None
_settings_lock = threading.Lock()

def _default_settings():
    return {
        'patient_display': {
            'rotate_ms': 12000,  # default 12s
            'show_room': True,
            'show_breaks': True,
            'show_start_time': True,
            'exclusive_multirow_specialty': False,
            'per_specialty_slides': False,  # each specialty on its own slide (even if small)
            'flat_mode': False              # show all doctors ignoring specialties (paginate raw list)
        }
    }

def load_settings():
    global _settings_cache
    with _settings_lock:
        if _settings_cache is not None:
            return _settings_cache
        try:
            with open(SETTINGS_PATH, 'r', encoding='utf-8') as f:
                _settings_cache = json.load(f)
        except Exception:
            _settings_cache = _default_settings()
        return _settings_cache

def save_settings(data=None):
    global _settings_cache
    with _settings_lock:
        if data is not None:
            _settings_cache = data
        else:
            data = _settings_cache or _default_settings()
        os.makedirs(os.path.dirname(SETTINGS_PATH), exist_ok=True)
        tmp = SETTINGS_PATH + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        os.replace(tmp, SETTINGS_PATH)
        return data

# --- Telegram polling state & helpers (relocated after globals) ---
_telegram_thread = None
_telegram_lock = threading.Lock()
_telegram_running = False
_telegram_offset = 0
_telegram_fail_streak = 0  # incremental backoff counter
_processed_commands = {}  # Track recent commands to prevent duplicates (chat_id -> {command: timestamp})


def maybe_start_telegram():
    """Start Telegram long-poll thread once if enabled and token present.
    Reads current environment on each call so changing .env or session envs
    takes effect without a full process restart. Safe to call often."""
    global _telegram_thread, _telegram_running
    global ENABLE_TELEGRAM, TELEGRAM_TOKEN, TELEGRAM_CHAT_ID, EXTRA_CHAT_IDS

    # Re-compute config from current environment each time
    ENABLE_TELEGRAM = os.environ.get('ENABLE_TELEGRAM', '').lower() in ('1','true','yes','on')
    token_env = _clean_env(os.environ.get('TELEGRAM_BOT_TOKEN'))
    group_env = _clean_env(os.environ.get('TELEGRAM_GROUP_ID'))
    if token_env:
        TELEGRAM_TOKEN = token_env
    if group_env:
        TELEGRAM_CHAT_ID = str(group_env)
    # Refresh allowlist from env
    EXTRA_CHAT_IDS = set()
    val = _clean_env(os.environ.get('TELEGRAM_CHAT_IDS'))
    if val:
        for v in val.split(','):
            v = (v or '').strip()
            if v:
                EXTRA_CHAT_IDS.add(str(v))
    if TELEGRAM_CHAT_ID:
        EXTRA_CHAT_IDS.add(str(TELEGRAM_CHAT_ID))

    # Do not start unless explicitly enabled AND well configured
    if not ENABLE_TELEGRAM or not TELEGRAM_TOKEN:
        return
    accept_any = os.environ.get('TELEGRAM_ACCEPT_FROM_ANY','').lower() in ('1','true','yes','on')
    # If a group allowlist is required but not configured, don't start unless ACCEPT_ANY_CHAT true or EXTRA_CHAT_IDS present
    if not (accept_any or TELEGRAM_CHAT_ID or EXTRA_CHAT_IDS):
        log('telegram disabled: no GROUP_ID/allowlist; set TELEGRAM_GROUP_ID or TELEGRAM_ACCEPT_FROM_ANY=true')
        return
    # In debug with reloader only run in the reloader (WERKZEUG_RUN_MAIN == 'true') process (guard if app not yet created)
    if 'app' in globals() and getattr(globals()['app'], 'debug', False) and os.environ.get('WERKZEUG_RUN_MAIN') != 'true':
        return
    with _telegram_lock:
        if _telegram_thread and _telegram_thread.is_alive():
            return
        _telegram_running = True
        _telegram_thread = threading.Thread(target=_telegram_loop, name="telegram-poll", daemon=True)
        _telegram_thread.start()


def _telegram_loop():
    """Background loop performing getUpdates long polling and dispatching messages.
    Enhanced: richer logging, 409 conflict mitigation, adaptive backoff, periodic webhook delete retry.
    """
    global _telegram_running, _telegram_offset, _telegram_fail_streak
    base = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
    log("telegram poll thread started")
    # Ensure webhook removed (once now, periodically later)
    try:
        urlopen(f"{base}/deleteWebhook")
    except Exception:
        pass
    import random, time as _t
    last_webhook_clear = _t.time()
    consecutive_409 = 0
    while _telegram_running and TELEGRAM_TOKEN:
        if _t.time() - last_webhook_clear > 600:  # every 10 minutes
            try:
                urlopen(f"{base}/deleteWebhook")
                last_webhook_clear = _t.time()
            except Exception:
                pass
        try:
            # Always send offset = last_seen + 1 per Telegram docs (omit if zero)
            params = {'timeout': 25}
            if _telegram_offset:
                params['offset'] = _telegram_offset + 1
            # Could scope updates in future: params['allowed_updates'] = json.dumps(['message','channel_post'])
            query = urlencode(params)
            url = f"{base}/getUpdates?{query}"
            with urlopen(Request(url, headers={'User-Agent':'schedule-bot'}), timeout=35) as resp:
                raw = resp.read().decode('utf-8', errors='replace')
            try:
                data = json.loads(raw)
            except Exception as je:
                _telegram_fail_streak += 1
                log(f"json decode error: {je}; raw(trunc)={raw[:120]!r}")
                sleep(min(30, 2 + _telegram_fail_streak))
                continue
            if not data.get('ok'):
                _telegram_fail_streak += 1
                log(f"api not ok: {data}")
                sleep(min(30, 2 + _telegram_fail_streak))
                continue
            # Success path resets fail streak & 409 counter
            _telegram_fail_streak = 0
            consecutive_409 = 0
            updates = data.get('result', []) or []
            if updates:
                log(f"fetched {len(updates)} updates (offset now {updates[-1].get('update_id')})")
            got_any = False
            for upd in updates:
                got_any = True
                upid = upd.get('update_id', 0)
                if upid <= _telegram_offset:
                    # Already processed (duplicate delivery) – skip
                    continue
                _telegram_offset = upid
                msg = upd.get('message') or upd.get('channel_post') or {}
                if not msg:
                    continue
                chat_id = str((msg.get('chat') or {}).get('id', ''))
                text = msg.get('text') or msg.get('caption') or ''
                if not text:
                    continue
                if not ACCEPT_ANY_CHAT and EXTRA_CHAT_IDS and (chat_id not in EXTRA_CHAT_IDS):
                    continue  # ignore unauthorized chats
                try:
                    process_telegram_message(text, chat_id)
                except Exception as e:
                    log(f"msg error: {e}")
            # If no updates returned, jitter sleep avoids tight loop
            if not got_any:
                sleep(0.9 + random.uniform(0, 0.9))
        except Exception as e:
            code = getattr(e, 'code', None)
            if code == 409:
                consecutive_409 += 1
                # Minimal delay approach – keep trying rapidly, silently most of the time
                if consecutive_409 % 5 == 1:
                    log(f"409 conflict (#{consecutive_409}) – fast retry")
                # Aggressively clear webhook every 10th 409
                if consecutive_409 % 10 == 0:
                    try: urlopen(f"{base}/deleteWebhook"); last_webhook_clear = _t.time()
                    except Exception: pass
                sleep(1.2)
                continue
            consecutive_409 = 0
            _telegram_fail_streak += 1
            # Softer backoff cap (do not exceed 12s) to keep bot responsive
            backoff = min(12, 2 ** min(4, _telegram_fail_streak))
            log(f"poll error (streak={_telegram_fail_streak}): {e} -> retry in {backoff}s")
            sleep(backoff + random.uniform(0, 0.8))


def _telegram_status():
    # Compute enabled from current environment as well as initial globals,
    # so the Admin page reflects .env or session variable changes without restart.
    enabled_env = os.environ.get('ENABLE_TELEGRAM', '').lower() in ('1','true','yes','on')
    token_env = os.environ.get('TELEGRAM_BOT_TOKEN') or TELEGRAM_TOKEN
    enabled = (ENABLE_TELEGRAM or enabled_env) and bool(token_env)
    return {
        'enabled': bool(enabled),
        'running': _telegram_thread.is_alive() if _telegram_thread else False,
        'offset': _telegram_offset,
        'fail_streak': _telegram_fail_streak
    }

# Ensure Flask app exists before applying route decorators
if 'app' not in globals():
    app = Flask(__name__)
    app.secret_key = SECRET_KEY

# ---- Patient display settings API ----
@app.get('/api/patient_display/settings')
def api_get_patient_display_settings():
    s = load_settings()
    return jsonify({'ok': True, 'settings': s.get('patient_display', {})})

@app.post('/api/patient_display/settings')
def api_update_patient_display_settings():
    g = require_roles(ROLE_ADMIN)
    if g: return g
    data = request.get_json(silent=True) or {}
    rotate_ms = data.get('rotate_ms')
    try:
        rotate_ms = int(rotate_ms)
    except Exception:
        return jsonify({'ok': False, 'error': 'rotate_ms integer required'}), 400
    rotate_ms = max(3000, min(300000, rotate_ms))
    # Optional display resolution and fill mode
    display_resolution = data.get('display_resolution')
    fill_mode = data.get('fill_mode')
    if display_resolution is not None:
        try:
            display_resolution = str(display_resolution)
        except Exception:
            display_resolution = None
    if fill_mode is not None:
        fill_mode = str(fill_mode)
        if fill_mode not in ('auto', 'cover', 'contain'):
            fill_mode = 'auto'
    # Optional boolean flags
    def as_bool(v, default=None):
        if isinstance(v, bool): return v
        if v is None: return default
        if isinstance(v, (int, float)): return bool(v)
        if isinstance(v, str):
            return v.lower() in ('1','true','yes','on')
        return default
    show_room = as_bool(data.get('show_room'), True)
    show_breaks = as_bool(data.get('show_breaks'), True)
    show_start_time = as_bool(data.get('show_start_time'), True)
    exclusive_multi = as_bool(data.get('exclusive_multirow_specialty'), False)
    per_spec = as_bool(data.get('per_specialty_slides'), False)
    flat_mode = as_bool(data.get('flat_mode'), False)
    s = load_settings()
    if 'patient_display' not in s:
        s['patient_display'] = {}
    s['patient_display']['rotate_ms'] = rotate_ms
    s['patient_display']['show_room'] = show_room
    s['patient_display']['show_breaks'] = show_breaks
    s['patient_display']['show_start_time'] = show_start_time
    s['patient_display']['exclusive_multirow_specialty'] = exclusive_multi
    s['patient_display']['per_specialty_slides'] = per_spec
    s['patient_display']['flat_mode'] = flat_mode
    if display_resolution is not None:
        s['patient_display']['display_resolution'] = display_resolution
    if fill_mode is not None:
        s['patient_display']['fill_mode'] = fill_mode
    save_settings(s)
    try:
        broker.publish_event('patient_display_settings', s['patient_display'])
    except Exception:
        pass
    return jsonify({'ok': True, 'settings': s['patient_display']})

@app.get('/api/schedule/template.xlsx')
def api_schedule_template():
    """Serve an Excel (.xlsx) template for bulk schedule upload.

    Updated spec (2025-09-28):
      - Removed columns: Status, Designation, Reason (status is implied by providing data; leaves managed via Telegram)
      - Renamed: 'Patient Count' -> 'Patient Details' (free-form text instead of number)
      - Current column order:
            Date, Doctor, Start Time, Room, Patient Details, Before Break OPD Patients, Breaks, After Break OPD Patients
      - Accept optional query params start=YYYY-MM-DD & end=YYYY-MM-DD to pre-populate blank rows for every doctor for each date in range.
        If omitted, include one sample row only.
      - For each doctor/date row only Date + Doctor filled; other cells blank for user to complete.
      - If a doctor is left entirely blank for a date in upload parsing (future implementation) it will be treated as 'awaiting schedule'.

    We attempt to decorate the sheet (header style, table style, column widths, freeze header).
    Fallback if openpyxl missing: generate CSV with the new columns.
    """
    cols = [
        'Date','Doctor','Start Time','Room','Total Patients',
        'Before Break OPD Patients (time/pts)','Breaks','After Break OPD Patients (time/pts)','Status'
    ]

    # Parse optional date range
    start_q = request.args.get('start')
    end_q = request.args.get('end')
    date_rows = []
    def parse_d(s):
        try:
            return datetime.strptime(s, '%Y-%m-%d').date()
        except Exception:
            return None
    start_d = parse_d(start_q) if start_q else None
    end_d = parse_d(end_q) if end_q else None
    if start_d and end_d and end_d >= start_d and (end_d - start_d).days <= 120:
        # Build list of (date, doctor_name) pairs
        try:
            data = load_data()
            doctors = sorted(data.get('doctors', []), key=lambda d: d.get('name',''))
        except Exception:
            doctors = []
        cur = start_d
        while cur <= end_d:
            # Skip Fridays (auto-closed)
            if cur.weekday() == 4:  # 0=Mon, 4=Fri
                cur += timedelta(days=1)
                continue
            for d in doctors:
                date_rows.append((cur.isoformat(), d.get('name','')))
            cur += timedelta(days=1)
    grouped = request.args.get('group','').lower() in ('1','true','yes','y','grouped') and date_rows

    # Try openpyxl for rich formatting
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
        wb = Workbook()
        ws = wb.active
        ws.title = 'Schedule'
        if not grouped:
            # ORIGINAL FLAT RANGE LAYOUT
            ws.append(cols)
            if date_rows:
                for dt, doc_name in date_rows:
                    ws.append([dt, doc_name, '', '', '', '', '', '', ''])
            else:
                ws.append([datetime.utcnow().date().isoformat(),'Dr. Example','08:00','K101','35','20','11:00-12:00','15',''])
            # Add data validation for Status column (column I, 9th column)
            from openpyxl.worksheet.datavalidation import DataValidation
            status_validation = DataValidation(type="list", formula1='"ON/CALL,POST ON/CALL,OFF"', allow_blank=True)
            status_validation.error = 'Please select from: ON/CALL, POST ON/CALL, OFF'
            status_validation.errorTitle = 'Invalid Status'
            ws.add_data_validation(status_validation)
            # Apply validation to Status column (I) for all data rows (skip header)
            status_validation.add(f'I2:I{ws.max_row+1000}')  # Apply to many rows
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            widths = [12, 34, 10, 10, 26, 22, 18, 24, 14]  # Widen before/after columns for (time/pts)
            for idx, w in enumerate(widths, start=1):
                ws.column_dimensions[chr(64+idx)].width = w
            ws.freeze_panes = 'A2'
            last_row = ws.max_row
            ref = f"A1:I{last_row}"
            tbl = Table(displayName='ScheduleTable', ref=ref)
            style = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True, showColumnStripes=False)
            tbl.tableStyleInfo = style
            ws.add_table(tbl)
            thin = Side(style='thin', color='CCCCCC')
            for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=9):
                for c in r:
                    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    if c.column == 1:
                        c.alignment = Alignment(horizontal='center')
        else:
            # GROUPED LAYOUT: Date headers & specialty sections without repeating date each row
            # Prepare doctors grouped by specialty once
            try:
                data = load_data()
                doctors_all = data.get('doctors', [])
                spec_order = data.get('specialty_order', [])
            except Exception:
                doctors_all = []
                spec_order = []
            title = f"Doctor Schedule Template ({start_d.isoformat()} to {end_d.isoformat()})" if (start_d and end_d) else 'Doctor Schedule Template'
            ws.append([title])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center')
            # (Removed extra blank row to ensure freeze pane behaves consistently)
            cur_row = ws.max_row
            # Column widths for grouped layout (added Status column)
            # With a repeated Date column (frozen), we now have 10 columns in grouped layout
            widths = [12, 34, 10, 10, 30, 22, 18, 24, 14, 4]
            for idx, w in enumerate(widths, start=1):
                ws.column_dimensions[chr(64+idx)].width = w
            # Modern professional color scheme
            date_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')  # Teal green
            date_font = Font(color='FFFFFF', bold=True, size=12)
            spec_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')  # Light green
            spec_font = Font(bold=True, color='065F46', size=11)  # Dark green
            header_fill2 = PatternFill(start_color='E0F2FE', end_color='E0F2FE', fill_type='solid')  # Light blue
            header_font2 = Font(bold=True, color='0C4A6E', size=10)  # Dark blue
            thin = Side(style='thin', color='9CA3AF')  # Gray borders
            medium = Side(style='medium', color='6B7280')  # Darker borders for headers
            # iterate dates
            if start_d and end_d:
                day = start_d
                while day <= end_d:
                    # Skip Fridays (auto-closed)
                    if day.weekday() == 4:
                        day += timedelta(days=1)
                        continue
                    # Format date with day name and ordinal: "1st Saturday 2025 1/11/2025"
                    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                    day_name = day_names[day.weekday()]
                    day_num = day.day
                    if 10 <= day_num % 100 <= 20:
                        suffix = 'th'
                    else:
                        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day_num % 10, 'th')
                    date_formatted = f"{day_num}{suffix} {day_name} {day.year} {day.strftime('%d/%m/%Y')}"
                    # Date header row
                    ws.append([date_formatted])
                    r = ws.max_row
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
                    cell = ws.cell(row=r, column=1)
                    cell.fill = date_fill; cell.font = date_font; cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(top=medium, left=medium, right=medium, bottom=medium)
                    # For each specialty add section
                    for spec in spec_order or sorted({d.get('specialty') for d in doctors_all}):
                        spec_docs = [d for d in doctors_all if d.get('specialty') == spec]
                        if not spec_docs:
                            continue
                        ws.append([spec])
                        rs = ws.max_row
                        ws.merge_cells(start_row=rs, start_column=1, end_row=rs, end_column=10)
                        scell = ws.cell(row=rs, column=1)
                        scell.fill = spec_fill; scell.font = spec_font; scell.alignment = Alignment(horizontal='left', vertical='center')
                        scell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        # Column headers for this specialty block (no Date column here, but includes Status)
                        ws.append(['Date','Doctor','Start Time','Room','Total Patients','Before Break OPD Patients (time/pts)','Breaks','After Break OPD Patients (time/pts)','Status',''])
                        hr = ws.max_row
                        for c in ws.iter_rows(min_row=hr, max_row=hr, min_col=1, max_col=10):
                            for cell2 in c:
                                cell2.fill = header_fill2; cell2.font = header_font2; cell2.alignment = Alignment(horizontal='center', vertical='center')
                                cell2.border = Border(top=thin, left=thin, right=thin, bottom=medium)
                        # Doctor rows
                        for d in sorted(spec_docs, key=lambda x: x.get('name','')):
                            # Repeat the date in the first column so it stays visible when we freeze column A
                            ws.append([day.strftime('%d/%m/%Y'), d.get('name',''),'','','','','','','',''])
                            rr = ws.max_row
                            for c in ws.iter_rows(min_row=rr, max_row=rr, min_col=1, max_col=10):
                                for cell3 in c:
                                    cell3.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        # Spacer row
                        ws.append([''])
                    # Extra spacer between dates
                    ws.append([''])
                    day += timedelta(days=1)
            else:
                # Fallback single sample block
                ws.append(['Sample Date'])
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=10)
                ws.append(['Sample Specialty'])
                ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=10)
                ws.append(['Date','Doctor','Start Time','Room','Total Patients','Before Break OPD Patients (time/pts)','Breaks','After Break OPD Patients (time/pts)','Status',''])
                ws.append(['01/01/2026','Dr. Example','08:00','K101','35','20 / 11:00-12:00','11:00-12:00','15','', ''])
            # Add data validation for Status column (I in grouped layout with Date column)
            from openpyxl.worksheet.datavalidation import DataValidation
            status_validation = DataValidation(type="list", formula1='"ON/CALL,POST ON/CALL,OFF"', allow_blank=True)
            status_validation.error = 'Please select from: ON/CALL, POST ON/CALL, OFF (leave blank for normal schedule)'
            status_validation.errorTitle = 'Invalid Status'
            status_validation.prompt = 'ON/CALL = on-call duty | POST ON/CALL = post on-call (can have OPD) | OFF = day off (no OPD)'
            status_validation.promptTitle = 'Status Options'
            ws.add_data_validation(status_validation)
            # Apply to Status column I for all rows
            status_validation.add(f'I2:I{ws.max_row+1000}')
            
            # Freeze the title row and the first column (Date) so the date is always visible while scrolling
            ws.freeze_panes = 'B2'
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        fname = 'template.xlsx'
        if start_d and end_d:
            fname = f"template_{start_d.isoformat()}_{end_d.isoformat()}{'_grouped' if grouped else ''}.xlsx"
        return send_file(bio, as_attachment=True, download_name=fname, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception:
        # Fallback: simple CSV (without formatting) – still provides rows.
        lines = [','.join(cols)]
        if date_rows:
            for dt, doc_name in date_rows:
                lines.append(f"{dt},{doc_name},,,,(leave blank or '11:00-12:00/20'),(leave blank),(leave blank or '14:00-15:00/15'),(ON/CALL or POST ON/CALL or OFF)")
        else:
            lines.append(f"{datetime.utcnow().date().isoformat()},Dr. Example,08:00,K101,(e.g. 35 tokens),20/11:00-12:00,11:00-12:00,15/14:00-15:00,")
        bio = BytesIO('\n'.join(lines).encode('utf-8'))
        return send_file(bio, as_attachment=True, download_name='template.csv', mimetype='text/csv')

@app.post('/api/schedule/bulk_excel')
def api_schedule_bulk_excel():
    """Upload an Excel (.xlsx) with rows: Date, Doctor, Start Time, Room, Total Patients,
    Before Break OPD Patients, Breaks, After Break OPD Patients.
    Applies per-date updates and notifies displays via SSE.
    """
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': 'empty filename'}), 400
    try:
        from openpyxl import load_workbook
    except Exception:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500

    # Helpers
    def _cell_str(v):
        if v is None:
            return ''
        if isinstance(v, (int, float)):
            # Excel times sometimes come as floats; we'll ignore here
            return str(v)
        return str(v).strip()

    def _parse_date(v) -> Optional[str]:
        from datetime import datetime, date
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, date):
            return v.isoformat()
        s = _cell_str(v)
        if not s:
            return None
        # try ISO first
        try:
            return datetime.strptime(s, '%Y-%m-%d').date().isoformat()
        except Exception:
            pass
        # try dd/mm/yyyy
        try:
            d, m, y = s.replace('.', '/').replace('-', '/').split('/')
            return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
        except Exception:
            return None

    def _parse_breaks(v):
        s = _cell_str(v)
        if not s:
            return None
        items = [p.strip() for p in s.split(',') if p.strip()]
        return items or None

    def _extract_time_and_count(v):
        s = _cell_str(v)
        if not s:
            return (None, None)
        # Find first time range like HH:MM(:SS)?-HH:MM(:SS)?
        m = re.search(r'(\d{1,2}:\d{2}(?::\d{2})?)\s*-\s*(\d{1,2}:\d{2}(?::\d{2})?)', s)
        time_txt = None
        rest = s
        if m:
            a, b = m.group(1), m.group(2)
            def norm(t):
                mm = re.match(r'^(\d{1,2}):(\d{2})', t)
                return f"{int(mm.group(1)):02d}:{mm.group(2)}" if mm else t
            time_txt = f"{norm(a)}-{norm(b)}"
            # Remove matched range for number search
            rest = (s[:m.start()] + ' ' + s[m.end():]).strip()
        # Find an integer not part of the time
        m2 = re.findall(r'\b\d{1,3}\b', rest)
        count = int(m2[-1]) if m2 else None
        return (time_txt, count)

    def _match_doctor(name: str, doctors: list):
        toks = set(name_tokens(name))
        if not toks:
            return None
        best = None; score = 0
        for d in doctors:
            dt = set(name_tokens(d.get('name','')))
            if dt and dt.issubset(toks) and len(dt) > score:
                best = d; score = len(dt)
        return best

    # Load workbook
    try:
        bio = BytesIO(f.read())
        wb = load_workbook(bio, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'ok': False, 'error': f'failed to read xlsx: {e}'}), 400

    # Map headers
    headers = {}
    first = True
    applied = 0
    errors = []
    data = load_data()
    doc_list = data.get('doctors', [])
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if first:
            first = False
            for idx, h in enumerate(row):
                key = _cell_str(h).lower()
                headers[key] = idx
            continue
        def col(name):
            idx = headers.get(name)
            if idx is None:
                return None
            try:
                return row[idx]
            except Exception:
                return None
        date_iso = _parse_date(col('date'))
        doc_name = _cell_str(col('doctor'))
        if not date_iso or not doc_name:
            # Allow empty rows
            continue
        # Exclude Fridays (auto-closed): ignore rows for Fridays
        try:
            if datetime.strptime(date_iso, '%Y-%m-%d').date().weekday() == 4:
                errors.append({'row': i, 'error': 'Friday is auto-closed; row ignored', 'doctor': doc_name})
                continue
        except Exception:
            pass
        doc = _match_doctor(doc_name, doc_list)
        if not doc:
            errors.append({'row': i, 'error': 'doctor not found', 'doctor': doc_name})
            continue
        patch = {}
        st = _cell_str(col('start time'))
        if st:
            patch['start_time'] = st
            patch['status'] = 'ON_DUTY'
        rm = _cell_str(col('room'))
        if rm:
            patch['room'] = rm
        tot = _cell_str(col('total patients'))
        if tot:
            try: patch['patient_count'] = int(float(tot))
            except Exception: pass
        breaks_accum = []
        bb_raw = _cell_str(col('before break opd patients')) or _cell_str(col('before break opd patients (time/pts)'))
        if bb_raw:
            rng, cnt = _extract_time_and_count(bb_raw)
            if cnt is not None:
                try: patch['before_break_opd_patients'] = int(float(cnt))
                except Exception: pass
            if rng:
                breaks_accum.append(rng)
        br = _parse_breaks(col('breaks'))
        if br:
            # Check for "NO BREAK" explicitly
            if any(re.search(r'no\s*break', b, re.I) for b in br):
                patch['breaks'] = ['NO BREAK']
                breaks_accum = []  # explicit override
            else:
                patch['breaks'] = br
        ab_raw = _cell_str(col('after break opd patients')) or _cell_str(col('after break opd patients (time/pts)'))
        if ab_raw:
            rng, cnt = _extract_time_and_count(ab_raw)
            if cnt is not None:
                try: patch['after_break_opd_patients'] = int(float(cnt))
                except Exception: pass
            if rng:
                breaks_accum.append(rng)
        if breaks_accum and patch.get('breaks') != ['NO BREAK']:
            # Merge with parsed breaks (if any) and de-dup
            existing = set(patch.get('breaks') or [])
            for r in breaks_accum:
                if r not in existing:
                    existing.add(r)
            if existing:
                patch['breaks'] = list(existing)
        
        # NEW: Status column (H) - special duty status markers
        status_val = _cell_str(col('status')).upper()
        if status_val:
            if status_val in ('ON/CALL', 'ON CALL', 'ONCALL', 'ON-CALL'):
                patch['status'] = 'ON_CALL'
                # ON_CALL can have OPD if other fields provided
            elif status_val in ('POST ON/CALL', 'POST ONCALL', 'POST ON CALL', 'POST-ONCALL', 'POST-ON-CALL'):
                patch['status'] = 'ON_DUTY'
                patch['post_oncall'] = True
                # Can have OPD details
            elif status_val == 'OFF':
                patch['status'] = 'OFF_DUTY'
                # Clear OPD fields for OFF status - doctor is not working
                patch['start_time'] = None
                patch['room'] = None
                patch['patient_count'] = None
                patch['before_break_opd_patients'] = None
                patch['after_break_opd_patients'] = None
                patch['breaks'] = None
        
        # If either before or after patient count is missing (empty cell), default to 0
        # This prevents the "no patient" issue when cells are left empty
        # BUT: only if status is not OFF_DUTY
        if 'start_time' in patch and 'patient_count' not in patch and patch.get('status') != 'OFF_DUTY':
            # If start time exists but counts are missing, set defaults
            if 'before_break_opd_patients' not in patch:
                patch['before_break_opd_patients'] = 0
            if 'after_break_opd_patients' not in patch:
                patch['after_break_opd_patients'] = 0
        if not patch:
            continue
        if apply_single_day_update(doc, date_iso, patch):
            applied += 1
    if applied:
        save_data(data)
        try:
            broker.publish_event('doctor_update', {'bulk': True})
        except Exception:
            pass
    return jsonify({'ok': True, 'applied': applied, 'errors': errors})

# ==================== NEW 2-STAGE EXCEL UPLOAD SYSTEM ====================

@app.post('/api/schedule/store_excel')
def api_schedule_store_excel():
    """Stage 1: Upload and store Excel file without processing."""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': 'empty filename'}), 400
    
    # Create uploads/schedules directory (optionally within a subfolder)
    import time
    schedule_root = os.path.join('uploads', 'schedules')
    os.makedirs(schedule_root, exist_ok=True)
    # Optional subpath where to store the file, e.g. '2025/Nov'
    subpath = (request.args.get('path') or request.form.get('path') or '').strip().replace('\\','/')
    def _safe_path(base, rel):
        rel = (rel or '').strip().strip('/')
        target = os.path.normpath(os.path.join(base, rel))
        base_abs = os.path.abspath(base)
        target_abs = os.path.abspath(target)
        if not target_abs.startswith(base_abs):
            raise ValueError('invalid path')
        return target_abs
    try:
        schedule_dir = _safe_path(schedule_root, subpath)
    except Exception:
        schedule_dir = schedule_root
    os.makedirs(schedule_dir, exist_ok=True)
    
    # Generate unique filename with timestamp
    timestamp = int(time.time())
    ext = os.path.splitext(f.filename)[1]
    safe_name = f.filename.replace(' ', '_').replace('..', '')
    stored_filename = f"{timestamp}_{safe_name}"
    filepath = os.path.join(schedule_dir, stored_filename)
    
    try:
        f.save(filepath)
        return jsonify({
            'ok': True,
            'filename': stored_filename,
            'original_name': f.filename,
            'uploaded_at': timestamp,
            'message': 'File uploaded successfully. Go to Stage 2 to review and process.',
            'path': os.path.relpath(schedule_dir, schedule_root).replace('\\','/')
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Failed to save file: {str(e)}'}), 500

@app.get('/api/schedule/list_excel')
def api_schedule_list_excel():
    """List uploaded Excel files under a given path (no directories)."""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    
    schedule_root = os.path.join('uploads', 'schedules')
    os.makedirs(schedule_root, exist_ok=True)
    subpath = (request.args.get('path') or '').strip().replace('\\','/')
    def _safe_path(base, rel):
        rel = (rel or '').strip().strip('/')
        target = os.path.normpath(os.path.join(base, rel))
        base_abs = os.path.abspath(base)
        target_abs = os.path.abspath(target)
        if not target_abs.startswith(base_abs):
            raise ValueError('invalid path')
        return target_abs
    try:
        schedule_dir = _safe_path(schedule_root, subpath)
    except Exception:
        schedule_dir = schedule_root
    if not os.path.exists(schedule_dir):
        return jsonify({'ok': True, 'files': []})
    
    files = []
    for fname in os.listdir(schedule_dir):
        if fname.endswith(('.xlsx', '.xls')):
            filepath = os.path.join(schedule_dir, fname)
            stat = os.stat(filepath)
            # Extract timestamp from filename
            try:
                timestamp = int(fname.split('_')[0])
                from datetime import datetime
                uploaded_date = datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M')
            except:
                uploaded_date = 'Unknown'
            
            files.append({
                'filename': fname,
                'uploaded_at': uploaded_date,
                'size_kb': round(stat.st_size / 1024, 2)
            })
    
    # Sort by filename (timestamp) descending
    files.sort(key=lambda x: x['filename'], reverse=True)
    return jsonify({'ok': True, 'files': files})

@app.post('/api/schedule/delete_excel')
def api_schedule_delete_excel():
    """Delete a previously uploaded Excel file by filename."""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    filename = (payload.get('filename') or '').strip()
    if not filename:
        return jsonify({'ok': False, 'error': 'filename required'}), 400
    schedule_root = os.path.join('uploads', 'schedules')
    # Support nested paths in filename like '2025/Nov/file.xlsx'
    filename = filename.replace('\\','/')
    target = os.path.normpath(os.path.join(schedule_root, filename))
    root_abs = os.path.abspath(schedule_root)
    target_abs = os.path.abspath(target)
    if not target_abs.startswith(root_abs):
        return jsonify({'ok': False, 'error': 'invalid path'}), 400
    filepath = target_abs
    if not os.path.isfile(filepath):
        return jsonify({'ok': False, 'error': 'file not found'}), 404
    try:
        os.remove(filepath)
        return jsonify({'ok': True, 'deleted': filename})
    except Exception as e:
        return jsonify({'ok': False, 'error': f'failed to delete: {e}'}), 500

@app.post('/api/schedule/preview_excel')
def api_schedule_preview_excel():
    """Stage 2a: Preview statistics from uploaded Excel file without applying."""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    
    filename = request.json.get('filename')
    if not filename:
        return jsonify({'ok': False, 'error': 'filename required'}), 400
    # Support nested path under schedules, with safety
    schedule_root = os.path.join('uploads', 'schedules')
    os.makedirs(schedule_root, exist_ok=True)
    fname_norm = filename.replace('\\','/')
    filepath = os.path.normpath(os.path.join(schedule_root, fname_norm))
    if not os.path.abspath(filepath).startswith(os.path.abspath(schedule_root)):
        return jsonify({'ok': False, 'error': 'invalid path'}), 400
    if not os.path.exists(filepath):
        return jsonify({'ok': False, 'error': 'File not found'}), 404
    
    try:
        from openpyxl import load_workbook
        from datetime import datetime, date
    except Exception:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500
    
    def _cell_str(v):
        if v is None: return ''
        return str(v).strip()
    
    def _parse_date(v):
        import re
        if v is None: return None
        if isinstance(v, datetime): return v.date().isoformat()
        if isinstance(v, date): return v.isoformat()
        s = _cell_str(v)
        if not s: return None
        try:
            return datetime.strptime(s, '%Y-%m-%d').date().isoformat()
        except:
            pass
        # Try to find a date anywhere in the string (handles e.g. "2nd Sunday 2025 02/11/2025")
        m = re.search(r'(\d{4})[./-](\d{1,2})[./-](\d{1,2})', s)
        if m:
            y, mo, d = map(int, m.groups())
            return f"{y:04d}-{mo:02d}-{d:02d}"
        m = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})', s)
        if m:
            d, mo, y = m.groups()
            y = int(y)
            if y < 100: y += 2000
            return f"{y:04d}-{int(mo):02d}-{int(d):02d}"
        return None
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        print(f"DEBUG Preview: Loaded workbook with {ws.max_row} rows and {ws.max_column} columns")
        # Print first few rows for debugging
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i <= 5:
                print(f"DEBUG Row {i}: {[_cell_str(v) for v in row]}")
    except Exception as e:
        print(f"ERROR Preview: Failed to read file: {str(e)}")
        return jsonify({'ok': False, 'error': f'Failed to read file: {str(e)}'}), 400
    
    # Parse file and collect statistics
    headers = {}
    first = True
    date_range = set()
    doctor_stats = {}
    total_rows = 0
    # For richer preview: map matched system doctor -> per-date status
    # States: ADDED (has schedule details), OFF (explicit off), EMPTY (row present but no details), MISSING (no row for that date)
    data_all = load_data()
    doc_list = data_all.get('doctors', [])

    def _match_doctor(name: str, doctors: list):
        toks = set(name_tokens(name))
        if not toks:
            return None
        best = None; score = 0
        for d in doctors:
            dt = set(name_tokens(d.get('name','')))
            if dt and dt.issubset(toks) and len(dt) > score:
                best = d; score = len(dt)
        return best

    per_doc_date = {}  # id -> { date_iso: state }
    
    # Header normalization map
    def norm_key(k: str) -> str:
        k = (k or '').strip().lower()
        aliases = {
            'doctor name': 'doctor', 'name': 'doctor', 'dr': 'doctor', 'dr.': 'doctor',
            'physician': 'doctor', 'doctor/physician': 'doctor',
            'starting time': 'start time', 'start': 'start time', 'starttime': 'start time',
            'room no': 'room', 'room number': 'room',
            'total no of patients': 'total patients', 'total patients': 'total patients', 'total': 'total patients',
            # Support both old combined format and new separate columns
            'before break opd patients (time/pts)': 'before break opd patients combined',
            'before break opd timing': 'before break opd timing',
            'before break opd patients': 'before break opd patients',
            'before break patients': 'before break opd patients',
            'after break opd patients (time/pts)': 'after break opd patients combined',
            'after break opd timing': 'after break opd timing',
            'after break opd patients': 'after break opd patients',
            'after break patients': 'after break opd patients',
            # New modern template columns
            'shift': 'shift',
            'specialty': 'specialty', 'speciality': 'specialty', 'department': 'specialty', 'dept': 'specialty',
        }
        return aliases.get(k, k)

    # Try to find header row (might not be first row)
    header_row_index = None
    # Load shift knowledge once
    shift_knowledge = {}
    try:
        shift_knowledge = load_shift_knowledge()
    except Exception:
        shift_knowledge = {}

    def _find_specialty_match(spec_val: str) -> str:
        if not spec_val:
            return None
        s = spec_val.strip().lower()
        for k in shift_knowledge.keys():
            if k.strip().lower() == s:
                return k
        return spec_val  # return as-is if not found; caller will still try

    def _apply_shift_defaults(patch: dict, spec: str, shift_name: str):
        if not shift_name:
            return
        s_key = _find_specialty_match(spec or '') if spec else None
        if not s_key:
            return
        spec_data = shift_knowledge.get(s_key) or {}
        sh = (shift_name or '').strip() or ''
        defaults = spec_data.get(sh)
        if not isinstance(defaults, dict):
            return
        applied_any = False
        # Start time
        if 'start_time' not in patch and defaults.get('start'):
            patch['start_time'] = defaults.get('start')
            applied_any = True
        # Before/After patients
        if 'before_break_opd_patients' not in patch and defaults.get('before_patients') not in (None, ''):
            try:
                patch['before_break_opd_patients'] = int(defaults.get('before_patients'))
            except Exception:
                pass
            else:
                applied_any = True
        if 'after_break_opd_patients' not in patch and defaults.get('after_patients') not in (None, ''):
            try:
                patch['after_break_opd_patients'] = int(defaults.get('after_patients'))
            except Exception:
                pass
            else:
                applied_any = True
        # Breaks/timings bucket
        if 'breaks' not in patch:
            accum = []
            bt = (defaults.get('before_timing') or '').strip()
            if bt:
                accum.append(bt)
            brk = (defaults.get('breaks') or '').strip()
            if brk:
                accum.append(brk)
            at = (defaults.get('after_timing') or '').strip()
            if at:
                accum.append(at)
            if accum:
                patch['breaks'] = accum
                applied_any = True
        # Total patients if not set
        if 'patient_count' not in patch:
            tp = defaults.get('total_patients')
            if tp not in (None, ''):
                try:
                    patch['patient_count'] = int(tp)
                    applied_any = True
                except Exception:
                    pass
            else:
                # compute from before/after if both available
                bp = patch.get('before_break_opd_patients')
                ap = patch.get('after_break_opd_patients')
                if isinstance(bp, int) and isinstance(ap, int):
                    patch['patient_count'] = bp + ap
                    applied_any = True
        if applied_any and patch.get('status') not in ('OFF_DUTY', 'ON_CALL'):
            patch['status'] = 'ON_DUTY'

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_vals = [_cell_str(v).lower() for v in row]
        # Check if this row contains key headers
        if any('doctor' in val or 'physician' in val or 'dr.' in val for val in row_vals):
            # This looks like a header row
            header_row_index = i
            print(f"DEBUG: Found header row at index {i}")
            for idx, h in enumerate(row):
                key = norm_key(_cell_str(h))
                if key:
                    headers[key] = idx
            print(f"DEBUG: Headers found: {headers}")
            break
    
    # If we found headers, parse data rows
    if headers and 'doctor' in headers:
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # Skip until after header row
            if header_row_index and i <= header_row_index:
                continue
            
            def col(name):
                idx = headers.get(name)
                if idx is None: return None
                try: return row[idx]
                except: return None
            
            date_iso = _parse_date(col('date'))
            doc_name = _cell_str(col('doctor'))
            
            if not date_iso or not doc_name:
                continue
            
            date_range.add(date_iso)
            if doc_name not in doctor_stats:
                doctor_stats[doc_name] = 0
            doctor_stats[doc_name] += 1
            total_rows += 1

            # Determine row state
            st_txt = _cell_str(col('start time'))
            bb_t = _cell_str(col('before break opd timing'))
            ab_t = _cell_str(col('after break opd timing'))
            status_val = (_cell_str(col('status')) or '').strip().upper()
            if status_val == 'OFF':
                state = 'OFF'
            elif st_txt or bb_t or ab_t:
                state = 'ADDED'
            else:
                state = 'EMPTY'
            match = _match_doctor(doc_name, doc_list)
            if match:
                dd = per_doc_date.setdefault(match.get('id') or match.get('name'), {})
                dd[date_iso] = state
    
    # If the simple column-based parse found nothing, try "grouped" template parsing
    if not total_rows:
        print("DEBUG: No rows found in column-based parse, trying grouped template...")
        current_date = None
        header_map = None

        def build_header_map(row_vals):
            m = {}
            for idx, val in enumerate(row_vals):
                k = norm_key(_cell_str(val))
                if k:
                    m[k] = idx
            return m if ('doctor' in m) else None

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_vals = [_cell_str(v) for v in row]
            if not any(row_vals):
                continue
            
            # First check if this is a header row before checking for dates
            if not header_map:
                hm = build_header_map(row_vals)
                if hm:
                    header_map = hm
                    print(f"DEBUG: Found header map at line {i}: {header_map}")
                    continue
            
            # If we have a header_map, try to process this as a data row first
            if current_date and header_map and len(row_vals) > header_map.get('doctor', -1):
                doc_name = row_vals[header_map['doctor']] if header_map.get('doctor', -1) < len(row_vals) else None
                doc_name = _cell_str(doc_name)
                
                # If there's a doctor name, this is a data row
                if doc_name and len(doc_name) > 0:
                    # Extract date from this row if 'date' column exists
                    row_date = None
                    if 'date' in header_map and header_map['date'] < len(row_vals):
                        row_date = _parse_date(row_vals[header_map['date']])
                    
                    # Use row date if available, otherwise use current_date
                    use_date = row_date if row_date else current_date
                    
                    if use_date:
                        total_rows += 1
                        date_range.add(use_date)
                        doctor_stats[doc_name] = doctor_stats.get(doc_name, 0) + 1
                        print(f"DEBUG: Data row {i}: doctor={doc_name}, date={use_date}")

                        # Determine state using available columns
                        def get_col(k):
                            j = header_map.get(k); return row_vals[j] if (j is not None and j < len(row_vals)) else ''
                        st_txt = _cell_str(get_col('start time'))
                        before_t = _cell_str(get_col('before break opd timing'))
                        after_t = _cell_str(get_col('after break opd timing'))
                        status_val = (_cell_str(get_col('status')) or '').strip().upper()
                        if status_val == 'OFF':
                            state = 'OFF'
                        elif st_txt or before_t or after_t:
                            state = 'ADDED'
                        else:
                            state = 'EMPTY'
                        match = _match_doctor(doc_name, doc_list)
                        if match:
                            dd = per_doc_date.setdefault(match.get('id') or match.get('name'), {})
                            dd[use_date] = state
                    continue
            
            # Detect a date declaration row (standalone date rows)
            # These typically have few non-empty cells and no doctor name
            non_empty = sum(1 for v in row_vals if v)
            det_date = None
            
            # Only treat as date declaration if it has <= 3 non-empty cells
            if non_empty <= 3:
                for v in row_vals:
                    iso = _parse_date(v)
                    if iso:
                        det_date = iso
                        break
                if det_date:
                    current_date = det_date
                    print(f"DEBUG: Found date declaration row at line {i}: {current_date}")
                    header_map = None  # Reset header to look for new one
                    continue
        print(f"DEBUG: Grouped template found {total_rows} rows")
        if not total_rows:
            print("ERROR: No valid schedule data found in file after both parsing methods")
            return jsonify({'ok': False, 'error': 'No valid schedule data found in file'}), 400
    
    sorted_dates = sorted(list(date_range))
    date_range_str = f"{sorted_dates[0]} to {sorted_dates[-1]}" if sorted_dates else "Unknown"
    # Build doctor matrix for UI
    doctor_matrix = []
    if sorted_dates:
        for d in doc_list:
            key = d.get('id') or d.get('name')
            per = per_doc_date.get(key, {})
            states = [ per.get(dt, 'MISSING') for dt in sorted_dates ]
            doctor_matrix.append({'name': d.get('name'), 'specialty': d.get('specialty'), 'states': states})
    
    return jsonify({
        'ok': True,
        'filename': filename,
        'total_schedules': total_rows,
        'unique_doctors': len(doctor_stats),
        'date_range': date_range_str,
        'dates_count': len(date_range),
        'dates': sorted_dates,
        'doctor_matrix': doctor_matrix,
        'doctor_breakdown': [
            {'doctor': name, 'schedules': count}
            for name, count in sorted(doctor_stats.items(), key=lambda x: x[1], reverse=True)
        ]
    })

@app.post('/api/schedule/apply_excel')
def api_schedule_apply_excel():
    """Stage 2b: Actually process and apply the Excel file to database."""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    
    filename = request.json.get('filename')
    if not filename:
        return jsonify({'ok': False, 'error': 'filename required'}), 400
    schedule_root = os.path.join('uploads', 'schedules')
    os.makedirs(schedule_root, exist_ok=True)
    fname_norm = filename.replace('\\','/')
    filepath = os.path.normpath(os.path.join(schedule_root, fname_norm))
    if not os.path.abspath(filepath).startswith(os.path.abspath(schedule_root)):
        return jsonify({'ok': False, 'error': 'invalid path'}), 400
    if not os.path.exists(filepath):
        return jsonify({'ok': False, 'error': 'File not found'}), 404
    
    try:
        from openpyxl import load_workbook
        from datetime import datetime, date
    except Exception:
        return jsonify({'ok': False, 'error': 'openpyxl not installed'}), 500
    
    def _cell_str(v):
        if v is None: return ''
        if isinstance(v, (int, float)): return str(v)
        return str(v).strip()
    
    def _parse_date(v):
        import re
        if v is None: return None
        if isinstance(v, datetime): return v.date().isoformat()
        if isinstance(v, date): return v.isoformat()
        s = _cell_str(v)
        if not s: return None
        try:
            return datetime.strptime(s, '%Y-%m-%d').date().isoformat()
        except:
            pass
        # Search for dates embedded in strings
        m = re.search(r'(\d{4})[./-](\d{1,2})[./-](\d{1,2})', s)
        if m:
            y, mo, d = map(int, m.groups())
            return f"{y:04d}-{mo:02d}-{d:02d}"
        m = re.search(r'(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})', s)
        if m:
            d, mo, y = m.groups()
            y = int(y)
            if y < 100: y += 2000
            return f"{y:04d}-{int(mo):02d}-{int(d):02d}"
        return None
    
    def _parse_breaks(v):
        s = _cell_str(v)
        if not s: return None
        items = [p.strip() for p in s.split(',') if p.strip()]
        return items or None
    def _extract_time_and_count(v):
        s = _cell_str(v)
        if not s:
            return (None, None)
        m = re.search(r'(\d{1,2}:\d{2}(?::\d{2})?)\s*-\s*(\d{1,2}:\d{2}(?::\d{2})?)', s)
        time_txt = None
        rest = s
        if m:
            a, b = m.group(1), m.group(2)
            def norm(t):
                mm = re.match(r'^(\d{1,2}):(\d{2})', t)
                return f"{int(mm.group(1)):02d}:{mm.group(2)}" if mm else t
            time_txt = f"{norm(a)}-{norm(b)}"
            rest = (s[:m.start()] + ' ' + s[m.end():]).strip()
        m2 = re.findall(r'\b\d{1,3}\b', rest)
        count = int(m2[-1]) if m2 else None
        return (time_txt, count)
    
    def _match_doctor(name: str, doctors: list):
        toks = set(name_tokens(name))
        if not toks: return None
        best = None; score = 0
        for d in doctors:
            dt = set(name_tokens(d.get('name','')))
            if dt and dt.issubset(toks) and len(dt) > score:
                best = d; score = len(dt)
        return best
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Failed to read file: {str(e)}'}), 400
    
    # Process file and apply schedules
    headers = {}
    first = True
    applied = 0
    errors = []
    data = load_data()
    doc_list = data.get('doctors', [])
    
    def norm_key(k: str) -> str:
        k = (k or '').strip().lower()
        aliases = {
            'doctor name': 'doctor', 'name': 'doctor', 'dr': 'doctor', 'dr.': 'doctor',
            'physician': 'doctor', 'doctor/physician': 'doctor',
            'starting time': 'start time', 'start': 'start time', 'starttime': 'start time',
            'room no': 'room', 'room number': 'room',
            'total no of patients': 'total patients', 'total patients': 'total patients', 'total': 'total patients',
            # Support both old combined format and new separate columns
            'before break opd patients (time/pts)': 'before break opd patients combined',
            'before break opd timing': 'before break opd timing',
            'before break opd patients': 'before break opd patients',
            'before break patients': 'before break opd patients',
            'after break opd patients (time/pts)': 'after break opd patients combined',
            'after break opd timing': 'after break opd timing',
            'after break opd patients': 'after break opd patients',
            'after break patients': 'after break opd patients',
            # New modern template columns
            'shift': 'shift',
            'specialty': 'specialty', 'speciality': 'specialty', 'department': 'specialty', 'dept': 'specialty',
        }
        return aliases.get(k, k)

    # Load shift knowledge and helpers (used to fill missing fields when Shift is selected)
    shift_knowledge = {}
    try:
        shift_knowledge = load_shift_knowledge()
    except Exception:
        shift_knowledge = {}

    def _find_specialty_match(spec_val: str) -> str:
        if not spec_val:
            return None
        s = spec_val.strip().lower()
        for k in shift_knowledge.keys():
            if k.strip().lower() == s:
                return k
        return spec_val

    def _apply_shift_defaults(patch: dict, spec: str, shift_name: str):
        if not shift_name:
            return
        s_key = _find_specialty_match(spec or '') if spec else None
        if not s_key:
            return
        spec_data = shift_knowledge.get(s_key) or {}
        sh = (shift_name or '').strip() or ''
        defaults = spec_data.get(sh)
        if not isinstance(defaults, dict):
            return
        applied_any = False
        if 'start_time' not in patch and defaults.get('start'):
            patch['start_time'] = defaults.get('start')
            applied_any = True
        if 'before_break_opd_patients' not in patch and defaults.get('before_patients') not in (None, ''):
            try:
                patch['before_break_opd_patients'] = int(defaults.get('before_patients'))
            except Exception:
                pass
            else:
                applied_any = True
        if 'after_break_opd_patients' not in patch and defaults.get('after_patients') not in (None, ''):
            try:
                patch['after_break_opd_patients'] = int(defaults.get('after_patients'))
            except Exception:
                pass
            else:
                applied_any = True
        if 'breaks' not in patch:
            accum = []
            bt = (defaults.get('before_timing') or '').strip()
            if bt:
                accum.append(bt)
            brk = (defaults.get('breaks') or '').strip()
            if brk:
                accum.append(brk)
            at = (defaults.get('after_timing') or '').strip()
            if at:
                accum.append(at)
            if accum:
                patch['breaks'] = accum
                applied_any = True
        if 'patient_count' not in patch:
            tp = defaults.get('total_patients')
            if tp not in (None, ''):
                try:
                    patch['patient_count'] = int(tp)
                    applied_any = True
                except Exception:
                    pass
            else:
                bp = patch.get('before_break_opd_patients')
                ap = patch.get('after_break_opd_patients')
                if isinstance(bp, int) and isinstance(ap, int):
                    patch['patient_count'] = bp + ap
                    applied_any = True
        if applied_any and patch.get('status') not in ('OFF_DUTY', 'ON_CALL'):
            patch['status'] = 'ON_DUTY'

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if first:
            first = False
            for idx, h in enumerate(row):
                key = norm_key(_cell_str(h))
                headers[key] = idx
            continue
        
        def col(name):
            idx = headers.get(name)
            if idx is None: return None
            try: return row[idx]
            except: return None
        
        date_iso = _parse_date(col('date'))
        doc_name = _cell_str(col('doctor'))
        
        if not date_iso or not doc_name:
            continue
        
        doc = _match_doctor(doc_name, doc_list)
        if not doc:
            errors.append(f"Row {i}: Doctor '{doc_name}' not found")
            continue
        
        patch = {}
        st = _cell_str(col('start time'))
        if st: patch['start_time'] = st
        rm = _cell_str(col('room'))
        if rm: patch['room'] = rm
        tp = _cell_str(col('total patients'))
        if tp:
            try: patch['patient_count'] = int(float(tp))
            except: pass
        breaks_accum = []
        bb_raw = _cell_str(col('before break opd patients')) or _cell_str(col('before break opd patients (time/pts)'))
        if bb_raw:
            rng, cnt = _extract_time_and_count(bb_raw)
            if cnt is not None:
                try: patch['before_break_opd_patients'] = int(float(cnt))
                except: pass
            if rng: breaks_accum.append(rng)
        br = _parse_breaks(col('breaks'))
        if br:
            if any(re.search(r'no\s*break', b, re.I) for b in br):
                patch['breaks'] = ['NO BREAK']; breaks_accum = []
            else:
                patch['breaks'] = br
        ab_raw = _cell_str(col('after break opd patients')) or _cell_str(col('after break opd patients (time/pts)'))
        if ab_raw:
            rng, cnt = _extract_time_and_count(ab_raw)
            if cnt is not None:
                try: patch['after_break_opd_patients'] = int(float(cnt))
                except: pass
            if rng: breaks_accum.append(rng)
        if breaks_accum and patch.get('breaks') != ['NO BREAK']:
            existing = set(patch.get('breaks') or [])
            for r in breaks_accum:
                if r not in existing:
                    existing.add(r)
            if existing:
                patch['breaks'] = list(existing)

        # Apply shift defaults if provided and fields are missing
        shift_val = _cell_str(col('shift'))
        spec_val = _cell_str(col('specialty')) or (doc.get('specialty') if isinstance(doc, dict) else None)
        if shift_val:
            _apply_shift_defaults(patch, spec_val, shift_val)
        
        if not patch:
            continue
        
        if apply_single_day_update(doc, date_iso, patch):
            applied += 1
    
    # If nothing applied with simple column-based approach, try grouped template parsing
    if applied == 0:
        current_date = None
        header_map = None

        def build_header_map(row_vals):
            m = {}
            for idx, val in enumerate(row_vals):
                k = norm_key(_cell_str(val))
                if k:
                    m[k] = idx
            return m if ('doctor' in m) else None

        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_vals = [_cell_str(v) for v in row]
            if not any(row_vals):
                continue
            
            # First check if this is a header row
            if not header_map:
                hm = build_header_map(row_vals)
                if hm:
                    header_map = hm
                    print(f"DEBUG Apply: Found header map at line {i}: {header_map}")
                    continue
            
            # Try to process as data row first (if we have header_map and current_date)
            if current_date and header_map:
                d_idx = header_map.get('doctor')
                if d_idx is not None and d_idx < len(row_vals):
                    doc_name = _cell_str(row_vals[d_idx])
                    
                    # Skip section header rows (where doctor column just says "Doctor")
                    if doc_name and doc_name.lower() == 'doctor':
                        continue
                    
                    # If there's a doctor name, this is a data row - process it
                    if doc_name and len(doc_name) > 0:
                        doc = None
                        # approximate match
                        try:
                            doc = _match_doctor(doc_name, doc_list)
                        except Exception:
                            doc = None
                        if not doc:
                            errors.append(f"Row {i}: Doctor '{doc_name}' not found")
                            continue
                        
                        # Extract date from this row if 'date' column exists
                        row_date = None
                        if 'date' in header_map and header_map['date'] < len(row_vals):
                            row_date = _parse_date(row_vals[header_map['date']])
                        
                        # Use row date if available, otherwise use current_date
                        use_date = row_date if row_date else current_date
                        
                        patch = {}
                        # Helper getter
                        def get_col(key):
                            j = header_map.get(key)
                            return row_vals[j] if j is not None and j < len(row_vals) else ''
                        st = _cell_str(get_col('start time'))
                        if st:
                            patch['start_time'] = st
                            patch['status'] = 'ON_DUTY'
                        rm = _cell_str(get_col('room'))
                        if rm:
                            patch['room'] = rm
                        tp = _cell_str(get_col('total patients'))
                        if tp:
                            try:
                                patch['patient_count'] = int(float(tp))
                            except:
                                pass
                        breaks_accum = []
                        # Check for separate timing and patient columns first (new format)
                        before_timing = _cell_str(get_col('before break opd timing'))
                        before_patients = _cell_str(get_col('before break opd patients'))
                        
                        # If separate columns exist, use them
                        if before_timing or before_patients:
                            if before_patients:
                                try: 
                                    patch['before_break_opd_patients'] = int(float(before_patients))
                                except: 
                                    pass
                            if before_timing:
                                breaks_accum.append(before_timing)
                        else:
                            # Fall back to old combined format
                            bb = _cell_str(get_col('before break opd patients combined'))
                            if bb:
                                rng, cnt = _extract_time_and_count(bb)
                                if cnt is not None:
                                    try: patch['before_break_opd_patients'] = int(float(cnt))
                                    except: pass
                                if rng: breaks_accum.append(rng)
                        
                        br = _parse_breaks(get_col('breaks'))
                        if br:
                            if any(re.search(r'no\s*break', b, re.I) for b in br):
                                patch['breaks'] = ['NO BREAK']; breaks_accum = []
                            else:
                                patch['breaks'] = br
                        
                        # Check for separate after break columns (new format)
                        after_timing = _cell_str(get_col('after break opd timing'))
                        after_patients = _cell_str(get_col('after break opd patients'))
                        
                        if after_timing or after_patients:
                            if after_patients:
                                try: 
                                    patch['after_break_opd_patients'] = int(float(after_patients))
                                except: 
                                    pass
                            if after_timing:
                                breaks_accum.append(after_timing)
                        else:
                            # Fall back to old combined format
                            ab = _cell_str(get_col('after break opd patients combined'))
                            if ab:
                                rng, cnt = _extract_time_and_count(ab)
                                if cnt is not None:
                                    try: patch['after_break_opd_patients'] = int(float(cnt))
                                    except: pass
                                if rng: breaks_accum.append(rng)
                        if breaks_accum and patch.get('breaks') != ['NO BREAK']:
                            existing = set(patch.get('breaks') or [])
                            for r in breaks_accum:
                                if r not in existing:
                                    existing.add(r)
                            if existing:
                                patch['breaks'] = list(existing)
                        # Apply shift defaults if provided and fields are missing
                        shift_val = _cell_str(get_col('shift'))
                        spec_val = _cell_str(get_col('specialty')) or (doc.get('specialty') if isinstance(doc, dict) else None)
                        if shift_val:
                            _apply_shift_defaults(patch, spec_val, shift_val)
                        status_val = _cell_str(get_col('status')).upper()
                        if status_val:
                            if status_val in ('ON/CALL','ON CALL','ONCALL','ON-CALL'):
                                patch['status'] = 'ON_CALL'
                            elif status_val in ('POST ON/CALL','POST ONCALL','POST ON CALL','POST-ONCALL','POST-ON-CALL'):
                                patch['status'] = 'ON_DUTY'
                                patch['post_oncall'] = True
                            elif status_val == 'OFF':
                                patch['status'] = 'OFF_DUTY'
                                patch['start_time'] = None
                                patch['room'] = None
                                patch['patient_count'] = None
                                patch['before_break_opd_patients'] = None
                                patch['after_break_opd_patients'] = None
                                patch['breaks'] = None
                        if patch:
                            if apply_single_day_update(doc, use_date, patch):
                                applied += 1
                                print(f"DEBUG Apply: Applied schedule for {doc_name} on {use_date}")
                        continue  # Processed as data row, skip date detection
            
            # Detect a date declaration row (only if not a data row)
            non_empty = sum(1 for v in row_vals if v)
            det_date = None
            
            # Only treat as date declaration if it has <= 3 non-empty cells
            if non_empty <= 3:
                for v in row_vals:
                    iso = _parse_date(v)
                    if iso:
                        det_date = iso
                        break
                if det_date:
                    current_date = det_date
                    print(f"DEBUG Apply: Found date declaration row at line {i}: {current_date}")
                    header_map = None  # Reset header to look for new one
                    continue

    if applied:
        save_data(data)
        try:
            broker.publish_event('doctor_update', {'bulk': True})
        except:
            pass
    
    return jsonify({
        'ok': True,
        'applied': applied,
        'errors': errors,
        'message': f'Successfully applied {applied} schedule updates!'
    })

# -------------- Simple File Explorer APIs (under uploads/schedules) --------------

def _sched_root():
    base = os.path.join('uploads', 'schedules')
    os.makedirs(base, exist_ok=True)
    return os.path.abspath(base)

def _safe_under_root(relpath: str) -> str:
    rel = (relpath or '').replace('\\','/').strip().strip('/')
    target = os.path.normpath(os.path.join(_sched_root(), rel))
    root = _sched_root()
    if not os.path.abspath(target).startswith(root):
        raise ValueError('invalid path')
    return target

# ---------------- Shift Knowledge Management APIs -----------------
def _shift_knowledge_path():
    return os.path.join('data','shift_knowledge.json')

def load_shift_knowledge():
    path = _shift_knowledge_path()
    try:
        with open(path,'r',encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def save_shift_knowledge(data_obj):
    path = _shift_knowledge_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path,'w',encoding='utf-8') as f:
        json.dump(data_obj, f, indent=2, ensure_ascii=False)

@app.get('/api/shift_knowledge')
def api_get_shift_knowledge():
    if not session.get('user'):
        return jsonify({'ok': False, 'error':'auth required'}), 401
    return jsonify({'ok': True, 'data': load_shift_knowledge()})

@app.post('/api/shift_knowledge')
def api_update_shift_knowledge():
    if not session.get('user'):
        return jsonify({'ok': False, 'error':'auth required'}), 401
    try:
        payload = request.json or {}
    except Exception:
        return jsonify({'ok': False, 'error':'invalid json'}), 400
    data_in = payload.get('data')
    if not isinstance(data_in, dict):
        return jsonify({'ok': False, 'error':'data must be object'}), 400
    # Basic validation of structure { specialty: { 'Shift 1': {...}, 'Shift 2': {...} } }
    allowed_keys = {'start','before_timing','before_patients','breaks','after_timing','after_patients','total_patients'}
    for spec, shifts in data_in.items():
        if not isinstance(shifts, dict):
            return jsonify({'ok': False, 'error': f'specialty {spec} must map to object'}), 400
        for sh_name, vals in shifts.items():
            if not isinstance(vals, dict):
                return jsonify({'ok': False, 'error': f'shift {sh_name} for {spec} must be object'}), 400
            for k in list(vals.keys()):
                if k not in allowed_keys:
                    vals.pop(k)
    save_shift_knowledge(data_in)
    return jsonify({'ok': True, 'saved': True})

@app.get('/api/files/list')
def api_files_list():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    rel = request.args.get('path','').strip()
    try:
        base = _safe_under_root(rel)
    except Exception:
        base = _sched_root()
        rel = ''
    items = []
    try:
        for name in os.listdir(base):
            p = os.path.join(base, name)
            try:
                st = os.stat(p)
                is_dir = os.path.isdir(p)
                items.append({
                    'name': name,
                    'type': 'dir' if is_dir else 'file',
                    'size': 0 if is_dir else st.st_size,
                    'modified': st.st_mtime
                })
            except Exception:
                pass
        # sort: dirs first, then files alpha
        items.sort(key=lambda x: (x['type']!='dir', x['name'].lower()))
        return jsonify({'ok': True, 'path': rel, 'items': items})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/files/mkdir')
def api_files_mkdir():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = request.get_json(silent=True) or {}
    rel = (data.get('path') or '').strip()
    if not rel:
        return jsonify({'ok': False, 'error': 'path required'}), 400
    try:
        p = _safe_under_root(rel)
        os.makedirs(p, exist_ok=True)
        return jsonify({'ok': True, 'created': rel})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.post('/api/files/delete')
def api_files_delete():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = request.get_json(silent=True) or {}
    rel = (data.get('path') or '').strip()
    recursive = bool(data.get('recursive'))
    if not rel:
        return jsonify({'ok': False, 'error': 'path required'}), 400
    try:
        p = _safe_under_root(rel)
        if os.path.isdir(p):
            if recursive:
                import shutil
                shutil.rmtree(p)
            else:
                os.rmdir(p)
        elif os.path.isfile(p):
            os.remove(p)
        else:
            return jsonify({'ok': False, 'error': 'not found'}), 404
        return jsonify({'ok': True, 'deleted': rel})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.post('/api/files/rename')
def api_files_rename():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = request.get_json(silent=True) or {}
    rel = (data.get('path') or '').strip()
    new_name = (data.get('new_name') or '').strip()
    if not rel or not new_name:
        return jsonify({'ok': False, 'error': 'path and new_name required'}), 400
    try:
        src = _safe_under_root(rel)
        parent = os.path.dirname(src)
        dst = _safe_under_root(os.path.relpath(os.path.join(parent, new_name), _sched_root()))
        os.rename(src, dst)
        return jsonify({'ok': True, 'renamed': rel, 'to': os.path.relpath(dst, _sched_root()).replace('\\','/')})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.post('/api/files/move')
def api_files_move():
    """Move a file or folder to a different location (for drag-and-drop)"""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = request.get_json(silent=True) or {}
    source_path = (data.get('source') or '').strip()
    dest_folder = (data.get('destination') or '').strip()
    if not source_path:
        return jsonify({'ok': False, 'error': 'source path required'}), 400
    try:
        # Get absolute paths
        src = _safe_under_root(source_path)
        if not os.path.exists(src):
            return jsonify({'ok': False, 'error': 'source not found'}), 404
        
        # Destination folder (empty string means root)
        if dest_folder:
            dest_dir = _safe_under_root(dest_folder)
        else:
            dest_dir = _sched_root()
        
        if not os.path.isdir(dest_dir):
            return jsonify({'ok': False, 'error': 'destination is not a folder'}), 400
        
        # Get the item name
        item_name = os.path.basename(src)
        dst = os.path.join(dest_dir, item_name)
        
        # Check if destination already exists
        if os.path.exists(dst):
            return jsonify({'ok': False, 'error': f'{item_name} already exists in destination'}), 400
        
        # Move the file/folder
        import shutil
        shutil.move(src, dst)
        
        # Return the new relative path
        new_path = os.path.relpath(dst, _sched_root()).replace('\\', '/')
        return jsonify({'ok': True, 'moved': source_path, 'to': new_path})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.get('/api/files/download')
def api_files_download():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    rel = request.args.get('path','').strip()
    if not rel:
        return jsonify({'ok': False, 'error': 'path required'}), 400
    try:
        abspath = _safe_under_root(rel)
        if not os.path.isfile(abspath):
            return jsonify({'ok': False, 'error': 'file not found'}), 404
        directory = os.path.dirname(abspath)
        fname = os.path.basename(abspath)
        return send_from_directory(directory, fname, as_attachment=True)
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.get('/api/telegram/status')
def api_telegram_status():
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    return jsonify(_telegram_status())

@app.post('/api/telegram/restart')
def api_telegram_restart():
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    global _telegram_running, _telegram_thread
    with _telegram_lock:
        if _telegram_thread and _telegram_thread.is_alive():
            _telegram_running = False  # signal old loop to exit
        # start fresh
        _telegram_running = True
        _telegram_thread = threading.Thread(target=_telegram_loop, name="telegram-poll", daemon=True)
        _telegram_thread.start()
    return jsonify({'ok': True, 'restarted': True, 'status': _telegram_status()})


def _send_telegram_message(chat_id: str, text: str, parse_mode: str = None):
    """Send a message to a Telegram chat."""
    if not TELEGRAM_TOKEN:
        return False
    try:
        base = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
        payload = {
            'chat_id': chat_id,
            'text': text
        }
        if parse_mode:
            payload['parse_mode'] = parse_mode
        
        data = urlencode(payload).encode('utf-8')
        req = Request(f"{base}/sendMessage", data=data, headers={'Content-Type': 'application/x-www-form-urlencoded'})
        with urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read().decode('utf-8'))
            return result.get('ok', False)
    except Exception as e:
        log(f"send message error: {e}")
        return False


def _process_bot_command(text: str, chat_id: str) -> bool:
    """Process bot commands like /start, /help, /format, /status"""
    global _processed_commands
    import time
    
    cmd = text.strip().lower()
    current_time = time.time()
    
    # Deduplication: Check if this exact command was processed in last 5 seconds
    if chat_id in _processed_commands:
        if cmd in _processed_commands[chat_id]:
            last_time = _processed_commands[chat_id][cmd]
            if current_time - last_time < 5:  # 5 second window
                log(f"Ignoring duplicate command {cmd} from {chat_id}")
                return True  # Already processed, don't process again
    else:
        _processed_commands[chat_id] = {}
    
    # Mark this command as processed
    _processed_commands[chat_id][cmd] = current_time
    
    # Clean up old entries (older than 10 seconds)
    for cid in list(_processed_commands.keys()):
        _processed_commands[cid] = {c: t for c, t in _processed_commands[cid].items() if current_time - t < 10}
        if not _processed_commands[cid]:
            del _processed_commands[cid]
    
    if cmd == '/start' or cmd == '/help':
        help_text = """╔══════════════════════════════════════╗
║  🏥 HULHUMALE HOSPITAL BOT          ║
║     Schedule Management System       ║
╚══════════════════════════════════════╝

👋 *Welcome!*
I help you manage doctor schedules efficiently.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⚡ *AVAILABLE COMMANDS*

📋 /format
   → View complete formatting guide
   → Copy-paste templates
   → See examples

📊 /status
   → Check current date
   → Verify bot status
   → System information

❓ /help
   → Show this help menu
   → Quick command reference

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🎯 *HOW TO USE*

1️⃣ Type /format to get templates
2️⃣ Copy the format you need
3️⃣ Fill in the doctor details
4️⃣ Send the message

*Important:* Always include the date!

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 *TIP:* Start with /format command
to see all available templates and examples.

╚══════════════════════════════════════╝"""
        _send_telegram_message(chat_id, help_text, parse_mode='Markdown')
        log(f"Sent help to {chat_id}")
        return True
    
    elif cmd == '/format':
        format_text = """╔══════════════════════════════════╗
║   📋 MESSAGE FORMAT GUIDE        ║
╚══════════════════════════════════╝

🎯 *DOCTOR DUTY SCHEDULE FORMAT*
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

```
Date:
Name:
Starting time:
Room:
Total no of patients:
Before break OPD:
Before break patients:
Break:
After break OPD:
After break patients:
```

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✨ *EXAMPLE - Complete Duty*
```
Date: 31/10/2025
Name: Dr. Moosa Manik
Starting time: 8:00
Room: 4
Total no of patients: 20
Before break OPD: [8:00 TO 11:00]
Before break patients: [10] pts
Break: [11:00-12:00]
After break OPD: [12:00 TO 14:00]
After break patients: [10] pts
```

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🏥 *LEAVE/SICK STATUS FORMAT*
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

```
Start date:
End date:
Name:
Leave type:
Reason:
```

✨ *EXAMPLE - Leave Status*
```
Start date: 01/11/2025
End date: 05/11/2025
Name: Dr. Ahmed
Leave type: Sick Leave
Reason: Medical leave
```

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 *IMPORTANT TIPS*

✓ *Date is MANDATORY* - Always include date
✓ Starting time formats: 15:00 OR 1500 OR 15
✓ Copy the format above, then fill details
✓ Be accurate with doctor names
✓ Include patient counts for better tracking

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📌 *Quick Reference:*
• For duty updates: Use full format above
• For leave: Use leave format above
• Questions? Use /help command

╚══════════════════════════════════╝"""
        _send_telegram_message(chat_id, format_text, parse_mode='Markdown')
        log(f"Sent format guide to {chat_id}")
        return True
    
    elif cmd == '/status':
        today = get_internet_today_iso()
        # Format date nicely
        try:
            from datetime import datetime
            date_obj = datetime.strptime(today, '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d/%m/%Y')
            day_name = date_obj.strftime('%A')
        except:
            formatted_date = today
            day_name = ''
        
        # Build day line separately to avoid nested f-strings
        day_line = f'\nDay: *{day_name}*' if day_name else ''
        
        status_text = f"""╔══════════════════════════════════╗
║      🤖 BOT STATUS REPORT        ║
╚══════════════════════════════════╝

✅ *SYSTEM STATUS*
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🟢 Bot Status: *ACTIVE*
⚡ Response: *ONLINE*
� Processing: *READY*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📅 *DATE INFORMATION*

Today: *{formatted_date}*{day_line}
ISO Format: `{today}`
Time Zone: *UTC+5 (Maldives)*

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

💡 *QUICK ACTIONS*

• Type /format for templates
• Type /help for command list
• Send updates anytime!

╚══════════════════════════════════╝"""
        _send_telegram_message(chat_id, status_text, parse_mode='Markdown')
        log(f"Sent status to {chat_id}")
        return True
    
    return False


def process_telegram_message(text: str, chat_id: str):
    # --- Check for bot commands first ---
    if text.strip().startswith('/'):
        if _process_bot_command(text.strip(), chat_id):
            return
    
    # --- Enhanced: multi-line, rich, and shift style parsing ---
    # Try multi-line rich parsing first
    if text:
        try:
            if _process_rich_multiline_message(text):
                return
        except Exception as e:
            log(f"rich multiline parse error: {e}")
        try:
            if _process_shift_style_message(text):
                return
        except Exception as e:
            log(f"shift style parse error: {e}")
    t_low = (text or '').lower()
    log(f"msg {chat_id}: {text.replace(chr(10),' ')[:160]}")
    # Closure detection
    for pat in CLOSURE_PATTERNS:
        if pat.search(t_low):
            date_iso = _extract_for_date_or_today(text)
            _apply_closure(date_iso, reason=text.strip())
            return
    # Single-line status (doctor leave / sick / opd cancelled) with optional date
    try:
        if _process_status_only_message(text):
            return
    except Exception as e:
        log(f"status-only parse error: {e}")
    # Fallback: single line parsing
    g = globals()
    parse_status_fn = g.get('parse_status_from_text')
    parse_rich_fn = g.get('parse_rich_fields')
    get_today_fn = g.get('get_internet_today_iso')
    parse_time_fn = g.get('parse_time_from_text')
    load_data_fn = g.get('load_data')
    apply_day_fn = g.get('apply_single_day_update')
    save_data_fn = g.get('save_data')
    broker_obj = g.get('broker')

    status = parse_status_fn(t_low) if parse_status_fn else None
    rich = parse_rich_fn(text, status) if parse_rich_fn else {}
    date_iso = (rich.get('for_date') if rich else None) or (get_today_fn() if get_today_fn else datetime.utcnow().date().isoformat())
    doc = _match_doctor_by_text(text)
    if not doc:
        return
    patch = {}
    if status:
        patch['status'] = status
    for key in ('status_reason','patient_count','breaks','opd','room','designation'):
        if rich and key in rich:
            patch[key] = rich[key]
    t_start = parse_time_fn(text) if parse_time_fn else None
    if t_start and 'opd' not in patch:
        patch['opd'] = [{'range': f"{t_start}-{t_start}"}]
    if not patch or not load_data_fn or not apply_day_fn:
        return
    data = load_data_fn()
    for d in data.get('doctors', []):
        if d.get('id') == doc.get('id'):
            doc = d; break
    if apply_day_fn(doc, date_iso, patch):
        if save_data_fn:
            save_data_fn(data)
        if broker_obj:
            broker_obj.publish_event('doctor_update', {'doctor_id': doc.get('id'), 'date': date_iso})
        log(f"updated {doc.get('name')} {date_iso}: {list(patch.keys())}")
def _process_rich_multiline_message(text: str) -> bool:
    """Parse rich multi-line Telegram messages for schedule updates.
    Enhanced: flexible time parsing (e.g. '80' -> 08:00) & leave/cancel detection.
    """
    import re
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    if not lines or len(lines) < 2:
        return False
    # Extract date or date range
    date_iso = None
    start_date = None
    end_date = None
    
    # Check for start_date and end_date fields (for leave/sick ranges)
    start_line = next((l for l in lines if l.lower().startswith('start date:')), None)
    end_line = next((l for l in lines if l.lower().startswith('end date:')), None)
    
    if start_line and end_line:
        # Parse start date
        m = re.search(r'(\d{1,2}[\/\-]\d{1,2}[\/\-]20\d{2})', start_line)
        if m:
            parts = re.split(r'[\/\-]', m.group(0))
            if len(parts)==3:
                d2,mn,y = parts
                start_date = f"{y}-{int(mn):02d}-{int(d2):02d}"
        
        # Parse end date
        m = re.search(r'(\d{1,2}[\/\-]\d{1,2}[\/\-]20\d{2})', end_line)
        if m:
            parts = re.split(r'[\/\-]', m.group(0))
            if len(parts)==3:
                d2,mn,y = parts
                end_date = f"{y}-{int(mn):02d}-{int(d2):02d}"
    
    # Fall back to single date field if no range
    if not start_date:
        date_lines = [l for l in lines if l.lower().startswith('date:')] or lines
        for dl in date_lines:
            m = re.search(r'(\d{1,2}[\/\-]\d{1,2}[\/\-]20\d{2}|20\d{2}-\d{2}-\d{2})', dl)
            if m:
                parts = re.split(r'[\/\-]', m.group(0))
                if len(parts)==3:
                    if len(parts[0])==4: date_iso = m.group(0)
                    else:
                        d2,mn,y = parts; date_iso = f"{y}-{int(mn):02d}-{int(d2):02d}"
                break
            hd = parse_human_date(dl)
            if hd: date_iso = hd; break
        if not date_iso:
            date_iso = get_internet_today_iso()
    # Doctor name
    name_line = next((l for l in lines if l.lower().startswith('name:')), None)
    doc_name = name_line.split(':',1)[1].strip() if name_line else None
    if not doc_name:
        return False
    data = load_data()
    # Exact match first
    doc = next((d for d in data['doctors'] if d['name'].lower() == doc_name.lower()), None)
    # Fuzzy / partial token match (handles 'Dr. Asish' vs 'Dr. Asish Rajak')
    if not doc:
        name_tokens_fn = globals().get('name_tokens')
        tokens = set(name_tokens_fn(doc_name)) if name_tokens_fn else set(t.lower() for t in re.findall(r"[A-Za-z']+", doc_name) if len(t)>1)
        best = None; best_score = 0.0
        for d in data['doctors']:
            d_tokens = set(name_tokens_fn(d['name'])) if name_tokens_fn else set(t.lower() for t in re.findall(r"[A-Za-z']+", d['name']) if len(t)>1)
            if not d_tokens: continue
            inter = tokens & d_tokens
            if not inter: continue
            # score: overlap over smaller set size so 'Asish' perfectly matches first name
            score = len(inter) / max(1, min(len(tokens), len(d_tokens)))
            # Prefer higher score; tie-break shorter levenshtein distance (optional simple length diff) / more tokens
            if score > best_score:
                best_score = score; best = d
        # Accept if at least half tokens overlap or full first token match
        if best and (best_score >= 0.5 or any(best['name'].lower().startswith(t) for t in tokens)):
            doc = best
            try: log(f"fuzzy matched '{doc_name}' -> {doc.get('name')} (score {best_score:.2f})")
            except Exception: pass
    if not doc:
        try: log(f"no doctor match for '{doc_name}' in rich multiline message")
        except Exception: pass
        return False
    patch = {}
    desig_line = next((l for l in lines if l.lower().startswith('designation:')), None)
    if desig_line: patch['designation'] = desig_line.split(':',1)[1].strip()
    st_line = next((l for l in lines if 'starting time' in l.lower()), None)
    if st_line:
        # Accept HH:MM or compressed forms like 8, 80, 800, 830 -> 08:00 / 08:30
        m = re.search(r'(\d{1,2}:\d{2})', st_line)
        if m:
            patch['start_time'] = m.group(1)
        else:
            m2 = re.search(r'starting time\s*:?\s*(\d{1,4})', st_line, flags=re.I)
            if m2:
                raw = m2.group(1)
                # Normalisation rules:
                #  - '80'  -> 08:00, '90' -> 09:00 (treat second 0 as minutes)
                #  - '8'   -> 08:00
                #  - '830' -> 08:30, '930' -> 09:30
                #  - '1230'-> 12:30, '945' -> 09:45
                hh = mm = 0
                if len(raw) == 1:  # '8'
                    hh = int(raw); mm = 0
                elif len(raw) == 2:
                    if raw[1] == '0' and raw[0].isdigit():  # '80' -> 8:00
                        hh = int(raw[0]); mm = 0
                    else:  # '12' -> 12:00
                        hh = int(raw); mm = 0
                elif len(raw) == 3:  # '830' -> 8:30, '945' -> 9:45
                    hh = int(raw[0]); mm = int(raw[1:])
                else:  # len 4: '1230' -> 12:30
                    hh = int(raw[:2]); mm = int(raw[2:])
                # Clamp ranges
                hh = max(0, min(23, hh))
                mm = max(0, min(59, mm))
                patch['start_time'] = f"{hh:02d}:{mm:02d}"
    # Support room lines with multi-word descriptions, e.g. 'Room: dental room 136' or 'Room dental room 136'
    room_line = next((l for l in lines if l.lower().startswith('room')), None)
    if room_line:
        raw_room = ''
        if ':' in room_line:
            raw_room = room_line.split(':',1)[1].strip()
        else:
            raw_room = room_line[4:].strip()  # remove leading 'Room'
        # Collapse multiple spaces
        raw_room = re.sub(r'\s+', ' ', raw_room)
        # Remove any leading separators
        raw_room = raw_room.lstrip('-:') .strip()
        # Accept full string (no token truncation). Trim to a sane max length.
        if raw_room:
            patch['room'] = raw_room[:60]
    pc_line = next((l for l in lines if 'no of patients' in l.lower()), None)
    if pc_line:
        m = re.search(r'(\d+)', pc_line)
        if m: patch['patient_count'] = int(m.group(1))
    opd_lines = [l for l in lines if l.lower().startswith('opd')]
    opd = []  # generic OPD list (if provided with 'OPD:' lines)
    for ol in opd_lines:
        m = re.search(r'(\d{1,2}:\d{2}-\d{1,2}:\d{2})[-\s]*(\d{1,3})?', ol)
        if m:
            rng = m.group(1); cnt = m.group(2)
            opd.append({'range': rng, **({'count': int(cnt)} if cnt else {})})

    # Capture explicit BEFORE / AFTER BREAK lines into separate arrays (robust multi-match)
    before_slots = []
    after_slots = []
    full_text = '\n'.join(lines)
    # Patterns allow optional punctuation/hyphens between label and time range, and optional trailing count tokens
    # More permissive: allow optional 'OPD' token, hyphens/colons, and flexible spacing before the time range
    # Examples matched: 'before breakOPD- 10:00-11:00- 20 pts', 'after break opd: 13:00 - 14:00 15'
    # Accept brackets [..], and either '-' or 'to' between times
    pat_before = re.compile(
        r'before\s*break[^0-9\n]{0,40}(?:opd)?[^0-9\n]{0,12}'
        r'\[?\s*(\d{1,2}\s*:?\s*\d{0,2})\s*(?:-|to)\s*(\d{1,2}\s*:?\s*\d{0,2})\s*\]?'
        r'(?:[^0-9\n]{0,12}(\d{1,3}))?\s*(?:pts|patients)?', re.I)
    pat_after = re.compile(
        r'after\s*break[^0-9\n]{0,40}(?:opd)?[^0-9\n]{0,12}'
        r'\[?\s*(\d{1,2}\s*:?\s*\d{0,2})\s*(?:-|to)\s*(\d{1,2}\s*:?\s*\d{0,2})\s*\]?'
        r'(?:[^0-9\n]{0,12}(\d{1,3}))?\s*(?:pts|patients)?', re.I)
    for m in pat_before.finditer(full_text):
        # Build normalized HH:MM-HH:MM
        a = re.sub(r'\s+', '', m.group(1))
        b = re.sub(r'\s+', '', m.group(2))
        if len(a) in (1,2): a = f"{int(a):02d}:00"
        if len(b) in (1,2): b = f"{int(b):02d}:00"
        if len(a)==3: a = f"{a[0]}:{a[1:]}"
        if len(b)==3: b = f"{b[0]}:{b[1:]}"
        rng = f"{a}-{b}"
        cnt = m.group(3)
        slot = {'range': rng}
        if cnt: 
            try: slot['count'] = int(cnt)
            except Exception: pass
        before_slots.append(slot)
    for m in pat_after.finditer(full_text):
        a = re.sub(r'\s+', '', m.group(1))
        b = re.sub(r'\s+', '', m.group(2))
        if len(a) in (1,2): a = f"{int(a):02d}:00"
        if len(b) in (1,2): b = f"{int(b):02d}:00"
        if len(a)==3: a = f"{a[0]}:{a[1:]}"
        if len(b)==3: b = f"{b[0]}:{b[1:]}"
        rng = f"{a}-{b}"
        cnt = m.group(3)
        slot = {'range': rng}
        if cnt:
            try: slot['count'] = int(cnt)
            except Exception: pass
        after_slots.append(slot)
    if before_slots:
        patch['before_break_opd'] = before_slots
        if 'before_break_opd_patients' not in patch:
            try: patch['before_break_opd_patients'] = sum(s.get('count', 0) for s in before_slots)
            except Exception: pass
    if after_slots:
        patch['after_break_opd'] = after_slots
        if 'after_break_opd_patients' not in patch:
            try: patch['after_break_opd_patients'] = sum(s.get('count', 0) for s in after_slots)
            except Exception: pass
    # If we captured explicit before/after slots, also publish a combined generic OPD list
    if (before_slots or after_slots) and 'opd' not in patch:
        try:
            patch['opd'] = []
            if before_slots:
                for s in before_slots:
                    rng = s.get('range')
                    if rng: patch['opd'].append({'range': rng, **({'count': s['count']} if 'count' in s else {})})
            if after_slots:
                for s in after_slots:
                    rng = s.get('range')
                    if rng: patch['opd'].append({'range': rng, **({'count': s['count']} if 'count' in s else {})})
        except Exception:
            pass
    if opd:
        patch['opd'] = opd
    break_lines = [l for l in lines if l.lower().startswith('break')]
    brks = []
    has_no_break = False
    for bl in break_lines:
        if re.search(r'no\s*break', bl, re.I):
            brks.append('NO BREAK')
            has_no_break = True
            continue
        m = re.search(r'\[?\s*(\d{1,2}:\d{2})\s*(?:-|to)\s*(\d{1,2}:\d{2})\s*\]?', bl, re.I)
        if m: brks.append(f"{m.group(1)}-{m.group(2)}")
    if brks: 
        patch['breaks'] = brks
    # If no break, ensure we don't split before/after - treat as single session
    if has_no_break and before_slots and not after_slots:
        # Move all before slots to after or keep as is
        pass
    # Extract before/after break patient counts if present on labeled lines
    for idx, l in enumerate(lines):
        if re.search(r'before\s*break\s*patients?', l, re.I):
            m = re.search(r'(\d+)', l)
            if m: patch['before_break_opd_patients'] = int(m.group(1))
        if re.search(r'after\s*break\s*patients?', l, re.I):
            m = re.search(r'(\d+)', l)
            if m: patch['after_break_opd_patients'] = int(m.group(1))
        # Also support generic 'No of patients' lines immediately after before/after OPD sections
        tl = l.strip().lower()
        if tl.startswith('before') and 'opd' in tl:
            # Look ahead one line for 'no of patients' or bare '10 pts'
            if idx+1 < len(lines):
                nxt = lines[idx+1]
                m1 = re.search(r'no\s*of\s*patients\s*[:\-]?\s*(\d+)', nxt, re.I)
                m2 = re.search(r'\b(\d{1,3})\s*(?:pts|patients)?\b', nxt, re.I)
                if m1 and 'before_break_opd_patients' not in patch:
                    patch['before_break_opd_patients'] = int(m1.group(1))
                elif m2 and 'before_break_opd_patients' not in patch:
                    patch['before_break_opd_patients'] = int(m2.group(1))
        if tl.startswith('after') and 'opd' in tl:
            if idx+1 < len(lines):
                nxt = lines[idx+1]
                m1 = re.search(r'no\s*of\s*patients\s*[:\-]?\s*(\d+)', nxt, re.I)
                m2 = re.search(r'\b(\d{1,3})\s*(?:pts|patients)?\b', nxt, re.I)
                if m1 and 'after_break_opd_patients' not in patch:
                    patch['after_break_opd_patients'] = int(m1.group(1))
                elif m2 and 'after_break_opd_patients' not in patch:
                    patch['after_break_opd_patients'] = int(m2.group(1))
    # Status detection (leave / sick / cancel / on call) + flags/notes
    text_lower = '\n'.join(lines).lower()
    if re.search(r'\b(sick|medical)\s+leave\b', text_lower):
        patch['status'] = 'SICK'; patch['status_reason'] = 'Medical leave'
    elif re.search(r'\bfamily\s+leave\b', text_lower):
        patch['status'] = 'LEAVE'; patch['status_reason'] = 'Family leave'
    elif re.search(r'\b(on call)\b', text_lower):
        patch['status'] = 'ON_CALL'
    elif re.search(r'opd.*(cancelled|canceled)', text_lower):
        patch['status'] = 'OFF_DUTY'; patch['status_reason'] = 'OPD cancelled'
    elif re.search(r'\bleave\b', text_lower):
        patch['status'] = 'LEAVE'
    # Post-oncall flag if mentioned anywhere (DATE line may carry '(Post-oncall)')
    if re.search(r'post[-\s]?on[-\s]?call', text_lower):
        patch['post_oncall'] = True
    # Detect explicit "no tokens"/"no after break" for AFTER BREAK section
    for l in lines:
        tl = l.strip().lower()
        if 'after' in tl and 'break' in tl:
            if re.search(r'no\s*tokens?', tl):
                patch['after_break_note'] = 'NO TOKENS'
            elif re.search(r'no\s*after\s*break', tl):
                patch['after_break_note'] = 'NO AFTER BREAK'
    # Default ON_DUTY if schedule present & no explicit inactive status
    if 'status' not in patch and (opd or patch.get('patient_count') or patch.get('room') or patch.get('start_time')):
        patch['status'] = 'ON_DUTY'
    reason_line = next((l for l in lines if l.lower().startswith('reason:')), None)
    if reason_line: patch['status_reason'] = reason_line.split(':',1)[1].strip()
    
    # Extract leave type if provided
    leave_type_line = next((l for l in lines if l.lower().startswith('leave type:')), None)
    if leave_type_line:
        leave_type = leave_type_line.split(':',1)[1].strip()
        # Update status based on leave type
        if 'sick' in leave_type.lower():
            patch['status'] = 'SICK'
            if 'status_reason' not in patch:
                patch['status_reason'] = leave_type
        else:
            if 'status_reason' not in patch:
                patch['status_reason'] = leave_type
    
    if not patch: return False
    
    # Determine which dates to apply to
    dates_to_apply = []
    if start_date and end_date:
        # Date range: expand to all dates from start to end
        dates_to_apply = _expand_date_range(start_date, end_date)
        # Add date range info to reason if status is leave/sick
        if patch.get('status') in ('LEAVE', 'SICK') and dates_to_apply:
            existing_reason = patch.get('status_reason', patch.get('status'))
            # Format dates for display using helper function
            try:
                start_fmt = _fmt_display_dmy(start_date)
                end_fmt = _fmt_display_dmy(end_date)
                patch['status_reason'] = f"{existing_reason} ({start_fmt} to {end_fmt})"
            except Exception:
                pass
    elif date_iso:
        # Single date
        dates_to_apply = [date_iso]
    
    if not dates_to_apply:
        return False
    
    # Apply update to all dates in range
    any_success = False
    for target_date in dates_to_apply:
        if apply_single_day_update(doc, target_date, patch.copy()):
            any_success = True
            # Publish SSE event for each date
            try:
                if 'broker' in globals():
                    broker.publish_event('doctor_update', {
                        'doctor_id': doc.get('id'),
                        'date': target_date,
                        'doctor': hydrate_doctor_for_date(doc, target_date)
                    })
            except Exception:
                pass
    
    if any_success:
        save_data(data)
        date_desc = f"{dates_to_apply[0]} to {dates_to_apply[-1]}" if len(dates_to_apply) > 1 else dates_to_apply[0]
        log(f"rich multiline updated {doc.get('name')} {date_desc}: {list(patch.keys())}")
        return True
    
    return False


def _match_doctor_by_text(text: str):
    import re, difflib, unicodedata
    g = globals(); name_tokens_fn = g.get('name_tokens'); load_data_fn = g.get('load_data')
    def norm(s: str) -> str:
        if not s:
            return ''
        s = unicodedata.normalize('NFKD', s)
        s = ''.join(ch for ch in s if not unicodedata.combining(ch))
        s = s.lower()
        s = re.sub(r"(?i)\b(dr|mr|mrs|ms|prof)\.?\b", " ", s)
        s = re.sub(r"[^a-z0-9\s]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        return s
    toks = set(name_tokens_fn(text)) if (name_tokens_fn and text) else set()
    msg_norm = norm(text)
    if (not toks and not msg_norm) or not load_data_fn:
        return None
    data = load_data_fn()
    best = None; best_score = 0.0
    for d in data.get('doctors', []):
        raw_name = d.get('name','') or ''
        name_norm = norm(raw_name)
        dtoks = set(name_tokens(raw_name))
        if not (name_norm or dtoks):
            continue
        overlap = toks & dtoks if toks else set()
        # Base overlap score normalized by doctor's token count (so extra words in message don't punish)
        base = (len(overlap) / max(1, len(dtoks))) if dtoks else 0.0
        # Prefix bonus using normalized name without honorifics
        prefix_bonus = 0.0
        for t in (toks or {""}):
            if t and len(t) > 2 and name_norm.startswith(t):
                prefix_bonus = 0.6
                break
        # Keyword / alias bonus
        kw_bonus = 0.0
        kws = d.get('keywords') or []
        for kw in kws:
            k = norm(str(kw))
            if not k:
                continue
            if (toks and k in toks) or (k and k in msg_norm):
                kw_bonus = 0.5
                break
        # Fuzzy ratio between normalized strings (bounded weight)
        ratio = difflib.SequenceMatcher(None, name_norm, msg_norm).ratio() if msg_norm and name_norm else 0.0
        ratio_bonus = 0.35 * ratio
        score = base + prefix_bonus + kw_bonus + ratio_bonus
        # Exact subset of doctor's tokens in message strongly boosts
        if dtoks and dtoks.issubset(toks):
            score = max(score, 1.0)
        if score > best_score:
            best = d; best_score = score
    return best


def _extract_for_date_or_today(text: str) -> str:
    g = globals(); parse_rich_fn = g.get('parse_rich_fields'); get_today_fn = g.get('get_internet_today_iso')
    rich = parse_rich_fn(text) if parse_rich_fn else {}
    return rich.get('for_date') or (get_today_fn() if get_today_fn else datetime.utcnow().date().isoformat())

# --- Simple single-line status message parser ---
def _process_status_only_message(text: str) -> bool:
    """Parse messages like:
        "Dr Asish 07/09/2025 sick leave"
        "Dr. Hatem 7 Sep OPD cancelled"
        "Dr Ahmed 2025-09-07 on call"
        "Dr Sameh leave 9 Sep"
    If a date is present it applies to that date, else today's date.
    Returns True if a doctor status was applied.
    """
    if not text or len(text) < 4:
        return False
    low = text.lower()
    # Fast check for any status keywords
    if not any(k in low for k in ('leave','sick','cancel','on call','on-call','oncall','off duty','off-duty','offduty')):
        return False
    date_iso = parse_human_date(text) or get_internet_today_iso()
    doc = _match_doctor_by_text(text)
    if not doc:
        return False
    status = None; reason = None
    # Determine status precedence
    if 'sick' in low:
        status = 'SICK'; reason = 'Medical leave'
    elif re.search(r'\b(on\s*call|on-call|oncall)\b', low):
        status = 'ON_CALL'
    elif 'cancel' in low and 'opd' in low:
        status = 'OFF_DUTY'; reason = 'OPD cancelled'
    elif re.search(r'\boff(\s+duty)?\b', low):
        status = 'OFF_DUTY'
    elif 'leave' in low:
        status = 'LEAVE'
    if not status:
        return False
    # Optional custom reason extraction (text after status keyword)
    if status in ('LEAVE','SICK'):
        m = re.search(r'(sick\s+leave|medical\s+leave|annual\s+leave|family\s+leave)', low)
        if m:
            reason = m.group(1).title()
        elif status == 'LEAVE' and not reason:
            reason = 'Leave'
        elif status == 'SICK' and not reason:
            reason = 'Medical leave'
    # Apply per-date patch (support date range)
    data = load_data()
    found = None
    for d in data.get('doctors', []):
        if d.get('id') == doc.get('id'):
            found = d; break
    if not found:
        return False
    days = [date_iso]
    rng = _extract_date_range(text)
    if rng:
        days = _expand_date_range(rng[0], rng[1])
    any_changed = False
    for di in days:
        patch = {'status': status}
        if reason:
            patch['status_reason'] = reason if len(days)==1 else f"{reason} ({_fmt_display_dmy(days[0])} till {_fmt_display_dmy(days[-1])})"
        if apply_single_day_update(found, di, patch):
            any_changed = True
            try:
                broker.publish_event('doctor_update', {'doctor_id': found.get('id'), 'date': di})
            except Exception:
                pass
    if any_changed:
        save_data(data)
        log(f"status-only updated {found.get('name')} {days[0]}..{days[-1]} -> {status}")
        return True
    return False

# Simple closure persistence (date -> list of reasons)

def _load_closures():
    if not os.path.exists(CLOSURE_PATH):
        return {}
    try:
        with open(CLOSURE_PATH,'r',encoding='utf-8') as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def _save_closures(data: dict):
    os.makedirs(os.path.dirname(CLOSURE_PATH), exist_ok=True)
    with open(CLOSURE_PATH,'w',encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _apply_closure(date_iso: str, reason: str):
    closures = _load_closures()
    reasons = closures.get(date_iso) or []
    if not isinstance(reasons, list):
        reasons = []
    if reason not in reasons:
        reasons.append(reason)
        closures[date_iso] = reasons
        _save_closures(closures)
        # publish event if broker already defined
        if 'broker' in globals():
            try:
                globals()['broker'].publish_event('closure_update', {'date': date_iso, 'reason': reason})
            except Exception:
                pass
        log(f"closure added {date_iso}")

# --- Application & Data Layer (reconstructed) ---
# (App already initialized above)
# from flask import Flask, request, jsonify, send_file, render_template  # removed duplicate import
# from io import BytesIO  # removed duplicate import
# app = Flask(__name__)
# app.secret_key = SECRET_KEY

# Simple in-memory event broker (SSE)
class EventBroker:
    def __init__(self):
        self.subscribers = []  # list of queues
        self.lock = threading.Lock()
    def subscribe(self):
        q = deque(maxlen=200)
        with self.lock:
            self.subscribers.append(q)
        return q
    def publish_event(self, event: str, payload):
        data = json.dumps(payload)
        with self.lock:
            dead = []
            for q in self.subscribers:
                try:
                    q.append((event, data))
                except Exception:
                    dead.append(q)
            for q in dead:
                try: self.subscribers.remove(q)
                except Exception: pass

broker = EventBroker()

# --- Timezone & date helpers (Maldives UTC+5) ---
TZ_OFFSET_MINUTES = 5 * 60  # UTC+5

def get_internet_today_iso(cache=[None, 0]):
    """Return today's date (local Maldives time) as YYYY-MM-DD.
    Cached ~55s to reduce recomputation."""
    now = time()
    if cache[0] is not None and now - cache[1] < 55:
        return cache[0]
    today = (datetime.utcnow() + timedelta(minutes=TZ_OFFSET_MINUTES)).date().isoformat()
    cache[0] = today; cache[1] = now
    return today

def iso_to_display_date(iso_str: str) -> str:
    try:
        dt = datetime.strptime(iso_str, '%Y-%m-%d').date()
        return f"{dt.day:02d} {dt.strftime('%b').upper()} {dt.year}"
    except Exception:
        return iso_str

# --- Per-date schedule patch helper ---
PER_DATE_FIELDS = {
    'status','status_reason','start_time','room','patient_count',
    'opd','breaks','designation',
    'before_break_opd_patients','after_break_opd_patients',
    # NEW explicit grouping arrays
    'before_break_opd','after_break_opd',
    # Additional flags/notes for display flexibility
    'post_oncall','after_break_note'
}

def apply_single_day_update(doc: dict, date_iso: str, patch: dict) -> bool:
    """Apply partial schedule patch only for given date.
    Only keys in PER_DATE_FIELDS processed. Returns True if changes applied."""
    if not doc or not date_iso or not patch:
        return False
    if 'schedule_by_date' not in doc:
        _migrate_doctor_multidate(doc)
    try:
        parsed_date = datetime.strptime(date_iso, '%Y-%m-%d')
        # BLOCK FRIDAY SCHEDULES - Friday is weekday 4
        if parsed_date.weekday() == 4:
            return False
    except Exception:
        return False
    sched_map = doc.setdefault('schedule_by_date', {})
    current = sched_map.get(date_iso, {})
    changed = False
    for k, v in patch.items():
        if k not in PER_DATE_FIELDS:
            continue
        if v in (None, ''):
            if k in current:
                del current[k]; changed = True
            continue
        if k == 'status' and isinstance(v, str):
            v = v.upper().strip()
        if k == 'opd' and isinstance(v, list):
            norm = []
            for item in v:
                if isinstance(item, dict) and 'range' in item:
                    norm.append({'range': item['range'], **{ik: iv for ik, iv in item.items() if ik != 'range'}})
                elif isinstance(item, str) and '-' in item:
                    norm.append({'range': item.strip()})
            v = norm
        if k in ('before_break_opd','after_break_opd') and isinstance(v, list):
            norm = []
            for item in v:
                if isinstance(item, dict) and 'range' in item:
                    norm.append({'range': item['range'], **{ik: iv for ik, iv in item.items() if ik != 'range'}})
                elif isinstance(item, str) and '-' in item:
                    norm.append({'range': item.strip()})
            v = norm
        if current.get(k) != v:
            current[k] = v; changed = True
    if changed:
        current['last_update'] = datetime.utcnow().isoformat()+'Z'
        sched_map[date_iso] = current
    return changed

def hydrate_doctor_for_date(doc: dict, date_iso: str):
    """Return a merged doctor dict including per-date schedule fields for a specific date.
    Includes new before/after break arrays & counts for front-end instant patch."""
    if not doc:
        return {}
    base = {k: v for k, v in doc.items() if k != 'schedule_by_date'}
    sched = {}
    try:
        sched = (doc.get('schedule_by_date') or {}).get(date_iso, {})
    except Exception:
        sched = {}
    merged = dict(base)
    for fld in ('designation','status','start_time','room','patient_count','opd','breaks','status_reason',
                'before_break_opd','after_break_opd','before_break_opd_patients','after_break_opd_patients',
                'post_oncall','after_break_note'):
        if fld in sched:
            merged[fld] = sched[fld]
    merged['for_date'] = date_iso
    return merged

# --- Shift-style multi-line message parser (placed here to exist before use) ---
SHIFT_START_TIMES = {
    '1': '08:00',
    '2': '14:00',
    '3': '20:00',
}

def _process_shift_style_message(text: str) -> bool:
    if not text:
        return False
    raw_lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not raw_lines:
        return False
    # Leading date line detection
    date_iso = None
    first_lower = raw_lines[0].lower()
    if first_lower.startswith('date') or re.search(r'\d{1,2}(st|nd|rd|th)?\s+[A-Za-z]{3,}', raw_lines[0]):
        date_iso = parse_human_date(raw_lines[0]) or None
        if not date_iso:
            mnum = re.search(r'(20\d{2}-\d{2}-\d{2})', raw_lines[0])
            if mnum: date_iso = mnum.group(1)
        if not date_iso:
            m2 = re.search(r'(\d{1,2}[\/-]\d{1,2}[\/-]20\d{2})', raw_lines[0])
            if m2:
                parts = re.split(r'[\/-]', m2.group(1))
                if len(parts)==3:
                    d_part, mn, y = parts
                    try: date_iso = f"{y}-{int(mn):02d}-{int(d_part):02d}"
                    except Exception: pass
        if date_iso:
            raw_lines = raw_lines[1:]
    if not date_iso:
        date_iso = get_internet_today_iso()
    joined_lower = '\n'.join(raw_lines).lower()
    if len(raw_lines) <= 1 and not any(k in joined_lower for k in ('shift',' off','leave','sick','on/call','on call')):
        return False
    data = load_data()
    changed = False
    for line in raw_lines:
        low = line.lower()
        status = None
        if 'on/call' in low or 'on call' in low: status = 'ON_CALL'
        elif re.search(r'\boff\b', low): status = 'OFF_DUTY'
        elif 'leave' in low: status = 'LEAVE'
        elif 'sick' in low: status = 'SICK'
        elif 'shift' in low: status = 'ON_DUTY'
        m_shift = re.search(r'shift\s*(\d)', low)
        shift_num = m_shift.group(1) if m_shift else None
        start_time = SHIFT_START_TIMES.get(shift_num) if shift_num else None
        parts = re.split(r'\b(shift|off|leave|sick|on/call|on call)\b', line, flags=re.I)
        name_part = parts[0] if parts else line
        name_part = re.sub(r'^dr[.,\s]+', '', name_part, flags=re.I).strip()
        if not name_part: continue
        tokens = set(name_tokens(name_part))
        if not tokens: continue
        match = None; score = 0
        for d in data['doctors']:
            dtoks = set(name_tokens(d.get('name','')))
            if dtoks and dtoks.issubset(tokens) and len(dtoks) > score:
                match = d; score = len(dtoks)
        if not match:
            continue
        patch = {}
        if status: patch['status'] = status
        if start_time: patch['start_time'] = start_time
        if not patch: continue
        if apply_single_day_update(match, date_iso, patch):
            changed = True
            log(f"shift-style applied {match.get('name')} {date_iso} -> {patch}")
    if changed:
        save_data(data)
    return changed

# --- Helpers added for multi-date schedule support & parsing ---
# Tokenize doctor name / text for fuzzy matching

def name_tokens(text: str):
    if not text:
        return []
    # Remove honorifics and punctuation, split on non letters
    text = re.sub(r"(?i)\b(dr|mr|mrs|ms|prof)\.?\b", " ", text)
    toks = re.findall(r"[A-Za-z']+", text.lower())
    return [t for t in toks if len(t) > 1]

# Flexible human date parser (supports '2025-01-31', '31/01/2025', '31-01-2025', '31 Jan 2025', '31st Jan', 'Jan 31', etc.)
MONTH_MAP = {m.lower(): i for i, m in enumerate(['January','February','March','April','May','June','July','August','September','October','November','December'], start=1)}
MONTH_ABBR = {m[:3].lower(): i for m, i in MONTH_MAP.items()}

def parse_human_date(text: str):
    if not text:
        return None
    t = text.strip()
    # ISO first
    m = re.search(r"(20\d{2}-\d{2}-\d{2})", t)
    if m:
        return m.group(1)
    # dd/mm/yyyy or dd-mm-yyyy or dd.mm.yyyy
    m = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](20\d{2})", t)
    if m:
        d, mo, y = m.groups()
        try:
            return f"{y}-{int(mo):02d}-{int(d):02d}"
        except Exception:
            pass
    # Patterns with month names
    # Remove ordinal suffixes 1st 2nd 3rd 4th
    t2 = re.sub(r"(\d{1,2})(st|nd|rd|th)", r"\1", t.lower())
    # e.g. 31 jan 2025 or 31 jan
    m = re.search(r"(\d{1,2})\s+([a-z]{3,9})\s+(20\d{2})", t2)
    if m:
        d, mon, y = m.groups()
        mon_i = MONTH_MAP.get(mon) or MONTH_ABBR.get(mon[:3])
        if mon_i:
            try:
                return f"{y}-{int(mon_i):02d}-{int(d):02d}"
            except Exception:
                pass
    m = re.search(r"([a-z]{3,9})\s+(\d{1,2})\s*(,?\s*(20\d{2}))?", t2)
    if m:
        mon, d, _, y = m.group(1), m.group(2), m.group(3), m.group(4) if len(m.groups()) >= 4 else None
        mon_i = MONTH_MAP.get(mon) or MONTH_ABBR.get(mon[:3])
        if mon_i:
            year = y or str((datetime.utcnow() + timedelta(minutes=TZ_OFFSET_MINUTES)).year)
            try:
                return f"{year}-{int(mon_i):02d}-{int(d):02d}"
            except Exception:
                pass
    return None

def _fmt_display_dmy(iso: str) -> str:
    try:
        d = datetime.strptime(iso, '%Y-%m-%d').date()
        return f"{d.day:02d}/{d.month:02d}/{d.year}"
    except Exception:
        return iso

def _expand_date_range(start_iso: str, end_iso: str):
    try:
        a = datetime.strptime(start_iso, '%Y-%m-%d').date()
        b = datetime.strptime(end_iso, '%Y-%m-%d').date()
        if b < a:
            a, b = b, a
        out = []
        cur = a
        while cur <= b:
            out.append(cur.isoformat())
            cur += timedelta(days=1)
        return out
    except Exception:
        return [start_iso]

def _extract_date_range(text: str):
    """Return (start_iso, end_iso) if text contains a date range like
    'from 23/09/2025 till 26/09/2025' or '23.09.2025 to 25.09.2025'. Else None.
    """
    if not text:
        return None
    m = re.search(r"(\d{1,2}[./-]\d{1,2}[./-](?:20)?\d{2}).{0,30}?(?:till|to|until|through|thru|-|–|—).{0,30}?(\d{1,2}[./-]\d{1,2}[./-](?:20)?\d{2})", text, re.I)
    if not m:
        return None
    a, b = m.group(1), m.group(2)
    def norm(s):
        s = s.replace('.', '/').replace('-', '/')
        m2 = re.match(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", s)
        if not m2:
            return None
        d, mo, y = m2.groups()
        y = ('20'+y) if len(y)==2 else y
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
    sa = norm(a); sb = norm(b)
    if sa and sb:
        return (sa, sb)
    return None

# Data persistence helpers
_data_cache = None
_data_cache_mtime = 0
_data_lock = threading.Lock()

LEGACY_PER_DATE_FIELDS = ['status','status_reason','start_time','room','patient_count','opd','breaks','designation','before_break_opd_patients','after_break_opd_patients','before_break_opd','after_break_opd']

def _migrate_doctor_multidate(doc: dict):
    if not isinstance(doc, dict):
        return
    if 'schedule_by_date' in doc and isinstance(doc['schedule_by_date'], dict):
        return  # already migrated
    # Collect legacy fields if present
    legacy_present = any(k in doc for k in LEGACY_PER_DATE_FIELDS)
    if not legacy_present:
        doc.setdefault('schedule_by_date', {})
        return
    date_iso = get_internet_today_iso()
    entry = {}
    for k in LEGACY_PER_DATE_FIELDS:
        if k in doc:
            entry[k] = doc.pop(k)
    if entry:
        entry['last_update'] = datetime.utcnow().isoformat()+'Z'
    doc['schedule_by_date'] = {date_iso: entry} if entry else {}


def load_data(force=False):
    global _data_cache, _data_cache_mtime
    with _data_lock:
        try:
            st = os.stat(DATA_PATH)
            if (not force) and _data_cache is not None and st.st_mtime == _data_cache_mtime:
                return _data_cache
            if not os.path.exists(DATA_PATH):
                data = {'doctors': []}
            else:
                with open(DATA_PATH, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            # Ensure list structure
            if not isinstance(data, dict):
                data = {'doctors': []}
            data.setdefault('doctors', [])
            # Migrate each doctor
            for d in data['doctors']:
                _migrate_doctor_multidate(d)
            _data_cache = data
            _data_cache_mtime = st.st_mtime if os.path.exists(DATA_PATH) else time()
            return data
        except Exception:
            if _data_cache is None:
                _data_cache = {'doctors': []}
            return _data_cache


def save_data(data):
    if not isinstance(data, dict):
        return
    os.makedirs(os.path.dirname(DATA_PATH), exist_ok=True)
    tmp_path = DATA_PATH + '.tmp'
    with _data_lock:
        with open(tmp_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, DATA_PATH)
        # update cache reference
        global _data_cache, _data_cache_mtime
        _data_cache = data
        try:
            _data_cache_mtime = os.stat(DATA_PATH).st_mtime
        except Exception:
            _data_cache_mtime = time()


def flatten_schedule(doc: dict, limit_days: int = 14):
    """Return a flattened view of upcoming schedules for UI/API.
    limit_days: number of future (including today) days to include.
    """
    today = datetime.strptime(get_internet_today_iso(), '%Y-%m-%d').date()
    sched_map = doc.get('schedule_by_date', {}) if isinstance(doc, dict) else {}
    items = []
    for d_iso, val in sched_map.items():
        try:
            d_obj = datetime.strptime(d_iso, '%Y-%m-%d').date()
        except Exception:
            continue
        if d_obj < today - timedelta(days=1):  # allow yesterday for debug
            continue
        if d_obj > today + timedelta(days=limit_days):
            continue
        item = {'date': d_iso}
        for k in PER_DATE_FIELDS:
            if k in val:
                item[k] = val[k]
        items.append(item)
    items.sort(key=lambda x: x['date'])
    return items

# ---------------- User storage, roles & session auth ----------------
# Roles
ROLE_ADMIN = 'ADMIN'
ROLE_PR = 'PR'
ROLE_MED = 'MEDICAL_ADMIN'
ROLE_VIEW = 'VIEW_ONLY'

def _load_users():
    if not os.path.exists(USERS_PATH):
        return {'users': []}
    try:
        with open(USERS_PATH,'r',encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            data = {'users': []}
        data.setdefault('users',[])
        # Backfill missing role property (default to ADMIN for safety in legacy files)
        for u in data['users']:
            if 'role' not in u:
                # Default the very first user to ADMIN, others to PR as a sane default
                u['role'] = ROLE_ADMIN if u is data['users'][0] else ROLE_PR
        return data
    except Exception:
        return {'users': []}

def _save_users(data):
    os.makedirs(os.path.dirname(USERS_PATH), exist_ok=True)
    tmp = USERS_PATH + '.tmp'
    with open(tmp,'w',encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, USERS_PATH)

def _hash_pw(pw: str) -> str:
    return sha256((pw or '').encode('utf-8')).hexdigest()

def _ensure_default_admin():
    """Ensure there is at least one ADMIN account. Creates Admin/1234 if file is empty."""
    try:
        data = _load_users()
        if not data.get('users'):
            data['users'] = [{
                'username': 'Admin',
                'password': _hash_pw('1234'),
                'role': ROLE_ADMIN
            }]
            _save_users(data)
    except Exception:
        pass

def current_user():
    return session.get('user') or None

def require_roles(*roles):
    u = current_user()
    if not u:
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    if roles and u.get('role') not in roles:
        return jsonify({'ok': False, 'error': 'forbidden'}), 403
    return None

@app.get('/login')
def login_page():
    # If already logged in, go to admin
    if session.get('user'): return redirect(url_for('root'))
    _ensure_default_admin()
    return render_template('login.html')

@app.post('/login')
def login_post():
    data = request.json if request.is_json else request.form
    username = (data.get('username') or data.get('user') or '').strip()
    password = (data.get('password') or data.get('pass') or '').strip()
    users = _load_users().get('users', [])
    for u in users:
        if u.get('username') == username and u.get('password') == _hash_pw(password):
            session['user'] = {'username': username, 'role': u.get('role') or ROLE_PR}
            return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'invalid credentials'}), 401

@app.post('/logout')
def logout_post():
    session.pop('user', None)
    return jsonify({'ok': True})

@app.get('/api/me')
def api_me():
    u = current_user()
    if not u:
        return jsonify({'ok': False}), 401
    return jsonify({'ok': True, 'user': u})

@app.get('/register')
def register_page():
    # Registration is managed by admin inside Settings; direct page disabled
    if not current_user() or current_user().get('role') != ROLE_ADMIN:
        return redirect(url_for('root'))
    return render_template('register.html')

@app.post('/register')
def register_post():
    # Only admin can create users via API; prefer /api/users
    g = require_roles(ROLE_ADMIN)
    if g: return g
    data = request.json if request.is_json else request.form
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    role = (data.get('role') or ROLE_PR).strip().upper()
    if role not in (ROLE_ADMIN, ROLE_PR, ROLE_MED, ROLE_VIEW):
        role = ROLE_PR
    if not username or not password:
        return jsonify({'ok': False, 'error': 'username and password required'}), 400
    users_data = _load_users()
    if any(u.get('username') == username for u in users_data.get('users', [])):
        return jsonify({'ok': False, 'error': 'username exists'}), 400
    users_data['users'].append({'username': username, 'password': _hash_pw(password), 'role': role})
    _save_users(users_data)
    return jsonify({'ok': True})

@app.get('/reset')
def reset_page():
    if not current_user() or current_user().get('role') != ROLE_ADMIN:
        return redirect(url_for('root'))
    return render_template('reset.html')

@app.post('/reset')
def reset_post():
    g = require_roles(ROLE_ADMIN)
    if g: return g
    data = request.json if request.is_json else request.form
    username = (data.get('username') or '').strip()
    new_pw = (data.get('password') or '').strip()
    if not username or not new_pw:
        return jsonify({'ok': False, 'error': 'username and password required'}), 400
    users = _load_users()
    found = False
    for u in users.get('users', []):
        if u.get('username') == username:
            u['password'] = _hash_pw(new_pw)
            found = True
            break
    if not found:
        return jsonify({'ok': False, 'error': 'username not found'}), 404
    _save_users(users)
    return jsonify({'ok': True})

# Admin-managed users API
@app.get('/api/users')
def api_users_list():
    g = require_roles(ROLE_ADMIN)
    if g: return g
    data = _load_users()
    out = [{'username': u.get('username'), 'role': u.get('role', ROLE_PR)} for u in data.get('users',[])]
    return jsonify({'ok': True, 'users': out})

@app.post('/api/users')
def api_users_create():
    g = require_roles(ROLE_ADMIN)
    if g: return g
    payload = request.json or {}
    username = (payload.get('username') or '').strip()
    password = (payload.get('password') or '').strip()
    role = (payload.get('role') or ROLE_PR).strip().upper()
    if role not in (ROLE_ADMIN, ROLE_PR, ROLE_MED, ROLE_VIEW):
        return jsonify({'ok': False, 'error': 'invalid role'}), 400
    if not username or not password:
        return jsonify({'ok': False, 'error': 'username and password required'}), 400
    data = _load_users()
    if any(u.get('username') == username for u in data.get('users', [])):
        return jsonify({'ok': False, 'error': 'username exists'}), 400
    data['users'].append({'username': username, 'password': _hash_pw(password), 'role': role})
    _save_users(data)
    return jsonify({'ok': True})

@app.patch('/api/users/<username>')
def api_users_patch(username):
    g = require_roles(ROLE_ADMIN)
    if g: return g
    payload = request.json or {}
    data = _load_users()
    for u in data.get('users', []):
        if u.get('username') == username:
            if 'password' in payload and payload['password']:
                u['password'] = _hash_pw(payload['password'])
            if 'role' in payload:
                role = str(payload['role']).upper()
                if role in (ROLE_ADMIN, ROLE_PR, ROLE_MED, ROLE_VIEW):
                    u['role'] = role
            _save_users(data)
            return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'not found'}), 404

@app.delete('/api/users/<username>')
def api_users_delete(username):
    g = require_roles(ROLE_ADMIN)
    if g: return g
    data = _load_users()
    before = len(data.get('users', []))
    data['users'] = [u for u in data.get('users', []) if u.get('username') != username]
    if len(data['users']) == before:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    # Make sure at least one admin remains; if none, add default Admin
    if not any(u.get('role') == ROLE_ADMIN for u in data['users']):
        data['users'].append({'username': 'Admin', 'password': _hash_pw('1234'), 'role': ROLE_ADMIN})
    _save_users(data)
    return jsonify({'ok': True})

# --- Basic API & HTML endpoints (minimal set) ---

@app.get('/')
def root():
    # Require session login; legacy token no longer used for UI access
    if not session.get('user'):
        return redirect(url_for('login_page'))
    return render_template('admin.html', admin_token=ADMIN_TOKEN, role=session['user'].get('role', ROLE_PR), logo_version=_LOGO_VERSION)

@app.get('/display')
def display_page():
    # Provide doctors + flattened schedules for rendering template.
    # Existing template expects richer structure (specialty ordering, etc.).
    data = load_data()
    today_iso = get_internet_today_iso()
    # Build per specialty order; fallback if missing
    specialty_order = data.get('specialty_order') or sorted({d.get('specialty','') for d in data.get('doctors',[])})
    # Minimal serialization for initial paint (client will poll /api/window & /api/doctors)
    base_doctors = []
    for d in data.get('doctors', []):
        # Merge today's per-date schedule so initial render reflects correct status/timings
        hd = hydrate_doctor_for_date(d, today_iso)
        base_doctors.append({
            'id': d.get('id'),
            'name': d.get('name'),
            'specialty': d.get('specialty'),
            'designation': hd.get('designation', d.get('designation')),
            'status': hd.get('status', d.get('status')),
            'start_time': hd.get('start_time', d.get('start_time')),
            'room': hd.get('room', d.get('room')),
            'patient_count': hd.get('patient_count', d.get('patient_count')),
            'opd': hd.get('opd', d.get('opd')),
            'breaks': hd.get('breaks', d.get('breaks')),
            # Include explicit before/after arrays and patient counts so groups render immediately
            'before_break_opd': hd.get('before_break_opd') or d.get('before_break_opd'),
            'after_break_opd': hd.get('after_break_opd') or d.get('after_break_opd'),
            'before_break_opd_patients': hd.get('before_break_opd_patients') if 'before_break_opd_patients' in hd else d.get('before_break_opd_patients'),
            'after_break_opd_patients': hd.get('after_break_opd_patients') if 'after_break_opd_patients' in hd else d.get('after_break_opd_patients'),
            'after_break_note': hd.get('after_break_note') if 'after_break_note' in hd else d.get('after_break_note'),
            'image_version': d.get('image_version', 1)
        })
    closures = _load_closures()
    closure_today = None
    # Check if today is Friday (weekday 4 in Python, 0=Monday)
    from datetime import datetime
    today_date = datetime.fromisoformat(today_iso)
    is_friday = today_date.weekday() == 4
    
    # ONLY show server banner for manually set closed days (not Fridays)
    # Friday overlay is handled by client-side JavaScript in display.js
    if today_iso in closures:
        # store last reason
        reason = closures[today_iso][-1] if isinstance(closures[today_iso], list) else closures[today_iso]
        closure_today = {'active': True, 'reason': reason}
    # Do NOT set closure banner for Friday - client-side handles it dynamically
    # Find display logo (fallback to regular logo if not found)
    display_logo_path = None
    for ext in ALLOWED_EXTENSIONS:
        candidate = os.path.join('static', 'img', 'display-logo' + ext)
        if os.path.exists(candidate):
            display_logo_path = '/static/img/display-logo' + ext
            break
    if not display_logo_path:
        for ext in ALLOWED_EXTENSIONS:
            candidate = os.path.join('static', 'img', 'hulhumale-logo' + ext)
            if os.path.exists(candidate):
                display_logo_path = '/static/img/hulhumale-logo' + ext
                break
    try:
        return render_template('display.html', doctors=base_doctors, today=today_iso,
                               today_iso=today_iso, today_str=today_iso,
                               data={'specialty_order': specialty_order, 'doctors': base_doctors},
                               closure=closure_today, logo_version=_LOGO_VERSION,
                               display_logo=display_logo_path,
                               contact_phone=os.environ.get('CONTACT_PHONE',''),
                               contact_address=os.environ.get('CONTACT_ADDRESS',''))
    except Exception:
        return jsonify({'doctors': base_doctors})

@app.get('/patient')
def patient_page():
    """Public patient-friendly display: simple cards like admin list showing photo, name,
    specialty/designation, start time, room and break times for today.
    Uses client-side fetch of /api/doctors to render; initial render provides specialty order
    and a minimal snapshot for first paint.
    """
    data = load_data()
    today_iso = get_internet_today_iso()
    # Build minimal snapshot for initial load
    base_doctors = []
    for d in data.get('doctors', []):
        hd = hydrate_doctor_for_date(d, today_iso)
        base_doctors.append({
            'id': d.get('id'),
            'name': d.get('name'),
            'specialty': d.get('specialty'),
            'designation': hd.get('designation', d.get('designation')),
            'start_time': hd.get('start_time', d.get('start_time')),
            'room': hd.get('room', d.get('room')),
            'breaks': hd.get('breaks', d.get('breaks')),
            'image_version': d.get('image_version', 1),
            'status': hd.get('status', d.get('status','PENDING'))
        })
    specialty_order = data.get('specialty_order') or sorted({d.get('specialty','') for d in data.get('doctors',[])})
    settings = load_settings()
    patient_cfg = settings.get('patient_display', {})
    return render_template('patient.html', today_iso=today_iso,
                           data={'specialty_order': specialty_order, 'doctors': base_doctors, 'patient_display': patient_cfg, 'today_iso': today_iso})

@app.get('/api/doctors')
def api_doctors():
    """Return doctors list for both admin & display.
    Includes specialty_order and each doctor with merged *today* fields while preserving schedule_by_date for admin editing.
    """
    data = load_data()
    today_iso = get_internet_today_iso()
    specialty_order = data.get('specialty_order') or []
    specialty_designations = data.get('specialty_designations', {})
    enriched = []
    for d in data.get('doctors', []):
        per_date = {}
        if isinstance(d.get('schedule_by_date'), dict):
            per_date = d['schedule_by_date'].get(today_iso, {})
        doc = dict(d)  # shallow copy of stored structure
        # Merge today's schedule fields (do not overwrite stored values permanently)
        for fld in (
            'designation','status','start_time','room','patient_count','opd','breaks','status_reason',
            'before_break_opd','after_break_opd','before_break_opd_patients','after_break_opd_patients'
        ):
            if fld in per_date:
                doc[fld] = per_date[fld]
        doc['for_date'] = today_iso
        enriched.append(doc)
    return jsonify({'specialty_order': specialty_order, 'specialty_designations': specialty_designations, 'doctors': enriched, 'today': today_iso})

@app.post('/api/doctors')
def api_doctors_create():
    payload = request.json if request.is_json else request.form.to_dict()
    name = (payload.get('name') or '').strip()
    specialty = (payload.get('specialty') or '').strip() or 'General'
    if not name:
        return jsonify({'ok': False, 'error': 'name required'}), 400
    data = load_data()
    # Determine new id (numeric increment)
    existing_ids = {str(d.get('id')) for d in data.get('doctors', [])}
    nid = 1
    while str(nid) in existing_ids:
        nid += 1
    # Default designation based on specialty mapping if available
    s_map = data.get('specialty_designations', {})
    designation = (payload.get('designation') or '').strip() or s_map.get(specialty, '')
    doc = {
        'id': str(nid),
        'name': name,
        'specialty': specialty,
        'keywords': payload.get('keywords') or [],
        'start_time': payload.get('start_time','08:00'),
        'designation': designation,
        'notes': payload.get('notes',''),
        'image_version': 1,
        'schedule_by_date': {}
    }
    data.setdefault('doctors', []).append(doc)
    # Add specialty if new
    if specialty and 'specialty_order' in data and specialty not in data['specialty_order']:
        data['specialty_order'].append(specialty)
    save_data(data)
    broker.publish_event('doctor_create', {'doctor_id': doc['id']})
    return jsonify({'ok': True, 'doctor': doc})

@app.patch('/api/doctors/<doc_id>')
def api_doctors_patch(doc_id):
    payload = request.json if request.is_json else request.form.to_dict()
    data = load_data()
    doc = next((d for d in data.get('doctors', []) if str(d.get('id')) == str(doc_id)), None)
    if not doc:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    # Basic fields
    for base_field in ('name','specialty','notes','keywords','start_time','designation','room','patient_count','opd','breaks','status_reason'):
        if base_field in payload and base_field not in ('patient_count','opd','breaks'):
            doc[base_field] = payload[base_field]
        elif base_field == 'patient_count' and payload.get('patient_count') is not None:
            try: doc['patient_count'] = int(payload['patient_count'])
            except: pass
        elif base_field == 'opd' and payload.get('opd') is not None:
            doc['opd'] = payload['opd']
        elif base_field == 'breaks' and payload.get('breaks') is not None:
            doc['breaks'] = payload['breaks']
    # If specialty changed and designation not explicitly provided, auto-assign default designation
    if 'specialty' in payload and 'designation' not in payload:
        data_map = data.get('specialty_designations', {})
        auto_desig = data_map.get(doc.get('specialty'))
        if auto_desig is not None:
            doc['designation'] = auto_desig
    # Per-date patch if for_date provided
    for_date = payload.get('for_date') or payload.get('date')
    per_patch = {k: payload[k] for k in PER_DATE_FIELDS if k in payload}
    changed = False
    if for_date and per_patch:
        if apply_single_day_update(doc, for_date, per_patch):
            changed = True
    save_data(data)
    if changed:
        broker.publish_event('doctor_update', {'doctor_id': doc_id, 'date': for_date})
    return jsonify({'ok': True, 'doctor': doc})

@app.delete('/api/doctors/<doc_id>')
def api_doctors_delete(doc_id):
    data = load_data()
    before = len(data.get('doctors', []))
    data['doctors'] = [d for d in data.get('doctors', []) if str(d.get('id')) != str(doc_id)]
    if len(data['doctors']) == before:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    save_data(data)
    broker.publish_event('doctor_update', {'doctor_id': doc_id, 'deleted': True})
    return jsonify({'ok': True})

@app.post('/api/schedule/clear-date')
def api_schedule_clear_date():
    """Clear all schedules for a specific date - completely removes schedule_by_date entry"""
    payload = request.json if request.is_json else {}
    date_iso = payload.get('date')
    
    if not date_iso:
        return jsonify({'ok': False, 'error': 'date required'}), 400
    
    # Validate date format
    try:
        datetime.strptime(date_iso, '%Y-%m-%d')
    except ValueError:
        return jsonify({'ok': False, 'error': 'invalid date format (use YYYY-MM-DD)'}), 400
    
    data = load_data()
    cleared_count = 0
    doctor_ids = []
    
    for doc in data.get('doctors', []):
        if 'schedule_by_date' in doc and date_iso in doc['schedule_by_date']:
            # Completely remove the date entry
            del doc['schedule_by_date'][date_iso]
            cleared_count += 1
            doctor_ids.append(doc['id'])
    
    if cleared_count > 0:
        save_data(data)
        # Trigger SSE update for all affected doctors
        broker.publish_event('schedule_cleared', {
            'date': date_iso,
            'doctor_ids': doctor_ids,
            'count': cleared_count
        })
        log(f"Cleared {cleared_count} schedules for {date_iso}")
    
    return jsonify({
        'ok': True,
        'date': date_iso,
        'cleared': cleared_count,
        'doctor_ids': doctor_ids
    })

def _save_media_history(doc_id, file_path, media_type='photo'):
    """Save uploaded media to history folder with timestamp"""
    try:
        os.makedirs(MEDIA_HISTORY_DIR, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        ext = os.path.splitext(file_path)[1]
        history_filename = f"{doc_id}_{media_type}_{timestamp}{ext}"
        history_path = os.path.join(MEDIA_HISTORY_DIR, history_filename)
        
        # Copy file to history
        import shutil
        shutil.copy2(file_path, history_path)
        
        # Also save a log entry
        log_entry = {
            'timestamp': datetime.now().isoformat(),
            'doctor_id': doc_id,
            'media_type': media_type,
            'filename': history_filename,
            'original_name': os.path.basename(file_path)
        }
        
        history_log_path = os.path.join(MEDIA_HISTORY_DIR, 'upload_history.json')
        history_log = []
        if os.path.exists(history_log_path):
            try:
                with open(history_log_path, 'r', encoding='utf-8') as f:
                    history_log = json.load(f)
            except:
                history_log = []
        
        history_log.append(log_entry)
        
        # Keep only last 1000 entries
        if len(history_log) > 1000:
            history_log = history_log[-1000:]
        
        with open(history_log_path, 'w', encoding='utf-8') as f:
            json.dump(history_log, f, indent=2, ensure_ascii=False)
            
        log(f"Saved media history: {history_filename}")
    except Exception as e:
        log(f"Failed to save media history: {e}")

@app.post('/api/doctors/<doc_id>/photo')
def api_doctor_photo(doc_id):
    # Basic token check (optional - skip if not supplied)
    token = request.headers.get('X-Admin-Token') or request.args.get('token') or request.form.get('token')
    if ADMIN_TOKEN and token and token != ADMIN_TOKEN:
        return jsonify({'ok': False, 'error': 'unauthorized'}), 403
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': 'empty filename'}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'ok': False, 'error': 'invalid file type'}), 400
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    target = os.path.join(UPLOAD_DIR, f"{doc_id}{ext}")
    # Remove old versions with different extension
    for e in list(ALLOWED_EXTENSIONS):
        old = os.path.join(UPLOAD_DIR, f"{doc_id}{e}")
        if os.path.exists(old) and old != target:
            try: os.remove(old)
            except: pass
    f.save(target)
    
    # Save to media history
    _save_media_history(doc_id, target, 'photo')
    
    data = load_data(True)
    for d in data.get('doctors', []):
        if str(d.get('id')) == str(doc_id):
            d['image_version'] = int(d.get('image_version', 1)) + 1
            break
    save_data(data)
    broker.publish_event('doctor_update', {'doctor_id': doc_id, 'image_version': d.get('image_version',1)})
    return jsonify({'ok': True, 'image_version': d.get('image_version',1)})

@app.get('/admin')
def admin_page():
    # Combined admin + display snapshot for today's date.
    data = load_data()
    today_iso = get_internet_today_iso()
    out = []
    for d in data.get('doctors', []):
        per_date = {}
        if isinstance(d.get('schedule_by_date'), dict):
            per_date = d['schedule_by_date'].get(today_iso, {})
        merged = dict(
            id=d.get('id'),
            name=d.get('name'),
            specialty=d.get('specialty'),
            keywords=d.get('keywords', []),
            notes=d.get('notes'),
            designation=per_date.get('designation', d.get('designation')),
            status=per_date.get('status', d.get('status','PENDING')),
            start_time=per_date.get('start_time', d.get('start_time')),
            room=per_date.get('room', d.get('room')),
            patient_count=per_date.get('patient_count', d.get('patient_count')),
            opd=per_date.get('opd', d.get('opd')),
            breaks=per_date.get('breaks', d.get('breaks')),
            status_reason=per_date.get('status_reason', d.get('status_reason')),
            before_break_opd_patients=per_date.get('before_break_opd_patients'),
            after_break_opd_patients=per_date.get('after_break_opd_patients'),
            before_break_opd=per_date.get('before_break_opd'),
            after_break_opd=per_date.get('after_break_opd'),
            image_version=d.get('image_version', 1),
            for_date=today_iso if per_date else None
        )
        out.append(merged)
    return jsonify({'specialty_order': data.get('specialty_order', []), 'doctors': out})

# ---------------- Admin REST Endpoints (token protected) ----------------
def _check_admin(req):
    # First, support session-based auth
    if session.get('user'):  # logged-in user
        return True
    # Fallback: legacy token for API compatibility
    token = req.headers.get('X-Admin-Token') or req.args.get('token') or (req.json or {}).get('token') if req.is_json else None
    return token == ADMIN_TOKEN

def _admin_guard():
    if not _check_admin(request):
        return jsonify({'ok': False, 'error': 'unauthorized'}), 401
    return None

@app.post('/api/doctors')
def admin_create_doctor():
    g = _admin_guard();
    if g: return g
    payload = request.json or {}
    name = (payload.get('name') or '').strip()
    specialty = (payload.get('specialty') or '').strip() or 'General'
    if not name:
        return jsonify({'ok': False, 'error': 'name required'}), 400
    data = load_data()
    existing_ids = {str(d.get('id')) for d in data.get('doctors',[])}
    nid = 1
    while str(nid) in existing_ids: nid += 1
    # Default designation from specialty mapping if not provided
    s_map = load_data().get('specialty_designations', {})
    designation = (payload.get('designation') or '').strip() or s_map.get(specialty, '')
    doc = {
        'id': nid,
        'name': name,
        'specialty': specialty,
        'keywords': payload.get('keywords', []),
        'notes': payload.get('notes'),
        'designation': designation,
        'start_time': payload.get('start_time','08:00'),
        'image_version': 1,
        'schedule_by_date': {}
    }
    data.setdefault('doctors', []).append(doc)
    if specialty and specialty not in data.get('specialty_order', []):
        data.setdefault('specialty_order', []).append(specialty)
    save_data(data)
    broker.publish_event('doctor_create', {'doctor_id': doc['id']})
    return jsonify({'ok': True, 'doctor': doc})

@app.patch('/api/doctors/<doc_id>')
def admin_patch_doctor(doc_id):
    g = _admin_guard();
    if g: return g
    payload = request.json or {}
    data = load_data()
    doc = next((d for d in data.get('doctors', []) if str(d.get('id')) == str(doc_id)), None)
    if not doc:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    # Basic fields
    for k in ('name','specialty','notes','keywords','designation'):
        if k in payload:
            doc[k] = payload[k]
    if 'keywords' in payload and isinstance(doc.get('keywords'), list):
        doc['keywords'] = [k for k in doc['keywords'] if isinstance(k,str)]
    # If specialty changed and designation not explicitly provided, auto-assign default designation
    if 'specialty' in payload and 'designation' not in payload:
        data_map = load_data().get('specialty_designations', {})
        auto_desig = data_map.get(doc.get('specialty'))
        if auto_desig is not None:
            doc['designation'] = auto_desig
    # Per-date update
    date_iso = payload.get('for_date') or get_internet_today_iso()
    patch = {k: payload[k] for k in PER_DATE_FIELDS if k in payload}
    changed = False
    if patch:
        if apply_single_day_update(doc, date_iso, patch):
            changed = True
    if changed:
        save_data(data)
        broker.publish_event('doctor_update', {'doctor_id': doc.get('id'), 'date': date_iso})
    return jsonify({'ok': True, 'doctor': doc, 'changed': changed})

@app.delete('/api/doctors/<doc_id>')
def admin_delete_doctor(doc_id):
    g = _admin_guard();
    if g: return g
    data = load_data()
    before = len(data.get('doctors', []))
    data['doctors'] = [d for d in data.get('doctors', []) if str(d.get('id')) != str(doc_id)]
    if len(data['doctors']) == before:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    save_data(data)
    broker.publish_event('doctor_delete', {'doctor_id': doc_id})
    return jsonify({'ok': True})

@app.post('/api/doctors/<doc_id>/photo')
def admin_upload_photo(doc_id):
    g = _admin_guard();
    if g: return g
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'ok': False, 'error': 'empty name'}), 400
    ext = os.path.splitext(file.filename)[1].lower().strip()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'ok': False, 'error': 'unsupported file type', 'allowed': sorted(list(ALLOWED_EXTENSIONS))}), 400
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    path = os.path.join(UPLOAD_DIR, f"{doc_id}{ext}")
    for e in ALLOWED_EXTENSIONS:
        old = os.path.join(UPLOAD_DIR, f"{doc_id}{e}")
        if os.path.exists(old) and old != path:
            try: os.remove(old)
            except Exception: pass
    file.save(path)
    data = load_data(force=True)
    for d in data.get('doctors', []):
        if str(d.get('id')) == str(doc_id):
            d['image_version'] = int(d.get('image_version',1)) + 1
            break
    save_data(data)
    return jsonify({'ok': True, 'image_version': d.get('image_version',1)})

@app.post('/api/doctors/<doc_id>/promo')
def admin_upload_promo(doc_id):
    """Upload full-screen promotional image for patient display slideshow."""
    g = _admin_guard();
    if g: return g
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'ok': False, 'error': 'empty name'}), 400
    ext = os.path.splitext(file.filename)[1].lower().strip()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'ok': False, 'error': 'unsupported file type', 'allowed': sorted(list(ALLOWED_EXTENSIONS))}), 400
    os.makedirs(PROMO_DIR, exist_ok=True)
    path = os.path.join(PROMO_DIR, f"{doc_id}{ext}")
    # Remove old promo images with different extensions
    for e in ALLOWED_EXTENSIONS:
        old = os.path.join(PROMO_DIR, f"{doc_id}{e}")
        if os.path.exists(old) and old != path:
            try: os.remove(old)
            except Exception: pass
    file.save(path)
    
    # Save to media history
    _save_media_history(doc_id, path, 'promo')
    
    data = load_data(force=True)
    for d in data.get('doctors', []):
        if str(d.get('id')) == str(doc_id):
            d['promo_version'] = int(d.get('promo_version',1)) + 1
            break
    save_data(data)
    return jsonify({'ok': True, 'promo_version': d.get('promo_version',1)})

@app.get('/doctor-promo/<doc_id>')
def get_promo_image(doc_id):
    """Serve promotional image for patient display, fallback to regular photo."""
    for ext in ('.png','.jpg','.jpeg','.webp','.gif'):
        candidate = os.path.join(PROMO_DIR, f"{doc_id}{ext}")
        if os.path.exists(candidate):
            return send_file(candidate)
    # Fallback to regular photo or default
    for ext in ('.png','.jpg','.jpeg','.webp','.gif'):
        candidate = os.path.join(UPLOAD_DIR, f"{doc_id}{ext}")
        if os.path.exists(candidate):
            return send_file(candidate)
    return send_file(os.path.join('static','img','default-doctor.png'))

@app.post('/api/specialties')
def admin_add_specialty():
    g = _admin_guard();
    if g: return g
    body = (request.json or {})
    name = (body.get('name') or '').strip()
    designation = (body.get('designation') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'name required'}), 400
    data = load_data()
    order = data.setdefault('specialty_order', [])
    if name not in order:
        order.append(name)
    # store default designation mapping
    s_map = data.setdefault('specialty_designations', {})
    if designation or name not in s_map:
        s_map[name] = designation
    # Auto-assign designation to existing doctors of this specialty if they don't have one
    if designation:
        for d in data.get('doctors', []):
            try:
                if (d.get('specialty') == name) and not (d.get('designation') or '').strip():
                    d['designation'] = designation
            except Exception:
                pass
    save_data(data)
    return jsonify({'ok': True, 'specialty_order': order, 'specialty_designations': s_map})

@app.patch('/api/specialties/<path:spec>')
def admin_rename_specialty(spec):
    g = _admin_guard();
    if g: return g
    body = (request.json or {})
    new_name = (body.get('name') or '').strip()
    new_designation = (body.get('designation') or '').strip()
    if not new_name:
        return jsonify({'ok': False, 'error': 'name required'}), 400
    data = load_data()
    changed = False
    # Update specialty order list
    for i, s in enumerate(data.get('specialty_order', [])):
        if s == spec:
            data['specialty_order'][i] = new_name; changed = True
    for d in data.get('doctors', []):
        if d.get('specialty') == spec:
            d['specialty'] = new_name; changed = True
    # Move/update designation mapping
    s_map = data.setdefault('specialty_designations', {})
    if spec in s_map and new_name != spec:
        s_map[new_name] = s_map.pop(spec)
        changed = True
    if new_designation:
        s_map[new_name] = new_designation
        changed = True
    if changed:
        save_data(data)
    return jsonify({'ok': True, 'changed': changed, 'specialty_order': data.get('specialty_order', []), 'specialty_designations': data.get('specialty_designations', {})})

@app.delete('/api/specialties/<path:spec>')
def admin_delete_specialty(spec):
    g = _admin_guard();
    if g: return g
    data = load_data()
    if any(d.get('specialty') == spec for d in data.get('doctors', [])):
        return jsonify({'ok': False, 'error': 'in use'}), 400
    order = data.get('specialty_order', [])
    if spec in order:
        order.remove(spec)
        # remove designation default if present
        if 'specialty_designations' in data and spec in data['specialty_designations']:
            data['specialty_designations'].pop(spec, None)
        save_data(data); return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'not found'}), 404

@app.post('/api/specialties/<path:spec>/advanced_delete')
def admin_adv_delete(spec):
    g = _admin_guard();
    if g: return g
    payload = request.json or {}
    mode = payload.get('mode')
    data = load_data()
    if mode == 'delete_all':
        data['doctors'] = [d for d in data.get('doctors', []) if d.get('specialty') != spec]
        if spec in data.get('specialty_order', []):
            data['specialty_order'].remove(spec)
        if 'specialty_designations' in data:
            data['specialty_designations'].pop(spec, None)
        save_data(data)
        return jsonify({'ok': True, 'deleted': spec})
    if mode == 'migrate':
        target = (payload.get('target') or '').strip()
        if not target:
            return jsonify({'ok': False, 'error': 'target required'}), 400
        for d in data.get('doctors', []):
            if d.get('specialty') == spec:
                d['specialty'] = target
        if target and target not in data.get('specialty_order', []):
            data['specialty_order'].append(target)
        if spec in data.get('specialty_order', []):
            data['specialty_order'].remove(spec)
        # Remove old designation mapping for spec; don't alter target mapping
        if 'specialty_designations' in data:
            data['specialty_designations'].pop(spec, None)
        save_data(data)
        return jsonify({'ok': True, 'migrated': spec, 'to': target})
    return jsonify({'ok': False, 'error': 'bad mode'}), 400

@app.post('/api/specialties/reorder')
def admin_reorder_specialties():
    g = _admin_guard();
    if g: return g
    payload = request.json or {}
    new_order = payload.get('order') or []
    if not isinstance(new_order, list) or not all(isinstance(x, str) for x in new_order):
        return jsonify({'ok': False, 'error': 'order must be list of strings'}), 400
    data = load_data()
    current = data.get('specialty_order', [])
    # Validate same set (allow ignoring whitespace / case exact match scenario)
    cur_set = set(current)
    new_set = set(new_order)
    if cur_set != new_set:
        return jsonify({'ok': False, 'error': 'order must contain exactly existing specialties'}), 400
    data['specialty_order'] = new_order
    save_data(data)
    broker.publish_event('specialty_order_updated', {'order': new_order})
    return jsonify({'ok': True, 'order': new_order})

@app.post('/api/logo')
def upload_logo():
    g = require_roles(ROLE_ADMIN)
    if g: return g
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    f = request.files['file']
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'ok': False, 'error': 'image file required (PNG, JPG, etc.)'}), 400
    os.makedirs(os.path.dirname(LOGO_PATH), exist_ok=True)
    # Remove old logos before saving new one
    for old_ext in ALLOWED_EXTENSIONS:
        old_file = os.path.join(os.path.dirname(LOGO_PATH), 'hulhumale-logo' + old_ext)
        if os.path.exists(old_file):
            try: os.remove(old_file)
            except: pass
    # Save with original extension
    target = os.path.join(os.path.dirname(LOGO_PATH), 'hulhumale-logo' + ext)
    f.save(target)
    global _LOGO_VERSION; _LOGO_VERSION += 1
    broker.publish_event('branding_updated', {'logo_version': _LOGO_VERSION, 'type': 'login_logo'})
    return jsonify({'ok': True, 'logo_version': _LOGO_VERSION})

@app.post('/api/login_background')
def upload_login_background():
    g = require_roles(ROLE_ADMIN)
    if g: return g
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    f = request.files['file']
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'ok': False, 'error': 'image file required'}), 400
    os.makedirs(os.path.join('static', 'img'), exist_ok=True)
    target = os.path.join('static', 'img', 'login-background' + ext)
    # Remove old backgrounds
    for old_ext in ALLOWED_EXTENSIONS:
        old_file = os.path.join('static', 'img', 'login-background' + old_ext)
        if os.path.exists(old_file) and old_file != target:
            try: os.remove(old_file)
            except: pass
    f.save(target)
    return jsonify({'ok': True, 'path': '/static/img/login-background' + ext})

@app.post('/api/display_logo')
def upload_display_logo():
    g = require_roles(ROLE_ADMIN)
    if g: return g
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'file required'}), 400
    f = request.files['file']
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({'ok': False, 'error': 'image file required'}), 400
    os.makedirs(os.path.join('static', 'img'), exist_ok=True)
    target = os.path.join('static', 'img', 'display-logo' + ext)
    # Remove old display logos
    for old_ext in ALLOWED_EXTENSIONS:
        old_file = os.path.join('static', 'img', 'display-logo' + old_ext)
        if os.path.exists(old_file) and old_file != target:
            try: os.remove(old_file)
            except: pass
    f.save(target)
    global _LOGO_VERSION; _LOGO_VERSION += 1
    broker.publish_event('branding_updated', {'logo_version': _LOGO_VERSION, 'type': 'display_logo'})
    return jsonify({'ok': True, 'path': '/static/img/display-logo' + ext, 'logo_version': _LOGO_VERSION})

@app.get('/api/branding')
def api_branding():
    """Return current logo and login background paths for client to fetch."""
    logo = None
    for ext in ALLOWED_EXTENSIONS:
        candidate = os.path.join('static', 'img', 'hulhumale-logo' + ext)
        if os.path.exists(candidate):
            logo = '/static/img/hulhumale-logo' + ext
            break
    bg = None
    for ext in ALLOWED_EXTENSIONS:
        candidate = os.path.join('static', 'img', 'login-background' + ext)
        if os.path.exists(candidate):
            bg = '/static/img/login-background' + ext
            break
    display_logo = None
    for ext in ALLOWED_EXTENSIONS:
        candidate = os.path.join('static', 'img', 'display-logo' + ext)
        if os.path.exists(candidate):
            display_logo = '/static/img/display-logo' + ext
            break
    return jsonify({'ok': True, 'logo': logo, 'background': bg, 'display_logo': display_logo, 'logo_version': _LOGO_VERSION})

# --- Closures (closed days) API ---
@app.get('/api/closures')
def api_closures_get():
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    raw = _load_closures()
    out = {}
    for k, v in raw.items():
        if isinstance(v, list) and v:
            out[k] = {'reasons': v, 'reason': v[-1]}
        else:
            out[k] = {'reasons': [v] if v else [], 'reason': v if isinstance(v, str) else ''}
    return jsonify({'dates': out})

@app.post('/api/closures')
def api_closures_add():
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    payload = request.json or {}
    dates = payload.get('dates') or []
    reason = (payload.get('reason') or 'TODAY IS A OPD CLOSED DAY').strip()
    if not isinstance(dates, list) or not dates:
        return jsonify({'ok': False, 'error': 'dates required'}), 400
    closures = _load_closures()
    for d in dates:
        if not isinstance(closures.get(d), list):
            closures[d] = [] if d not in closures else closures[d]
        if reason and reason not in closures[d]:
            closures[d].append(reason)
    _save_closures(closures)
    broker.publish_event('closure_update', {'dates': dates, 'reason': reason})
    return jsonify({'ok': True})

@app.patch('/api/closures')
def api_closures_bulk_reopen():
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    payload = request.json or {}
    dates = payload.get('dates') or []
    closures = _load_closures()
    removed = []
    for d in dates:
        if d in closures:
            closures.pop(d, None); removed.append(d)
    _save_closures(closures)
    if removed:
        for d in removed:
            broker.publish_event('closure_update', {'date': d, 'reopen': True})
    return jsonify({'ok': True, 'removed': removed})

@app.delete('/api/closures/<date_iso>')
def api_closure_delete(date_iso):
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    closures = _load_closures()
    if date_iso in closures:
        closures.pop(date_iso, None); _save_closures(closures); broker.publish_event('closure_update', {'date': date_iso, 'reopen': True})
        return jsonify({'ok': True})
    return jsonify({'ok': False, 'error': 'not found'}), 404

# ============================================================================
# OPD INFORMATION API
# ============================================================================
OPDINFO_PATH = os.path.join('data', 'opdinfo.json')

def _load_opdinfo():
    if not os.path.exists(OPDINFO_PATH):
        return {}
    try:
        with open(OPDINFO_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return {}

def _save_opdinfo(data):
    with open(OPDINFO_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.get('/api/opdinfo')
def api_opdinfo_get():
    """Get all OPD information"""
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    data = _load_opdinfo()
    return jsonify({'ok': True, 'data': data})

@app.post('/api/opdinfo')
def api_opdinfo_save():
    """Save OPD information"""
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    payload = request.json or {}
    data = payload.get('data', {})
    _save_opdinfo(data)
    broker.publish_event('opdinfo_update', {'specialties': list(data.keys())})
    return jsonify({'ok': True})

@app.get('/debug/telegram_status')
def debug_telegram_status():
    # Provide simple plaintext snapshot for admin panel
    lines = [f"Telegram enabled: {ENABLE_TELEGRAM}", f"Running: {_telegram_thread.is_alive() if _telegram_thread else False}", f"Offset: {_telegram_offset}", f"Fail streak: {_telegram_fail_streak}"]
    try:
        if os.path.exists(LOG_PATH):
            with open(LOG_PATH,'r',encoding='utf-8') as f:
                recent = f.readlines()[-100:]
            lines.append('\nLast 100 lines:\n' + ''.join(recent))
    except Exception as e:
        lines.append(f"(log read error: {e})")
    return Response('\n'.join(lines), mimetype='text/plain')

@app.get('/api/window')
def api_window():
    """Return window of days with doctors merged per date for date navigation UI."""
    start = request.args.get('start') or 'auto'
    if start.lower() == 'auto':
        base = datetime.strptime(get_internet_today_iso(), '%Y-%m-%d').date()
    else:
        try:
            base = datetime.strptime(start, '%Y-%m-%d').date()
        except Exception:
            base = datetime.strptime(get_internet_today_iso(), '%Y-%m-%d').date()
    days = max(1, min(30, int(request.args.get('days', '14'))))
    data = load_data()
    window = []
    today_iso = get_internet_today_iso()
    # Iterate through more days to compensate for skipped Fridays (add ~20% more)
    max_iterations = days + (days // 4) + 5
    for i in range(max_iterations):
        if len(window) >= days:
            break
        day = base + timedelta(days=i)
        iso = day.isoformat()
        # Skip past days (older than today) to avoid showing stale previous day as first entry
        if iso < today_iso:
            continue
        # SKIP FRIDAYS COMPLETELY - Don't include them in the window at all
        if day.weekday() == 4:
            continue
        day_doctors = []
        for d in data.get('doctors', []):
            per_date = {}
            if isinstance(d.get('schedule_by_date'), dict):
                per_date = d['schedule_by_date'].get(iso, {})
            if per_date:
                merged = dict(
                    id=d.get('id'), name=d.get('name'), specialty=d.get('specialty'),
                    designation=per_date.get('designation'), status=per_date.get('status','PENDING'),
                    start_time=per_date.get('start_time'), room=per_date.get('room'),
                    patient_count=per_date.get('patient_count'), opd=per_date.get('opd'),
                    breaks=per_date.get('breaks'), status_reason=per_date.get('status_reason'),
                    before_break_opd_patients=per_date.get('before_break_opd_patients'),
                    after_break_opd_patients=per_date.get('after_break_opd_patients'),
                    before_break_opd=per_date.get('before_break_opd'),
                    after_break_opd=per_date.get('after_break_opd'),
                    post_oncall=per_date.get('post_oncall'),
                    after_break_note=per_date.get('after_break_note'),
                    image_version=d.get('image_version',1)
                )
            else:
                # If no per-date schedule for this day:
                # Always show as clean PENDING (no carry-over) to prevent previous-day bleed-through
                merged = dict(
                    id=d.get('id'), name=d.get('name'), specialty=d.get('specialty'),
                    designation=None, status='PENDING', start_time=None, room=None,
                    patient_count=None, opd=[], breaks=[], status_reason=None,
                    before_break_opd=[], after_break_opd=[],
                    before_break_opd_patients=None, after_break_opd_patients=None,
                    after_break_note=None,
                    image_version=d.get('image_version',1)
                )
            day_doctors.append(merged)
        # Add day object (Fridays are already skipped above)
        window.append({'date': iso, 'doctors': day_doctors})
    return jsonify({'window': window})

@app.get('/api/day')
def api_day():
    """Return a single day snapshot for refresh button."""
    date_iso = request.args.get('date') or get_internet_today_iso()
    try:
        datetime.strptime(date_iso, '%Y-%m-%d')
    except Exception:
        date_iso = get_internet_today_iso()
    data = load_data()
    today_iso = get_internet_today_iso()
    # Important: Only return doctors that have an explicit per-date entry for date_iso.
    # This API is used by displays as an "override diff". Returning baseline values here
    # caused every doctor to appear as scheduled. Keep it strict to real overrides.
    docs = []
    for d in data.get('doctors', []):
        if not isinstance(d.get('schedule_by_date'), dict):
            continue
        per_date = d['schedule_by_date'].get(date_iso)
        if not per_date:
            continue
        merged = dict(id=d.get('id'), name=d.get('name'), specialty=d.get('specialty'),
                      designation=per_date.get('designation'), status=per_date.get('status','PENDING'),
                      start_time=per_date.get('start_time'), room=per_date.get('room'),
                      patient_count=per_date.get('patient_count'), opd=per_date.get('opd'),
                      breaks=per_date.get('breaks'), status_reason=per_date.get('status_reason'),
                      before_break_opd_patients=per_date.get('before_break_opd_patients'),
                      after_break_opd_patients=per_date.get('after_break_opd_patients'),
                      post_oncall=per_date.get('post_oncall'),
                      after_break_note=per_date.get('after_break_note'),
                      image_version=d.get('image_version',1))
        docs.append(merged)
    return jsonify({'date': date_iso, 'doctors': docs})

@app.post('/api/schedule/clear_status')
def api_schedule_clear_status():
    """Bulk clear or set statuses for a specific day or a rolling window.

    Request JSON:
      - date: YYYY-MM-DD (clears/sets only this day)
      - OR start: YYYY-MM-DD and days: int (default 14) to apply over window (Fridays skipped like /api/window)
      - action: 'set' (default) to set a single status and wipe other fields, or 'remove' to delete per-date overrides
      - set: desired status when action=='set' (default 'OFF_DUTY')

    Behavior when action=='set': for each doctor and each target date, keep only {'status': set}.
    """
    g = require_roles(ROLE_ADMIN, ROLE_MED)
    if g: return g
    payload = request.get_json(silent=True) or {}
    action = (payload.get('action') or 'set').strip().lower()
    set_status = (payload.get('set') or 'OFF_DUTY').strip().upper()

    # Determine target dates
    dates = []
    if payload.get('date'):
        try:
            datetime.strptime(payload['date'], '%Y-%m-%d')
            dates = [payload['date']]
        except Exception:
            return jsonify({'ok': False, 'error': 'bad date'}), 400
    else:
        start = payload.get('start') or get_internet_today_iso()
        try:
            base = datetime.strptime(start, '%Y-%m-%d').date()
        except Exception:
            base = datetime.strptime(get_internet_today_iso(), '%Y-%m-%d').date()
        days = max(1, min(30, int(payload.get('days', 14))))
        today_iso = get_internet_today_iso()
        max_iterations = days + (days // 4) + 5
        for i in range(max_iterations):
            if len(dates) >= days: break
            day = base + timedelta(days=i)
            iso = day.isoformat()
            if iso < today_iso: continue
            if day.weekday() == 4:  # skip Fridays to match window
                continue
            dates.append(iso)

    if not dates:
        return jsonify({'ok': False, 'error': 'no dates resolved'}), 400

    data = load_data()
    changed = 0
    for d in data.get('doctors', []):
        sched = d.setdefault('schedule_by_date', {})
        for iso in dates:
            if action == 'remove':
                if iso in sched:
                    try:
                        del sched[iso]
                        changed += 1
                    except Exception:
                        pass
            else:
                # Set only status; clear other fields for cleanliness
                entry = {'status': set_status}
                if sched.get(iso) != entry:
                    sched[iso] = entry
                    changed += 1
    save_data(data)
    try:
        broker.publish_event('bulk_schedule_clear', {'dates': dates, 'action': action, 'set': set_status})
    except Exception:
        pass
    return jsonify({'ok': True, 'changed': changed, 'dates': dates})

@app.get('/doctor-photo/<doc_id>')
def photo_alias(doc_id):
    """Backward compatible alias for template path /doctor-photo/<id>."""
    # Map id to file in uploads/doctor_photos/<id>.png (or .jpg etc.)
    # Attempt common extensions; fall back to default image served by static path
    for ext in ('.png','.jpg','.jpeg','.webp','.gif'):
        candidate = os.path.join(UPLOAD_DIR, f"{doc_id}{ext}")
        if os.path.exists(candidate):
            return send_file(candidate)
    return send_file(os.path.join('static','img','default-doctor.png'))

@app.post('/api_create_doctor')
def api_create_doctor():
    name = (request.json or {}).get('name') if request.is_json else (request.form.get('name'))
    if not name:
        return jsonify({'ok': False, 'error': 'name required'}), 400
    data = load_data()
    # generate simple id
    existing_ids = {d.get('id') for d in data['doctors']}
    n = 1
    while str(n) in existing_ids:
        n += 1
    doc = {'id': str(n), 'name': name, 'schedule_by_date': {}}
    data['doctors'].append(doc)
    save_data(data)
    broker.publish_event('doctor_create', {'doctor_id': doc['id']})
    return jsonify({'ok': True, 'doctor': doc})

@app.post('/api_update_doctor')
def api_update_doctor():
    payload = request.json if request.is_json else request.form.to_dict()
    if not payload:
        return jsonify({'ok': False, 'error': 'no payload'}), 400
    doc_id = payload.get('id') or payload.get('doctor_id')
    if not doc_id:
        return jsonify({'ok': False, 'error': 'id required'}), 400
    date_iso = payload.get('date') or get_internet_today_iso()
    data = load_data()
    doc = next((d for d in data['doctors'] if d.get('id') == doc_id), None)
    if not doc:
        return jsonify({'ok': False, 'error': 'not found'}), 404
    patch = {k: payload[k] for k in PER_DATE_FIELDS if k in payload}
    changed = False
    if patch:
        changed = apply_single_day_update(doc, date_iso, patch)
    # Non per-date fields (rename etc.)
    if 'name' in payload and payload['name'] != doc.get('name'):
        doc['name'] = payload['name']; changed = True
    if changed:
        save_data(data)
        broker.publish_event('doctor_update', {'doctor_id': doc_id, 'date': date_iso})
    return jsonify({'ok': True, 'changed': changed, 'doctor': doc})

# ============== PR PORTAL BACKEND APIs ==============

PR_STAFF_PATH = os.path.join('data', 'pr_staff.json')
PR_CLINICAL_ROSTER_PATH = os.path.join('data', 'pr_clinical_roster.json')

def load_pr_staff(auto_sync_specialties: bool = True):
    try:
        with open(PR_STAFF_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception:
        data = {"staff": [], "stations": {"clinical": [], "front": []}, "leave_types": {}}

    changed = ensure_station_structure(data)
    specialty_mutated = False
    if auto_sync_specialties:
        _, _, specialty_mutated = sync_doctor_specialty_stations(data)
    normalize_changed = normalize_pr_staff_records(data)

    if changed or specialty_mutated or normalize_changed:
        save_pr_staff(data)

    return data

def save_pr_staff(data):
    os.makedirs(os.path.dirname(PR_STAFF_PATH), exist_ok=True)
    with open(PR_STAFF_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def load_pr_clinical_roster():
    try:
        with open(PR_CLINICAL_ROSTER_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {"rosters": {}}

def save_pr_clinical_roster(data):
    os.makedirs(os.path.dirname(PR_CLINICAL_ROSTER_PATH), exist_ok=True)
    with open(PR_CLINICAL_ROSTER_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def slugify_station_name(name: str) -> str:
    base = re.sub(r'[^a-z0-9]+', '-', (name or '').lower()).strip('-')
    if not base:
        base = f'station-{int(time() * 1000)}'
    return base

def ensure_station_structure(pr_data: dict) -> bool:
    """Normalize station structure and inject default fields. Returns True if mutated."""
    changed = False
    stations = pr_data.setdefault('stations', {'clinical': [], 'front': []})
    if not isinstance(stations, dict):
        stations = {'clinical': [], 'front': []}
        pr_data['stations'] = stations
        changed = True

    for key in ('clinical', 'front'):
        if key not in stations or not isinstance(stations.get(key), list):
            stations[key] = []
            changed = True

    def _ensure_id(existing_ids: set, candidate: str) -> str:
        base = candidate or 'station'
        if base not in existing_ids:
            existing_ids.add(base)
            return base
        suffix = 1
        while f"{base}-{suffix}" in existing_ids:
            suffix += 1
        final_id = f"{base}-{suffix}"
        existing_ids.add(final_id)
        return final_id

    clinical_ids = {s.get('id') for s in stations['clinical'] if s.get('id')}
    front_ids = {s.get('id') for s in stations['front'] if s.get('id')}

    for station in stations['clinical']:
        original_id = station.get('id')
        if not original_id:
            station['id'] = _ensure_id(clinical_ids, slugify_station_name(station.get('name', 'clinical')))
            changed = True
        elif original_id not in clinical_ids:
            clinical_ids.add(original_id)
        station_type = station.get('type')
        if station_type != 'clinical':
            station['type'] = 'clinical'
            changed = True
        if 'color' not in station or not station.get('color'):
            station['color'] = '#10b981'
            changed = True
        if station.get('source') not in ('specialty', 'custom'):
            station['source'] = 'custom'
            changed = True
        auto_generated = station.get('source') == 'specialty'
        if station.get('auto_generated') != auto_generated:
            station['auto_generated'] = auto_generated
            changed = True
        if 'specialty' not in station and auto_generated:
            station['specialty'] = station.get('name', '')
            changed = True
        if 'allow_ai_assignment' not in station:
            station['allow_ai_assignment'] = True
            changed = True

    for station in stations['front']:
        original_id = station.get('id')
        if not original_id:
            station['id'] = _ensure_id(front_ids, slugify_station_name(station.get('name', 'front')))
            changed = True
        elif original_id not in front_ids:
            front_ids.add(original_id)
        station_type = station.get('type')
        if station_type != 'front':
            station['type'] = 'front'
            changed = True
        if 'color' not in station or not station.get('color'):
            station['color'] = '#0ea5e9'
            changed = True
        if station.get('source') not in ('custom', 'specialty'):
            station['source'] = 'custom'
            changed = True
        if 'auto_generated' not in station:
            station['auto_generated'] = False
            changed = True

    return changed

def sync_doctor_specialty_stations(pr_data: dict, doctor_data: Optional[dict] = None) -> tuple[list[str], int, bool]:
    """Ensure specialty-driven stations exist. Returns (specialties, added_count, mutated)."""
    stations = pr_data.setdefault('stations', {'clinical': [], 'front': []})
    doctor_data = doctor_data or load_data()
    specialties = sorted({(doc.get('specialty') or '').strip() for doc in doctor_data.get('doctors', []) if (doc.get('specialty') or '').strip()})

    existing_lookup = {}
    for station in stations['clinical']:
        key = (station.get('specialty') or station.get('name') or '').strip().lower()
        if key and key not in existing_lookup:
            existing_lookup[key] = station

    added = 0
    mutated = False
    existing_ids = {s.get('id') for s in stations['clinical'] if s.get('id')}
    for specialty in specialties:
        key = specialty.lower()
        if key in existing_lookup:
            # Ensure metadata stays aligned
            station = existing_lookup[key]
            if station.get('source') != 'specialty':
                station['source'] = 'specialty'
                station['auto_generated'] = True
                mutated = True
            if not station.get('specialty'):
                station['specialty'] = specialty
                mutated = True
            if 'allow_ai_assignment' not in station:
                station['allow_ai_assignment'] = True
                mutated = True
            continue

        station_id_base = slugify_station_name(specialty)
        station_id = station_id_base
        suffix = 1
        while station_id in existing_ids:
            station_id = f"{station_id_base}-{suffix}"
            suffix += 1
        existing_ids.add(station_id)

        new_station = {
            'id': station_id,
            'name': specialty,
            'type': 'clinical',
            'color': '#10b981',
            'source': 'specialty',
            'specialty': specialty,
            'allow_ai_assignment': True,
            'auto_generated': True
        }
        stations['clinical'].append(new_station)
        existing_lookup[key] = new_station
        added += 1
        mutated = True

    return specialties, added, mutated

def normalize_pr_staff_records(pr_data: dict) -> bool:
    """Normalize staff entries (roles, station IDs). Returns True if mutated."""
    changed = False
    staff_list = pr_data.get('staff') or []
    stations = pr_data.get('stations', {})

    station_lookup: dict[str, str] = {}
    for station in stations.get('clinical', []) + stations.get('front', []):
        station_id = station.get('id')
        station_name = (station.get('name') or '').strip()
        specialty = (station.get('specialty') or '').strip()
        for key in filter(None, [station_id, station_name, specialty]):
            key_lower = str(key).strip().lower()
            if key_lower:
                station_lookup[key_lower] = station_id
                station_lookup[key_lower.replace(' ', '-')] = station_id

    for staff in staff_list:
        roles = staff.get('roles')
        if roles is None:
            roles = []
        if isinstance(roles, str):
            roles = [roles]
            staff['roles'] = roles
            changed = True
        elif not isinstance(roles, list):
            roles = list(roles)
            staff['roles'] = roles
            changed = True
        if roles:
            if staff.get('role') != roles[0]:
                staff['role'] = roles[0]
                changed = True
        elif staff.get('role'):
            staff['roles'] = [staff['role']]
            changed = True

        stations_list = staff.get('stations') or []
        normalized: list[str] = []
        stations_changed = False
        for entry in stations_list:
            if entry in (None, ''):
                continue
            value = str(entry).strip()
            if not value:
                continue
            key_lower = value.lower()
            candidate = station_lookup.get(key_lower) or station_lookup.get(value.replace(' ', '-').lower())
            resolved = candidate or value
            if resolved not in normalized:
                normalized.append(resolved)
            if resolved != entry:
                stations_changed = True
        if stations_changed:
            staff['stations'] = normalized
            changed = True

    return changed

SHIFT_KNOWLEDGE_PATH = os.path.join('data', 'shift_knowledge.json')

def load_shift_knowledge():
    try:
        with open(SHIFT_KNOWLEDGE_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def save_shift_knowledge(data):
    os.makedirs(os.path.dirname(SHIFT_KNOWLEDGE_PATH), exist_ok=True)
    with open(SHIFT_KNOWLEDGE_PATH, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

@app.get('/pr-portal')
def pr_portal():
    if not session.get('user'):
        return redirect(url_for('login_page'))
    # Restrict to Admin and PR roles only (case-insensitive)
    user_role = (session['user'].get('role') or '').upper()
    if user_role not in ['ADMIN', 'PR']:
        return redirect(url_for('login_page'))
    return render_template('pr_portal_light.html')

@app.get('/pr-schedule-fullscreen')
@app.get('/pr_schedule_fullscreen')
def pr_schedule_fullscreen():
    if not session.get('user'):
        return redirect(url_for('login_page'))
    user_role = (session['user'].get('role') or '').upper()
    if user_role not in ['ADMIN', 'PR']:
        return redirect(url_for('login_page'))
    return render_template('pr_schedule_fullscreen.html')

@app.get('/api/shift-knowledge')
def api_shift_knowledge_get():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_shift_knowledge()
    return jsonify({'ok': True, 'knowledge': data})

@app.post('/api/shift-knowledge')
def api_shift_knowledge_save():
    user_role = (session.get('user', {}).get('role') or '').upper()
    if not session.get('user') or user_role != 'ADMIN':
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    knowledge = payload.get('knowledge', {})
    if not isinstance(knowledge, dict):
        return jsonify({'ok': False, 'error': 'invalid payload'}), 400
    save_shift_knowledge(knowledge)
    return jsonify({'ok': True})

@app.post('/api/pr/station-custom-names')
def api_pr_station_custom_names():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    mapping = payload.get('station_custom_names', {})
    if not isinstance(mapping, dict):
        return jsonify({'ok': False, 'error': 'invalid payload'}), 400
    data = load_pr_staff()
    data['station_custom_names'] = mapping
    save_pr_staff(data)
    return jsonify({'ok': True})

# Staff Management APIs
@app.get('/api/pr/staff')
def api_pr_staff_list():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_pr_staff()
    return jsonify({'ok': True, 'staff': data.get('staff', [])})

@app.post('/api/pr/staff')
def api_pr_staff_create():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    name = payload.get('name', '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'name required'}), 400
    
    data = load_pr_staff()
    if 'staff' not in data:
        data['staff'] = []
    
    staff_id = str(int(time() * 1000))
    roles = payload.get('roles', [])
    new_staff = {
        'id': staff_id,
        'name': name,
        'roles': roles,  # ['clinical', 'front', 'training']
        'stations': payload.get('stations', []),  # list of station names
        'active': True,
        'role': roles[0] if roles else payload.get('role', ''),
        'created_at': datetime.now().isoformat()
    }
    data['staff'].append(new_staff)
    save_pr_staff(data)
    return jsonify({'ok': True, 'staff': new_staff})

@app.put('/api/pr/staff/<staff_id>')
def api_pr_staff_update(staff_id):
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    data = load_pr_staff()
    staff = next((s for s in data['staff'] if s['id'] == staff_id), None)
    if not staff:
        return jsonify({'ok': False, 'error': 'staff not found'}), 404
    
    if 'name' in payload:
        staff['name'] = payload['name']
    if 'roles' in payload:
        staff['roles'] = payload['roles']
        if payload['roles']:
            staff['role'] = payload['roles'][0]
    if 'stations' in payload:
        staff['stations'] = payload['stations']
    if 'active' in payload:
        staff['active'] = payload['active']
    
    save_pr_staff(data)
    return jsonify({'ok': True, 'staff': staff})

@app.delete('/api/pr/staff/<staff_id>')
def api_pr_staff_delete(staff_id):
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_pr_staff()
    data['staff'] = [s for s in data['staff'] if s['id'] != staff_id]
    save_pr_staff(data)
    return jsonify({'ok': True})

# Station/Knowledge Management APIs
@app.get('/api/pr/stations')
def api_pr_stations_list():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_pr_staff()
    changed = ensure_station_structure(data)
    doctor_data = load_data()
    specialties, added, specialty_mutated = sync_doctor_specialty_stations(data, doctor_data)

    if data.get('stations'):
        data['stations']['clinical'] = sorted(data['stations'].get('clinical', []), key=lambda s: (s.get('name') or '').lower())
        data['stations']['front'] = sorted(data['stations'].get('front', []), key=lambda s: (s.get('name') or '').lower())

    if changed or specialty_mutated:
        save_pr_staff(data)

    return jsonify({
        'ok': True,
        'stations': data.get('stations', {'clinical': [], 'front': []}),
        'station_custom_names': data.get('station_custom_names', {'clinical': {}, 'front': {}}),
        'doctor_specialties': specialties,
        'auto_specialties_added': added
    })

@app.post('/api/pr/stations')
def api_pr_stations_create():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    bulk = payload.get('stations')
    data = load_pr_staff()

    if bulk is not None:
        if not isinstance(bulk, dict):
            return jsonify({'ok': False, 'error': 'invalid stations payload'}), 400

        sanitized = {'clinical': [], 'front': []}

        def _sanitize_list(items, stype):
            result = []
            seen_ids: set[str] = set()
            for item in items or []:
                name = (item.get('name') or '').strip()
                if not name:
                    continue
                station_id = (item.get('id') or '').strip()
                if not station_id:
                    station_id = slugify_station_name(name)
                if station_id in seen_ids:
                    suffix = 1
                    base_id = station_id
                    while f"{base_id}-{suffix}" in seen_ids:
                        suffix += 1
                    station_id = f"{base_id}-{suffix}"
                seen_ids.add(station_id)

                station = {
                    'id': station_id,
                    'name': name,
                    'type': stype,
                    'color': item.get('color') or ('#10b981' if stype == 'clinical' else '#0ea5e9'),
                    'source': item.get('source') if item.get('source') in ('specialty', 'custom') else 'custom'
                }
                if stype == 'clinical':
                    station['specialty'] = (item.get('specialty') or name).strip()
                    station['allow_ai_assignment'] = bool(item.get('allow_ai_assignment', True))
                    station['auto_generated'] = bool(item.get('auto_generated', station['source'] == 'specialty'))
                else:
                    station['auto_generated'] = bool(item.get('auto_generated', False))
                result.append(station)
            return result

        sanitized['clinical'] = _sanitize_list(bulk.get('clinical'), 'clinical')
        sanitized['front'] = _sanitize_list(bulk.get('front'), 'front')

        data['stations'] = sanitized
        changed = ensure_station_structure(data)
        specialties, added, specialty_mutated = sync_doctor_specialty_stations(data)
        if data.get('stations'):
            data['stations']['clinical'] = sorted(data['stations'].get('clinical', []), key=lambda s: (s.get('name') or '').lower())
            data['stations']['front'] = sorted(data['stations'].get('front', []), key=lambda s: (s.get('name') or '').lower())
        save_pr_staff(data)
        return jsonify({'ok': True, 'stations': data['stations'], 'doctor_specialties': specialties, 'auto_specialties_added': added, 'mutated': changed or specialty_mutated})

    station_type = payload.get('type', 'clinical')  # 'clinical' or 'front'
    name = payload.get('name', '').strip()

    if not name or station_type not in ['clinical', 'front']:
        return jsonify({'ok': False, 'error': 'invalid input'}), 400

    if 'stations' not in data:
        data['stations'] = {'clinical': [], 'front': []}

    station_id = slugify_station_name(name)
    existing_ids = {s.get('id') for s in data['stations'].get(station_type, [])}
    suffix = 1
    original_id = station_id
    while station_id in existing_ids:
        station_id = f"{original_id}-{suffix}"
        suffix += 1

    new_station = {
        'id': station_id,
        'name': name,
        'type': station_type,
        'color': payload.get('color', '#10b981' if station_type == 'clinical' else '#0ea5e9'),
        'source': 'custom',
        'auto_generated': False
    }
    if station_type == 'clinical':
        new_station['specialty'] = name
        new_station['allow_ai_assignment'] = bool(payload.get('allow_ai_assignment', True))

    data['stations'][station_type].append(new_station)
    ensure_station_structure(data)
    specialties, added, _ = sync_doctor_specialty_stations(data)
    save_pr_staff(data)
    return jsonify({'ok': True, 'station': new_station, 'doctor_specialties': specialties, 'auto_specialties_added': added})

@app.delete('/api/pr/stations/<station_id>')
def api_pr_stations_delete(station_id):
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_pr_staff()
    stations = data.get('stations', {})
    changed = False
    for stype in ['clinical', 'front']:
        original = stations.get(stype, [])
        filtered = [s for s in original if s.get('id') != station_id]
        if len(filtered) != len(original):
            stations[stype] = filtered
            changed = True
    if changed:
        data['stations'] = stations
        ensure_station_structure(data)
        save_pr_staff(data)
    return jsonify({'ok': True})

@app.post('/api/pr/stations/sync')
def api_pr_stations_sync():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_pr_staff()
    changed = ensure_station_structure(data)
    specialties, added, specialty_mutated = sync_doctor_specialty_stations(data)
    if data.get('stations'):
        data['stations']['clinical'] = sorted(data['stations'].get('clinical', []), key=lambda s: (s.get('name') or '').lower())
        data['stations']['front'] = sorted(data['stations'].get('front', []), key=lambda s: (s.get('name') or '').lower())
    if changed or specialty_mutated:
        save_pr_staff(data)
    return jsonify({
        'ok': True,
        'stations': data.get('stations', {'clinical': [], 'front': []}),
        'doctor_specialties': specialties,
        'added': added
    })

# Clinical Roster APIs
@app.get('/api/pr/roster/clinical')
def api_pr_roster_clinical_get():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    data = load_pr_clinical_roster()
    rosters = data.get('rosters', {})
    
    # Filter by date range if provided
    if start_date and end_date:
        filtered = {k: v for k, v in rosters.items() if start_date <= k <= end_date}
        return jsonify({'ok': True, 'rosters': filtered})
    
    return jsonify({'ok': True, 'rosters': rosters})

@app.post('/api/pr/roster/clinical')
def api_pr_roster_clinical_save():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    date = payload.get('date')
    staff_id = payload.get('staff_id')
    duty_type = payload.get('duty_type')  # 'duty', 'leave', 'off'
    station_id = payload.get('station_id')  # for duty assignments
    leave_type = payload.get('leave_type')  # AL, FRL, etc.
    
    if not date or not staff_id:
        return jsonify({'ok': False, 'error': 'date and staff_id required'}), 400
    
    data = load_pr_clinical_roster()
    if 'rosters' not in data:
        data['rosters'] = {}
    if date not in data['rosters']:
        data['rosters'][date] = {}
    
    data['rosters'][date][staff_id] = {
        'duty_type': duty_type,
        'station_id': station_id,
        'leave_type': leave_type,
        'updated_at': datetime.now().isoformat()
    }
    
    save_pr_clinical_roster(data)
    return jsonify({'ok': True, 'roster': data['rosters'][date]})

@app.post('/api/pr/roster/clinical/generate')
def api_pr_roster_clinical_generate():
    """Enhanced AI-based duty generation algorithm with smart balancing"""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    
    payload = request.json or {}
    start_date = payload.get('start_date')
    end_date = payload.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({'ok': False, 'error': 'start_date and end_date required'}), 400
    
    # Load doctor schedules and PR staff
    doctor_data = load_data()
    pr_data = load_pr_staff()
    roster_data = load_pr_clinical_roster()
    
    # Get active clinical staff
    clinical_staff = [s for s in pr_data.get('staff', []) if 'clinical' in s.get('roles', []) and s.get('active', True)]
    
    if not clinical_staff:
        return jsonify({'ok': False, 'error': 'no active clinical staff'}), 400
    
    # Generate dates
    start = datetime.strptime(start_date, '%Y-%m-%d').date()
    end = datetime.strptime(end_date, '%Y-%m-%d').date()
    current = start
    dates = []
    while current <= end:
        dates.append(current.isoformat())
        current += timedelta(days=1)
    
    # Enhanced AI Algorithm with smart workload balancing
    import random
    generated_rosters = {}
    
    # Track weekly stats per staff (resets each week)
    staff_weekly_duty_count = {s['id']: 0 for s in clinical_staff}
    staff_total_duty_count = {s['id']: 0 for s in clinical_staff}
    staff_consecutive_days = {s['id']: 0 for s in clinical_staff}
    staff_last_specialty = {s['id']: None for s in clinical_staff}
    
    # Check existing leave assignments
    existing_rosters = roster_data.get('rosters', {})
    
    for date_str in dates:
        # Get doctors scheduled for this date
        doctors_on_duty = []
        for doc in doctor_data.get('doctors', []):
            # Check both 'schedule' and 'schedule_by_date' for compatibility
            schedule = None
            if 'schedule_by_date' in doc:
                schedule = doc['schedule_by_date'].get(date_str, {})
            elif 'schedule' in doc:
                schedule = doc['schedule'].get(date_str, {})
            
            if schedule:
                status = schedule.get('status', 'MISSING')
                if status == 'ON_DUTY':
                    specialty = doc.get('specialty', 'General')
                    doctors_on_duty.append({
                        'name': doc['name'], 
                        'specialty': specialty,
                        'room': schedule.get('room', 'N/A')
                    })
        
        # Get unique specialties for today
        specialties_today = list(set([d['specialty'] for d in doctors_on_duty]))
        
        generated_rosters[date_str] = {}
        
        # Check for existing leave requests
        if date_str in existing_rosters:
            for staff_id, assignment in existing_rosters[date_str].items():
                if assignment.get('duty_type') == 'leave':
                    generated_rosters[date_str][staff_id] = assignment
                    continue
        
        # Get available staff (not on leave, not exceeding 5 days/week)
        available_staff = []
        for s in clinical_staff:
            # Skip if on leave
            if s['id'] in generated_rosters[date_str]:
                continue
            # Skip if already worked 5 days this week
            if staff_weekly_duty_count[s['id']] >= 5:
                continue
            # Skip if worked 3 consecutive days (needs a break)
            if staff_consecutive_days[s['id']] >= 3:
                continue
            available_staff.append(s)
        
        if len(doctors_on_duty) > 0 and len(available_staff) > 0:
            # Calculate how many staff needed (match doctor count)
            num_assignments = min(len(doctors_on_duty), len(available_staff))
            
            # Sort staff by total duty count (assign least worked first for fairness)
            available_staff.sort(key=lambda x: (staff_total_duty_count[x['id']], random.random()))
            
            assigned_specialties = []
            for i in range(num_assignments):
                staff = available_staff[i]
                
                # Smart specialty assignment: avoid same specialty consecutively
                available_specialties = [s for s in specialties_today if s not in assigned_specialties]
                if not available_specialties:
                    available_specialties = specialties_today
                
                # Prefer different specialty than last time
                last_spec = staff_last_specialty[staff['id']]
                if last_spec and last_spec in available_specialties and len(available_specialties) > 1:
                    available_specialties = [s for s in available_specialties if s != last_spec]
                
                assigned_specialty = random.choice(available_specialties)
                assigned_specialties.append(assigned_specialty)
                
                generated_rosters[date_str][staff['id']] = {
                    'duty_type': 'duty',
                    'station_id': assigned_specialty,
                    'leave_type': None,
                    'updated_at': datetime.now().isoformat()
                }
                
                # Update counters
                staff_weekly_duty_count[staff['id']] += 1
                staff_total_duty_count[staff['id']] += 1
                staff_consecutive_days[staff['id']] += 1
                staff_last_specialty[staff['id']] = assigned_specialty
        
        # Mark off days for staff not assigned
        for s in clinical_staff:
            if s['id'] not in generated_rosters[date_str]:
                # Reset consecutive counter when off
                staff_consecutive_days[s['id']] = 0
                # Only mark as "off" if not already marked as leave
                generated_rosters[date_str][s['id']] = {
                    'duty_type': 'off',
                    'station_id': None,
                    'leave_type': None,
                    'updated_at': datetime.now().isoformat()
                }
        
        # Reset weekly counters on Mondays
        if datetime.strptime(date_str, '%Y-%m-%d').weekday() == 0:  # Monday
            staff_weekly_duty_count = {s['id']: 0 for s in clinical_staff}
    
    # Save generated rosters
    if 'rosters' not in roster_data:
        roster_data['rosters'] = {}
    roster_data['rosters'].update(generated_rosters)
    save_pr_clinical_roster(roster_data)
    
    return jsonify({
        'ok': True,
        'generated': len(generated_rosters),
        'rosters': generated_rosters,
        'stats': {
            'dates': len(dates),
            'staff': len(clinical_staff),
            'assignments': sum(len(r) for r in generated_rosters.values()),
            'doctors_per_day': {d: len([doc for doc in doctor_data.get('doctors', []) if (doc.get('schedule_by_date', {}).get(d, {}).get('status') or doc.get('schedule', {}).get(d, {}).get('status')) == 'ON_DUTY']) for d in dates[:5]}
        }
    })

# GOPD Configuration APIs
@app.get('/api/pr/gopd-config')
def api_pr_gopd_config_get():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_pr_staff()
    # Support both old and new format
    gopd_config = data.get('gopd_config', {'staff_count': 2, 'shifts': ['Morning', 'Afternoon']})
    return jsonify({'ok': True, 'gopd_config': gopd_config})

@app.post('/api/pr/gopd-config')
def api_pr_gopd_config_save():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    data = load_pr_staff()
    
    # NEW: Support categorized format
    if 'categories' in payload:
        data['gopd_config'] = {
            'staff_count': payload.get('staff_count', 2),
            'categories': payload.get('categories', {})
        }
    else:
        # OLD: Backward compatibility with flat shifts
        data['gopd_config'] = {
            'staff_count': payload.get('staff_count', 2),
            'shifts': payload.get('shifts', ['Morning', 'Afternoon'])
        }
    
    save_pr_staff(data)
    return jsonify({'ok': True, 'gopd_config': data['gopd_config']})

# AI Rules Configuration APIs
@app.get('/api/pr/ai-rules')
def api_pr_ai_rules_get():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_pr_staff()
    # Default AI rules if not set
    ai_rules = data.get('ai_rules', {
        'minWeeklyOffs': 2,
        'maxConsecutiveDays': 6,
        'allowEDAssignment': True,
        'maxEDPerWeek': 1,
        'noRepeatedED': True,
        'noEDAfterOff': True,
        'noEDOnFriday': False,
        'staffStartOffset': 30,
        'minBreakHours': 12,
        'enableOnCallAdjust': True,
        'balanceWorkload': True,
        'respectPreferences': True
    })
    return jsonify(ai_rules)

@app.post('/api/pr/ai-rules')
def api_pr_ai_rules_save():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    data = load_pr_staff()
    
    # Save AI rules
    data['ai_rules'] = {
        'minWeeklyOffs': int(payload.get('minWeeklyOffs', 2)),
        'maxConsecutiveDays': int(payload.get('maxConsecutiveDays', 6)),
        'allowEDAssignment': bool(payload.get('allowEDAssignment', True)),
        'maxEDPerWeek': int(payload.get('maxEDPerWeek', 1)),
        'noRepeatedED': bool(payload.get('noRepeatedED', True)),
        'noEDAfterOff': bool(payload.get('noEDAfterOff', True)),
        'noEDOnFriday': bool(payload.get('noEDOnFriday', False)),
        'staffStartOffset': int(payload.get('staffStartOffset', 30)),
        'minBreakHours': int(payload.get('minBreakHours', 12)),
        'enableOnCallAdjust': bool(payload.get('enableOnCallAdjust', True)),
        'balanceWorkload': bool(payload.get('balanceWorkload', True)),
        'respectPreferences': bool(payload.get('respectPreferences', True))
    }
    
    save_pr_staff(data)
    return jsonify({'ok': True, 'ai_rules': data['ai_rules']})

# Staff Shift Templates with Categories
@app.get('/api/shift-templates')
def api_shift_templates_get():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = load_pr_staff()
    # Support both old and new format
    templates = data.get('shift_templates', {
        'clinical': {'Morning Shifts': [], 'Evening Shifts': [], 'Night Shifts': []},
        'front': {'Morning Shifts': [], 'Evening Shifts': [], 'Night Shifts': []}
    })
    return jsonify({'ok': True, 'templates': templates})

@app.post('/api/shift-templates')
def api_shift_templates_save():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    data = load_pr_staff()
    
    # NEW: Support categorized format {clinical: {Morning: [...], Evening: [...]}, front: {...}}
    # OLD: Also supports flat format {clinical: [...], front: [...]} for backward compatibility
    data['shift_templates'] = payload
    
    save_pr_staff(data)
    return jsonify({'ok': True, 'templates': data['shift_templates']})

# Month View APIs
@app.get('/api/pr/roster/month')
def api_pr_roster_month_get():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    
    if not year or not month:
        return jsonify({'ok': False, 'error': 'year and month required'}), 400
    
    # Generate all dates in the month
    from calendar import monthrange
    num_days = monthrange(year, month)[1]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{num_days:02d}"
    
    data = load_pr_clinical_roster()
    rosters = data.get('rosters', {})
    filtered = {k: v for k, v in rosters.items() if start_date <= k <= end_date}
    
    return jsonify({'ok': True, 'rosters': filtered, 'year': year, 'month': month, 'days': num_days})

# Bulk Roster Update API
@app.post('/api/pr/roster/bulk-update')
def api_pr_roster_bulk_update():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    updates = payload.get('updates', [])  # List of {date, staff_id, duty_type, station_id, leave_type}
    
    data = load_pr_clinical_roster()
    if 'rosters' not in data:
        data['rosters'] = {}
    
    updated_count = 0
    for update in updates:
        date = update.get('date')
        staff_id = update.get('staff_id')
        if not date or not staff_id:
            continue
        
        if date not in data['rosters']:
            data['rosters'][date] = {}
        
        data['rosters'][date][staff_id] = {
            'duty_type': update.get('duty_type'),
            'station_id': update.get('station_id'),
            'leave_type': update.get('leave_type'),
            'updated_at': datetime.now().isoformat()
        }
        updated_count += 1
    
    save_pr_clinical_roster(data)
    return jsonify({'ok': True, 'updated': updated_count})

# Clear Staff Schedule API
@app.post('/api/pr/roster/clear-staff')
def api_pr_roster_clear_staff():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    payload = request.json or {}
    staff_id = payload.get('staff_id')
    start_date = payload.get('start_date')
    end_date = payload.get('end_date')
    
    if not staff_id:
        return jsonify({'ok': False, 'error': 'staff_id required'}), 400
    
    data = load_pr_clinical_roster()
    rosters = data.get('rosters', {})
    cleared_count = 0
    
    for date_str in list(rosters.keys()):
        if start_date and end_date:
            if not (start_date <= date_str <= end_date):
                continue
        if staff_id in rosters[date_str]:
            del rosters[date_str][staff_id]
            cleared_count += 1
    
    save_pr_clinical_roster(data)
    return jsonify({'ok': True, 'cleared': cleared_count})

@app.post('/api/pr/roster/clear')
def api_pr_roster_clear():
    """Clear roster data - specific day, specific staff, or all"""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    user_role = (session.get('user', {}).get('role') or '').upper()
    if user_role != 'ADMIN':
        return jsonify({'ok': False, 'error': 'admin only'}), 403
    
    try:
        payload = request.json or {}
        clear_type = payload.get('type')
        
        if not clear_type:
            return jsonify({'ok': False, 'error': 'type required (day, staff, or all)'}), 400
        
        roster_path = 'data/pr_roster.json'
        
        # Load current roster
        if os.path.exists(roster_path):
            with open(roster_path, 'r', encoding='utf-8') as f:
                roster = json.load(f)
        else:
            roster = {}
        
        if clear_type == 'day':
            # Clear specific day for all staff
            date = payload.get('date')
            if not date:
                return jsonify({'ok': False, 'error': 'date required'}), 400
            
            cleared_count = 0
            for staff_id in list(roster.keys()):
                if date in roster[staff_id]:
                    del roster[staff_id][date]
                    cleared_count += 1
            
            message = f'Cleared {cleared_count} assignments for {date}'
            
        elif clear_type == 'staff':
            # Clear specific staff member completely
            staff_id = payload.get('staff_id')
            if not staff_id:
                return jsonify({'ok': False, 'error': 'staff_id required'}), 400
            
            if staff_id in roster:
                days_cleared = len(roster[staff_id])
                del roster[staff_id]
                message = f'Cleared {days_cleared} days for staff member'
            else:
                message = 'No data found for this staff member'
                
        elif clear_type == 'all':
            # Clear everything
            days_count = sum(len(dates) for dates in roster.values())
            roster = {}
            message = f'Cleared all roster data ({days_count} total assignments)'
            
        else:
            return jsonify({'ok': False, 'error': 'Invalid type. Use: day, staff, or all'}), 400
        
        # Save updated roster
        with open(roster_path, 'w', encoding='utf-8') as f:
            json.dump(roster, f, indent=2, ensure_ascii=False)
        
        return jsonify({'ok': True, 'message': message})
        
    except Exception as e:
        print(f'Error clearing roster: {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500

# Export to Excel API
@app.get('/api/pr/roster/export-excel')
def api_pr_roster_export_excel():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({'ok': False, 'error': 'date range required'}), 400
    
    # Load data
    roster_data = load_pr_clinical_roster()
    pr_data = load_pr_staff()
    rosters = roster_data.get('rosters', {})
    staff_list = pr_data.get('staff', [])
    
    # Generate dates
    dates = []
    current = datetime.strptime(start_date, '%Y-%m-%d').date()
    end = datetime.strptime(end_date, '%Y-%m-%d').date()
    while current <= end:
        dates.append(current.isoformat())
        current += timedelta(days=1)
    
    # Create CSV
    output = BytesIO()
    import csv
    writer = csv.writer(output)
    
    # Header row
    header = ['Staff Name'] + [datetime.strptime(d, '%Y-%m-%d').strftime('%d-%b') for d in dates]
    writer.writerow(header)
    
    # Data rows
    for staff in staff_list:
        if 'clinical' not in staff.get('roles', []):
            continue
        row = [staff['name']]
        for date in dates:
            assignment = rosters.get(date, {}).get(staff['id'], {})
            duty_type = assignment.get('duty_type', '-')
            if duty_type == 'duty':
                cell = assignment.get('station_id', 'DUTY')
            elif duty_type == 'leave':
                cell = assignment.get('leave_type', 'LEAVE')
            elif duty_type == 'off':
                cell = 'OFF'
            else:
                cell = '-'
            row.append(cell)
        writer.writerow(row)
    
    output.seek(0)
    return send_file(
        output,
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'pr_roster_{start_date}_to_{end_date}.csv'
    )

# PR Telegram Bot Configuration
PR_TELEGRAM_TOKEN = _clean_env(os.environ.get('PR_TELEGRAM_BOT_TOKEN') or '')
PR_TELEGRAM_CHAT_ID = _clean_env(os.environ.get('PR_TELEGRAM_GROUP_ID') or '')

# Server Sent Events endpoint
@app.get('/events')
def sse_events():
    def gen():
        q = broker.subscribe()
        heartbeat_interval = 15  # seconds
        last_beat = time()
        while True:
            # Flush all queued events first for low latency
            flushed = False
            while q:
                ev, data_json = q.popleft()
                yield f"event: {ev}\ndata: {data_json}\n\n"
                flushed = True
            now = time()
            # Heartbeat comment (ignored by EventSource but keeps connection open across proxies)
            if now - last_beat >= heartbeat_interval:
                yield f": ping {int(now)}\n\n"
                last_beat = now
            # When idle, slight sleep to avoid tight loop
            if not flushed:
                sleep(0.5)
    headers = {
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no'  # for nginx disabling buffering if ever used
    }
    return Response(stream_with_context(gen()), mimetype='text/event-stream', headers=headers)

# Static file for uploaded photos
@app.get('/uploads/doctor_photos/<path:fname>')
def serve_doctor_photo(fname):
    return send_from_directory(UPLOAD_DIR, fname)

# Ensure telegram thread starts when app context ready
@app.before_request
def _ensure_telegram():
    maybe_start_telegram()

# --- END appended helpers ---

# ========== PR PORTAL ADDITIONAL APIs ==========

# PR Roster Management (All Staff - Excel View 16th-15th)
@app.get('/api/pr/roster/all')
def api_pr_roster_all():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    roster_path = 'data/pr_roster.json'
    if os.path.exists(roster_path):
        with open(roster_path, 'r', encoding='utf-8') as f:
            roster = json.load(f)
    else:
        roster = {}
    return jsonify(roster)

@app.post('/api/pr/generate-roster')
def api_pr_generate_roster():
    """Generate PR staff roster based on doctor schedules with comprehensive fallback"""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    user_role = (session.get('user', {}).get('role') or '').upper()
    if user_role != 'ADMIN':
        return jsonify({'ok': False, 'error': 'admin only'}), 403
    
    try:
        data = request.get_json()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'ok': False, 'error': 'start_date and end_date required'}), 400
        
        from datetime import datetime, timedelta
        
        # Load doctor schedules from multiple sources
        doctor_schedules_by_date = {}
        
        # Source 1: PR clinical roster
        clinical_roster_path = 'data/pr_clinical_roster.json'
        if os.path.exists(clinical_roster_path):
            with open(clinical_roster_path, 'r', encoding='utf-8') as f:
                clinical_data = json.load(f)
                # Handle both formats: direct dates or nested rosters
                if 'rosters' in clinical_data:
                    doctor_schedules_by_date = clinical_data.get('rosters', {})
                else:
                    doctor_schedules_by_date = clinical_data
        
        # Source 2: Main doctors.json (for active ON_DUTY schedules)
        doctors_path = 'data/doctors.json'
        if os.path.exists(doctors_path):
            with open(doctors_path, 'r', encoding='utf-8') as f:
                doctors_data = json.load(f)
                doctors_list = doctors_data.get('doctors', [])
                
                # Extract schedules from doctor records
                start = datetime.strptime(start_date, '%Y-%m-%d')
                end = datetime.strptime(end_date, '%Y-%m-%d')
                current = start
                
                while current <= end:
                    date_str = current.strftime('%Y-%m-%d')
                    if date_str not in doctor_schedules_by_date:
                        doctor_schedules_by_date[date_str] = {}
                    
                    # Check each doctor's schedule for this date
                    for doctor in doctors_list:
                        schedule_by_date = doctor.get('schedule_by_date', {})
                        if date_str in schedule_by_date:
                            day_schedule = schedule_by_date[date_str]
                            if day_schedule.get('status') == 'ON_DUTY':
                                specialty = doctor.get('specialty', '')
                                # Create schedule entry
                                if not isinstance(doctor_schedules_by_date[date_str], dict):
                                    doctor_schedules_by_date[date_str] = {}
                                
                                # Use doctor ID as key
                                doctor_schedules_by_date[date_str][str(doctor['id'])] = {
                                    'duty_type': 'duty',
                                    'station_id': specialty,
                                    'specialty': specialty,
                                    'leave_type': '',
                                    'start_time': day_schedule.get('start_time', '08:00'),
                                    'name': doctor.get('name', ''),
                                    'opd': day_schedule.get('opd', [])
                                }
                    
                    current += timedelta(days=1)
        
        # Load PR staff & stations
        pr_staff = load_pr_staff()
        ensure_changed = ensure_station_structure(pr_staff)
        _, _, specialty_mutated = sync_doctor_specialty_stations(pr_staff)
        if ensure_changed or specialty_mutated:
            save_pr_staff(pr_staff)

        station_config = pr_staff.get('stations', {'clinical': [], 'front': []})
        staff_records = pr_staff.get('staff', [])
        shift_templates = pr_staff.get('shift_templates', {})
        opdinfo = _load_opdinfo()
        
        # Helper functions
        def has_role(staff_member, role_name: str) -> bool:
            roles = staff_member.get('roles')
            if isinstance(roles, list):
                return role_name in roles
            return (staff_member.get('role') or '').lower() == role_name

        def time_to_minutes(time_str):
            try:
                if not time_str:
                    return 0
                hours, mins = time_str.split(':')
                return int(hours) * 60 + int(mins)
            except:
                return 0
        
        def find_matching_clinical_shift(doctor_opd_start, clinical_shifts):
            """Find clinical shift that starts 30-90 min before doctor OPD"""
            doctor_minutes = time_to_minutes(doctor_opd_start)
            best_match = None
            best_diff = 999
            
            for shift in clinical_shifts:
                shift_minutes = time_to_minutes(shift.get('start', ''))
                diff = doctor_minutes - shift_minutes
                
                # Clinical should start 30-90 min before doctor
                if 30 <= diff <= 90:
                    if abs(diff - 30) < best_diff:  # Prefer 30 min offset
                        best_diff = abs(diff - 30)
                        best_match = shift
            
            return best_match or (clinical_shifts[0] if clinical_shifts else None)
        
        # Get active staff
        clinical_staff = [s for s in staff_records if s.get('active', True) and has_role(s, 'clinical')]
        front_staff = [s for s in staff_records if s.get('active', True) and has_role(s, 'front')]
        
        if not clinical_staff and not front_staff:
            return jsonify({'ok': False, 'error': 'No active staff members found'}), 400
        
        # Build clinical shift pool
        clinical_shifts = []
        clinical_team_templates = shift_templates.get('clinical', {})
        for category_name, category_shifts in clinical_team_templates.items():
            if isinstance(category_shifts, list):
                clinical_shifts.extend(category_shifts)
        
        # Build doctor-to-clinical mapping
        doctor_clinical_mapping = {}
        for specialty, opd_data in opdinfo.items():
            if not isinstance(opd_data, dict):
                continue
            doctor_clinical_mapping[specialty] = {}
            for shift_key in ['shift1', 'shift2']:
                shift_info = opd_data.get(shift_key, {})
                if not shift_info or not isinstance(shift_info, dict):
                    continue
                opd_start = shift_info.get('start', '')
                if opd_start:
                    matching_shift = find_matching_clinical_shift(opd_start, clinical_shifts)
                    if matching_shift:
                        doctor_clinical_mapping[specialty][opd_start] = matching_shift
        
        # Generate roster
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        generated = {}
        
        clinical_station_pool = [s for s in station_config.get('clinical', []) if s.get('allow_ai_assignment', True)]
        clinical_station_lookup = {
            (s.get('specialty') or s.get('name') or '').strip().lower(): s
            for s in clinical_station_pool
            if (s.get('specialty') or s.get('name'))
        }
        
        # Pool-based day-off distribution for fair rotation
        # Track consecutive work days and per-station counts for each staff
        staff_work_days = {s['id']: 0 for s in staff_records if s.get('active', True)}
        staff_station_counts = {s['id']: {} for s in staff_records if s.get('active', True)}  # {staff_id: {station_id: count}}
        
        current_date = start
        clinical_staff_rotation = {s.get('id'): 0 for s in clinical_staff}  # Track rotation index
        front_staff_rotation = {s.get('id'): 0 for s in front_staff}  # Track rotation index
        
        stats = {
            'total_days': 0,
            'days_with_doctors': 0,
            'days_without_doctors': 0,
            'clinical_assigned': 0,
            'front_assigned': 0,
            'days_off_distributed': 0
        }
        
        while current_date <= end:
            date_str = current_date.strftime('%Y-%m-%d')
            stats['total_days'] += 1
            
            # Get doctor schedules for this date
            date_doctor_schedules = doctor_schedules_by_date.get(date_str, {})
            has_doctors = bool(date_doctor_schedules and len(date_doctor_schedules) > 0)
            
            if has_doctors:
                stats['days_with_doctors'] += 1
            else:
                stats['days_without_doctors'] += 1
            
            # ONLY assign clinical staff when doctors are present
            if clinical_staff and has_doctors:
                # Calculate day of week for smart day-off distribution
                day_of_week = current_date.weekday()  # 0=Monday, 6=Sunday
                days_since_start = (current_date - start).days
                
                # Rotate which pool gets priority day off (changes every 2-3 days)
                day_off_pool = ['A', 'B', 'C', 'D'][(days_since_start // 2) % 4]
                
                # Group doctors by specialty to assign multiple staff if needed
                specialty_assignments = {}
                for doctor_id, schedule in date_doctor_schedules.items():
                    specialty = schedule.get('specialty') or schedule.get('station_id', '')
                    specialty_key = specialty.strip().lower()
                    
                    if specialty_key not in specialty_assignments:
                        specialty_assignments[specialty_key] = []
                    specialty_assignments[specialty_key].append(schedule)
                
                # Assign staff for each specialty
                for specialty_key, doctor_schedules in specialty_assignments.items():
                    # Find matching station
                    station = clinical_station_lookup.get(specialty_key)
                    if not station:
                        continue
                    
                    station_id = station.get('id', '')
                    
                    # Determine how many staff needed (at least 1 per 2 doctors)
                    staff_needed = max(1, len(doctor_schedules))
                    
                    # Find all available staff for this station
                    available_staff = []
                    for staff in clinical_staff:
                        staff_stations = staff.get('stations', [])
                        staff_pools = staff.get('pools', [])
                        staff_id = staff['id']
                        
                        # Staff MUST have this station
                        if station_id not in staff_stations:
                            continue
                        
                        # Skip if already assigned today
                        if generated.get(staff_id) and date_str in generated[staff_id]:
                            continue

                        # Hard rule: do not assign same clinical station more than 2 times in the whole range
                        station_counts = staff_station_counts.get(staff_id, {})
                        if station_counts.get(station_id, 0) >= 2:
                            continue
                        
                        # Smart day-off logic: give rest to overworked staff from priority pool
                        consecutive_days = staff_work_days.get(staff_id, 0)
                        
                        # Skip if this staff is in day-off pool and worked 5+ days straight
                        if day_off_pool in staff_pools and consecutive_days >= 5:
                            stats['days_off_distributed'] += 1
                            continue
                        
                        # Calculate priority score (lower = higher priority)
                        priority = consecutive_days * 10
                        
                        # Bonus priority for staff who haven't worked recently
                        if consecutive_days == 0:
                            priority -= 50
                        
                        available_staff.append({
                            'staff': staff,
                            'priority': priority,
                            'consecutive_days': consecutive_days
                        })
                    
                    if not available_staff:
                        continue
                    
                    # Sort by priority (lowest first = least worked recently)
                    available_staff.sort(key=lambda x: x['priority'])
                    
                    # Assign staff (up to staff_needed)
                    assigned_count = 0
                    for staff_info in available_staff[:staff_needed]:
                        staff = staff_info['staff']
                        staff_id = staff['id']
                        
                        if not generated.get(staff_id):
                            generated[staff_id] = {}
                        
                        # Get first doctor's timing for this specialty
                        first_schedule = doctor_schedules[0]
                        opd_start = first_schedule.get('start_time', '08:00')
                        opd_end = first_schedule.get('end_time', '16:00')
                        
                        # Find matching clinical shift
                        specialty_mapping = doctor_clinical_mapping.get(specialty_key, {})
                        matching_shift = specialty_mapping.get(opd_start) or (clinical_shifts[0] if clinical_shifts else None)
                        
                        # Calculate shift times
                        if matching_shift:
                            shift_start = matching_shift.get('start', '')
                            shift_end = matching_shift.get('end', '')
                        else:
                            # Default: 30 min before doctor
                            try:
                                doc_h, doc_m = opd_start.split(':')
                                staff_start_min = int(doc_h) * 60 + int(doc_m) - 30
                                shift_start = f"{staff_start_min // 60:02d}:{staff_start_min % 60:02d}"
                                shift_end = opd_end
                            except:
                                shift_start = '08:00'
                                shift_end = '16:00'
                        
                        generated[staff_id][date_str] = {
                            'station': station_id,
                            'leave': '',
                            'shift_start': shift_start,
                            'shift_end': shift_end,
                            'notes': f'Auto: {station.get("name", "")} - {len(doctor_schedules)} doctor(s)'
                        }
                        
                        # Track work days and station usage
                        staff_work_days[staff_id] = staff_work_days.get(staff_id, 0) + 1
                        station_counts = staff_station_counts.setdefault(staff_id, {})
                        station_counts[station_id] = station_counts.get(station_id, 0) + 1
                        assigned_count += 1
                        stats['clinical_assigned'] += 1
            
            # Front staff: ONLY assign when doctors are present
            if front_staff and has_doctors:
                front_stations = station_config.get('front', [])
                num_front_to_assign = min(len(front_staff), max(1, len(front_stations)))
                
                # Get front desk shift templates
                front_team_templates = shift_templates.get('front', {})
                front_shifts = []
                for category_name, category_shifts in front_team_templates.items():
                    if isinstance(category_shifts, list):
                        front_shifts.extend(category_shifts)
                
                # Day-off pool rotation
                days_since_start = (current_date - start).days
                day_off_pool = ['A', 'B', 'C', 'D'][(days_since_start // 2) % 4]
                
                # Build priority list of available staff
                available_front = []
                for staff in front_staff:
                    staff_id = staff['id']
                    staff_pools = staff.get('pools', [])
                    
                    # Skip if already assigned
                    if generated.get(staff_id) and date_str in generated[staff_id]:
                        continue

                    # Hard rule: do not assign same front station more than 2 times in the whole range
                    front_counts = staff_station_counts.get(staff_id, {})
                    # We cannot yet know which station index they will take, but if they already
                    # hit 2+ for *all* front stations we skip them to force rotation.
                    front_ids = [fs.get('id') for fs in station_config.get('front', [])]
                    if front_ids and all(front_counts.get(fid, 0) >= 2 for fid in front_ids if fid):
                        continue
                    
                    consecutive_days = staff_work_days.get(staff_id, 0)
                    
                    # Give day off to priority pool after 5+ days
                    if day_off_pool in staff_pools and consecutive_days >= 5:
                        stats['days_off_distributed'] += 1
                        continue
                    
                    # Priority: least worked gets assigned first
                    priority = consecutive_days * 10
                    if consecutive_days == 0:
                        priority -= 50
                    
                    available_front.append({
                        'staff': staff,
                        'priority': priority
                    })
                
                # Sort by priority
                available_front.sort(key=lambda x: x['priority'])
                
                # Assign up to num_front_to_assign staff
                for i in range(min(num_front_to_assign, len(available_front))):
                    staff = available_front[i]['staff']
                    staff_id = staff['id']
                    
                    if not generated.get(staff_id):
                        generated[staff_id] = {}
                    
                    # Assign to specific front station (try to pick the one used least by this staff)
                    station = None
                    if front_stations:
                        front_counts = staff_station_counts.setdefault(staff_id, {})
                        station = min(
                            front_stations,
                            key=lambda s: front_counts.get(s.get('id'), 0)
                        )
                    station_id = station.get('id', '') if station else ''
                    
                    # Get shift timing
                    default_shift = front_shifts[0] if front_shifts else None
                    shift_start = default_shift.get('start', '07:30') if default_shift else '07:30'
                    shift_end = default_shift.get('end', '15:30') if default_shift else '15:30'
                    
                    generated[staff_id][date_str] = {
                        'station': station_id,
                        'leave': '',
                        'shift_start': shift_start,
                        'shift_end': shift_end,
                        'notes': f'Auto: {station.get("name", "Front desk")}' if station else 'Auto-assigned'
                    }
                    
                    # Track work days and station usage
                    staff_work_days[staff_id] = staff_work_days.get(staff_id, 0) + 1
                    front_counts = staff_station_counts.setdefault(staff_id, {})
                    if station_id:
                        front_counts[station_id] = front_counts.get(station_id, 0) + 1
                    stats['front_assigned'] += 1
            
            # Reset work day counter for staff not assigned today (day off)
            for staff in staff_records:
                if staff.get('active', True):
                    staff_id = staff['id']
                    if not (generated.get(staff_id) and date_str in generated[staff_id]):
                        staff_work_days[staff_id] = 0  # Reset counter on day off
            
            current_date += timedelta(days=1)
        
        # Save generated roster
        roster_path = 'data/pr_roster.json'
        with open(roster_path, 'w', encoding='utf-8') as f:
            json.dump(generated, f, indent=2, ensure_ascii=False)
        
        return jsonify({
            'ok': True,
            'roster': generated,
            'message': f'Roster generated successfully for {stats["total_days"]} days',
            'stats': stats,
            'info': {
                'date_range': f'{start_date} to {end_date}',
                'clinical_staff_count': len(clinical_staff),
                'front_staff_count': len(front_staff),
                'has_doctor_schedules': stats['days_with_doctors'] > 0,
                'days_off_given': stats['days_off_distributed']
            }
        })
    except Exception as e:
        print(f'Error generating roster: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/pr/export-excel')
def api_pr_export_excel():
    """Export roster to Excel in OneDrive format"""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    
    try:
        data = request.get_json()
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        
        if not start_date or not end_date:
            return jsonify({'ok': False, 'error': 'start_date and end_date required'}), 400
        
        # Import excel export module
        from excel_export import export_to_excel
        
        # Generate filename
        filename = f"PR_Roster_{start_date}_to_{end_date}.xlsx"
        filepath = os.path.join('uploads', 'schedules', filename)
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        # Export to Excel
        export_to_excel(start_date, end_date, filepath)
        
        return jsonify({
            'ok': True,
            'message': 'Excel file generated successfully',
            'filename': filename,
            'download_url': f'/uploads/schedules/{filename}'
        })
    except Exception as e:
        print(f'Error exporting to Excel: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/pr/parse-off-requests')
def api_pr_parse_off_requests():
    """Parse off request mail format and extract staff leave requests"""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    
    try:
        data = request.get_json()
        mail_text = data.get('mail_text', '')
        
        if not mail_text:
            return jsonify({'ok': False, 'error': 'mail_text required'}), 400
        
        # Parse format like:
        # STAFF NAME - DATE RANGE - LEAVE TYPE
        # Example: "FAALIH - 20-25 Dec - AL"
        
        parsed_requests = []
        lines = mail_text.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            
            parts = [p.strip() for p in line.split('-')]
            if len(parts) >= 3:
                staff_name = parts[0].strip()
                date_range = parts[1].strip()
                leave_type = parts[2].strip().upper()
                
                parsed_requests.append({
                    'staff_name': staff_name,
                    'date_range': date_range,
                    'leave_type': leave_type,
                    'status': 'pending'
                })
        
        return jsonify({
            'ok': True,
            'parsed_requests': parsed_requests,
            'count': len(parsed_requests)
        })
    except Exception as e:
        print(f'Error parsing off requests: {e}')
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/pr-bot/chat')
def api_pr_bot_chat():
    """PR Bot Assistant Chat Endpoint"""
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    try:
        data = request.get_json()
        message = data.get('message', '').lower()
        responses = {
            'roster': '📋 The roster system manages staff assignments across clinical and front desk duties. You can generate rosters automatically using AI, or manually assign staff to specific stations and dates.',
            'generation': '🤖 AI roster generation uses smart algorithms to distribute staff fairly. Choose:\n• Full Generation - all staff\n• Special Generation - custom patterns\n• Selective Generation - specific staff only',
            'staff': '👥 Staff members can have multiple roles: Clinical, Front Desk, or Training. Manage in Staff Management section.',
            'gopd': '⚙️ GOPD (General OPD) is configured for public holidays like Fridays. Set requirements in GOPD Configuration.',
            'station': '🧠 Stations represent duty locations. Clinical stations link to doctor specialties. Manage in Station Knowledge Base.',
            'leave': '🏖️ Leave management tracks AL, SL, ML, and CL leaves. View in Leave Management section.',
            'friday': '🕌 Fridays are closed days (public holiday in Maldives). Add more via Doctor Portal → Closed Days.',
            'pattern': '🔄 Special patterns:\n• Mixed Weekly - alternate weekly\n• Mixed Daily - switch daily\n• Balanced - equal distribution',
            'help': '❓ I can help with roster generation, staff management, stations, GOPD, leaves, and system features. Just ask!'
        }
        response = None
        for keyword, reply in responses.items():
            if keyword in message:
                response = reply
                break
        if not response:
            response = f"🤔 I'm not sure. Try asking about: roster generation, staff, stations, GOPD, leave, Friday closures, or patterns. Type 'help' for more!"
        return jsonify({'ok': True, 'response': response})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.post('/api/pr/roster/update')
def api_pr_roster_update():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    data = request.get_json()
    staff_id = data.get('staff_id')
    date = data.get('date')
    assignment = data.get('assignment')
    roster_path = 'data/pr_roster.json'
    if os.path.exists(roster_path):
        with open(roster_path, 'r', encoding='utf-8') as f:
            roster = json.load(f)
    else:
        roster = {}
    if staff_id not in roster:
        roster[staff_id] = {}
    roster[staff_id][date] = assignment
    os.makedirs(os.path.dirname(roster_path), exist_ok=True)
    with open(roster_path, 'w', encoding='utf-8') as f:
        json.dump(roster, f, indent=2, ensure_ascii=False)
    return jsonify({'ok': True})

@app.get('/api/pr/roster/export')
def api_pr_roster_export():
    start_date = request.args.get('start')
    end_date = request.args.get('end')
    return jsonify({'ok': True, 'message': 'Export coming soon'})

# Leave Management
@app.get('/api/pr/leaves/all')
def api_pr_leaves_all():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    leaves_path = 'data/pr_leaves.json'
    if os.path.exists(leaves_path):
        with open(leaves_path, 'r', encoding='utf-8') as f:
            leaves = json.load(f)
    else:
        leaves = {}
    return jsonify(leaves)

@app.post('/api/pr/leaves/update')
def api_pr_leaves_update():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    leaves = request.get_json()
    leaves_path = 'data/pr_leaves.json'
    os.makedirs(os.path.dirname(leaves_path), exist_ok=True)
    with open(leaves_path, 'w', encoding='utf-8') as f:
        json.dump(leaves, f, indent=2, ensure_ascii=False)
    return jsonify({'ok': True})

@app.get('/api/pr/leaves/export')
def api_pr_leaves_export():
    return jsonify({'ok': True, 'message': 'Export coming soon'})

# Telegram Integration
@app.post('/api/telegram/config')
def api_telegram_config():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    config = request.get_json()
    bot_token = config.get('bot_token')
    group_id = config.get('group_id')
    settings_path = 'data/settings.json'
    if os.path.exists(settings_path):
        with open(settings_path, 'r', encoding='utf-8') as f:
            settings = json.load(f)
    else:
        settings = {}
    settings['telegram'] = {'bot_token': bot_token, 'group_id': group_id, 'enabled': True}
    with open(settings_path, 'w', encoding='utf-8') as f:
        json.dump(settings, f, indent=2)
    return jsonify({'ok': True})

@app.get('/api/telegram/test')
def api_telegram_test():
    settings_path = 'data/settings.json'
    if os.path.exists(settings_path):
        with open(settings_path, 'r', encoding='utf-8') as f:
            settings = json.load(f)
            if 'telegram' in settings and settings['telegram'].get('enabled'):
                return jsonify({'ok': True, 'message': 'Configured but not tested'})
    return jsonify({'ok': False, 'message': 'Not configured'})

@app.post('/api/telegram/sync')
def api_telegram_sync():
    return jsonify({'ok': True, 'last_sync': 'Never'})

# OneDrive Integration
@app.post('/api/onedrive/config')
def api_onedrive_config():
    if not session.get('user'):
        return jsonify({'ok': False, 'error': 'auth required'}), 401
    config = request.get_json()
    token = config.get('token')
    sheet_id = config.get('sheet_id')
    settings_path = 'data/settings.json'
    if os.path.exists(settings_path):
        with open(settings_path, 'r', encoding='utf-8') as f:
            settings = json.load(f)
    else:
        settings = {}
    settings['onedrive'] = {'token': token, 'sheet_id': sheet_id}
    with open(settings_path, 'w', encoding='utf-8') as f:
        json.dump(settings, f, indent=2)
    return jsonify({'ok': True})

@app.get('/api/onedrive/test')
def api_onedrive_test():
    return jsonify({'ok': False, 'message': 'Not configured'})

if __name__ == '__main__':
    # Allow host/port override via env
    host = os.environ.get('FLASK_RUN_HOST','0.0.0.0')
    port = int(os.environ.get('FLASK_RUN_PORT','5000'))
    debug = os.environ.get('FLASK_DEBUG','1') == '1'
    maybe_start_telegram()
    app.run(host=host, port=port, debug=debug, threaded=True)